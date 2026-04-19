#!/usr/bin/env python3
from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import re

from lxml import etree as ET
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import column_index_from_string

import build_allo_mapping_workbook as allo_map


ROOT = Path("/Users/Kezhunya/Documents/New project/АЛЛО")
OUT_WORKBOOK = ROOT / "update_allo_color_fallbacks_review.xlsx"
OUTPUT_XML = ROOT / "update_allo.xml"
TMP_SOURCE_XML = Path("/tmp/allo_feed/aquafavorit_source.xml")
BACKUP_SOURCE_XML = Path("/Volumes/X-Files/Загрузки рабочие/Остатки/Backup/aquafavorit_last.xml")
MAPPING_WORKBOOK = ROOT / "ALLO_сопоставление.xlsx"
TEMP_SECTIONS_WORKBOOK = ROOT / "Временное_сопоставление_разделов_ALLO.xlsx"
ALLO_TEMPLATE_DIR = ROOT / "Шаблоны фильтров с АЛЛО"

COLOR_FIELDS = ("Цвет", "Цвет стекла", "Цвет профиля")
COLOR_FIELD_KEYS = {allo_map.normalize_key(x) for x in COLOR_FIELDS}
CYR_RE = re.compile(r"[А-Яа-яЁёІіЇїЄєҐґ]")
SOURCE_COLOR_ALIASES = {
    "Цвет": ("Цвет", "Колір"),
    "Цвет стекла": ("Цвет стекла", "Тон стекла", "Тон скла", "Колір скла"),
    "Цвет профиля": ("Цвет профиля", "Колір профілю"),
}
KNOWN_GOOD_NORMALIZATIONS = {
    (allo_map.normalize_key("тонированный"), allo_map.normalize_key("тонированное")),
    (allo_map.normalize_key("тонированное"), allo_map.normalize_key("тонированное")),
    (allo_map.normalize_key("прозрачный"), allo_map.normalize_key("прозрачное")),
    (allo_map.normalize_key("прозрачное"), allo_map.normalize_key("прозрачное")),
    (allo_map.normalize_key("прозрачный"), allo_map.normalize_key("прозрачный")),
}


def clean(value) -> str:
    return " ".join(str(value or "").strip().split())


def normalize(value) -> str:
    return allo_map.normalize_key(value)


def has_cyrillic(value: str) -> bool:
    return bool(CYR_RE.search(value or ""))


def color_sort_key(value: str) -> tuple[int, str]:
    # Для UX в выпадающем списке сначала показываем кириллические каноны.
    return (0 if has_cyrillic(value) else 1, normalize(value))


def parse_offer_params(node: ET._Element) -> dict[str, str]:
    result: dict[str, str] = {}
    for p in node.findall("param"):
        name = clean(p.get("name"))
        value = clean(p.text)
        if not name or not value:
            continue
        if name not in result:
            result[name] = value
    return result


def get_source_xml_path() -> Path:
    if TMP_SOURCE_XML.is_file():
        return TMP_SOURCE_XML
    if BACKUP_SOURCE_XML.is_file():
        return BACKUP_SOURCE_XML
    raise FileNotFoundError("Не найден исходный AquaFavorit XML ни в /tmp, ни в backup.")


def load_source_id_name_fallback() -> dict[str, str]:
    out: dict[str, str] = {}

    if TEMP_SECTIONS_WORKBOOK.is_file():
        wb = load_workbook(TEMP_SECTIONS_WORKBOOK, read_only=True, data_only=True)
        if "Новые_разделы_сайта" in wb.sheetnames:
            ws = wb["Новые_разделы_сайта"]
            headers = [clean(x) for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            idx = {h: i for i, h in enumerate(headers)}
            for row in ws.iter_rows(min_row=2, values_only=True):
                sid = clean(row[idx.get("ID категории сайта", -1)])
                sname = clean(row[idx.get("Новая категория сайта", -1)])
                if sid and sname and sid not in out:
                    out[sid] = sname

    if MAPPING_WORKBOOK.is_file():
        wb = load_workbook(MAPPING_WORKBOOK, read_only=True, data_only=True)
        if "Tech" in wb.sheetnames:
            ws = wb["Tech"]
            headers = [clean(x) for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            idx = {h: i for i, h in enumerate(headers)}
            for row in ws.iter_rows(min_row=2, values_only=True):
                sid = clean(row[idx.get("ID категории", -1)])
                sname = clean(row[idx.get("Категория в исходном фиде", -1)])
                if sid and sname and sid not in out:
                    out[sid] = sname
    return out


def load_source_data() -> tuple[dict[str, dict], dict[str, str]]:
    source_path = get_source_xml_path()
    root = ET.parse(str(source_path)).getroot()
    id_name_fallback = load_source_id_name_fallback()

    category_names: dict[str, str] = {}
    for c in root.findall(".//category"):
        cid = clean(c.findtext("id"))
        name = clean(c.findtext("name"))
        if cid and name:
            category_names[cid] = name

    by_code: dict[str, dict] = {}
    by_offer_id: dict[str, dict] = {}

    for offer in root.findall(".//offer"):
        params = parse_offer_params(offer)
        code = clean(params.get("Артикул") or offer.findtext("vendorCode"))
        offer_id = clean(offer.get("id") or offer.findtext("id"))
        source_cat_id = clean(offer.findtext("categoryId"))
        source_cat_name = category_names.get(source_cat_id) or id_name_fallback.get(source_cat_id) or source_cat_id
        payload = {
            "params": params,
            "source_category_name": source_cat_name,
        }
        if code and normalize(code) not in by_code:
            by_code[normalize(code)] = payload
        if offer_id and normalize(offer_id) not in by_offer_id:
            by_offer_id[normalize(offer_id)] = payload

    return by_code, by_offer_id


def load_output_items() -> tuple[list[ET._Element], dict[str, str]]:
    root = ET.parse(str(OUTPUT_XML)).getroot()
    cat_names: dict[str, str] = {}
    for c in root.findall(".//categories/category"):
        cid = clean(c.findtext("id"))
        name = clean(c.findtext("name"))
        if cid and name:
            cat_names[cid] = name
    items = root.findall(".//items/item")
    return items, cat_names


def load_exact_color_mappings() -> dict[tuple[str, str, str], str]:
    wb = load_workbook(MAPPING_WORKBOOK, read_only=True, data_only=True)
    ws = wb["Значения"]
    headers = [clean(x) for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(headers)}

    out: dict[tuple[str, str, str], str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        target_cat = clean(row[idx.get("Категория ALLO", -1)])
        param_allo = clean(row[idx.get("Параметр ALLO", -1)])
        source_val = clean(row[idx.get("Значение исходника", -1)])
        confirmed_val = clean(row[idx.get("Подтвердить значение ALLO", -1)])
        status = clean(row[idx.get("Статус", -1)])
        if not target_cat or not param_allo or not source_val or not confirmed_val:
            continue
        if normalize(param_allo) not in {normalize(x) for x in COLOR_FIELDS}:
            continue
        if "точно сопоставлено" not in normalize(status):
            continue
        out[(normalize(target_cat), normalize(param_allo), normalize(source_val))] = confirmed_val
    return out


def _load_color_canons_from_catalog() -> dict[tuple[str, str], list[str]]:
    allo_categories, _, _ = allo_map.load_allo_catalog()
    out: dict[tuple[str, str], set[str]] = defaultdict(set)
    for cat in allo_categories.values():
        title = clean(cat.get("title"))
        if not title:
            continue
        filters = cat.get("filters", {}) or {}
        for color_field in COLOR_FIELDS:
            values = (filters.get(color_field, {}) or {}).get("values", {}) or {}
            for v in values.keys():
                vv = clean(v)
                if vv:
                    out[(normalize(title), normalize(color_field))].add(vv)
    return {k: sorted(v, key=color_sort_key) for k, v in out.items()}


def _load_color_canons_from_templates() -> dict[tuple[str, str], list[str]]:
    result: dict[tuple[str, str], list[str]] = {}
    if not ALLO_TEMPLATE_DIR.exists():
        return result

    for path in sorted(ALLO_TEMPLATE_DIR.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        try:
            wb = load_workbook(path, data_only=False)
        except Exception:
            continue
        if "Шаблон импорта" not in wb.sheetnames:
            continue

        ws = wb["Шаблон импорта"]
        category_title = clean(ws["D3"].value)
        if not category_title:
            continue

        example_ws = wb["Пример заполнения"] if "Пример заполнения" in wb.sheetnames else None
        example_col_by_code: dict[str, int] = {}
        example_col_by_header: dict[str, int] = {}
        if example_ws is not None:
            for col in range(1, example_ws.max_column + 1):
                example_header = clean(example_ws.cell(1, col).value)
                example_code = clean(example_ws.cell(3, col).value)
                if example_header and example_header not in example_col_by_header:
                    example_col_by_header[example_header] = col
                if example_code and example_code not in example_col_by_code:
                    example_col_by_code[example_code] = col

        dv_by_col: dict[int, object] = {}
        for dv in ws.data_validations.dataValidation:
            sqrefs = str(dv.sqref).split()
            for sqref in sqrefs:
                if ":" in sqref or not sqref.endswith("3"):
                    continue
                col_letters = "".join(ch for ch in sqref if ch.isalpha())
                if not col_letters:
                    continue
                dv_by_col[column_index_from_string(col_letters)] = dv

        grouped: dict[str, list[dict]] = defaultdict(list)
        for col in range(1, ws.max_column + 1):
            raw_header = clean(ws.cell(1, col).value)
            field_type = clean(ws.cell(2, col).value)
            code = clean(ws.cell(3, col).value)
            if not raw_header:
                continue
            display_name = clean(raw_header.replace("*", ""))
            canonical_name = allo_map.canonical_filter_name(display_name)
            if normalize(canonical_name) not in COLOR_FIELD_KEYS:
                continue

            dv = dv_by_col.get(col)
            values = allo_map.parse_template_validation_values(wb, getattr(dv, "formula1", None))
            if not values and example_ws is not None:
                example_col = example_col_by_code.get(code) or example_col_by_header.get(display_name)
                if example_col:
                    example_type = clean(example_ws.cell(2, example_col).value) or field_type
                    values = allo_map.parse_template_example_values(example_ws, example_col, example_type)

            cleaned_values: list[str] = []
            seen: set[str] = set()
            for value in values:
                text = clean(allo_map.canonical_value_name(value))
                if not text:
                    continue
                key = normalize(text)
                if key in seen:
                    continue
                seen.add(key)
                cleaned_values.append(text)

            grouped[canonical_name].append(
                {
                    "required": "*" in raw_header,
                    "values": cleaned_values,
                }
            )

        for color_name, options in grouped.items():
            if not options:
                continue

            def option_score(option: dict) -> tuple[int, int, int]:
                values = option["values"]
                cyr_count = sum(1 for v in values if has_cyrillic(v))
                return (cyr_count, 1 if option["required"] else 0, len(values))

            best = max(options, key=option_score)
            best_values = sorted(best["values"], key=color_sort_key)
            if best_values:
                result[(normalize(category_title), normalize(color_name))] = best_values

    return result


def load_color_canons() -> dict[tuple[str, str], list[str]]:
    # Основа — каноны из шаблонов ALLO (по запросу пользователя).
    template_canons = _load_color_canons_from_templates()
    # Фолбэк — каталог (если в шаблоне нет цветового списка).
    catalog_canons = _load_color_canons_from_catalog()

    merged = dict(template_canons)
    for key, vals in catalog_canons.items():
        if key not in merged or not merged[key]:
            merged[key] = vals
    return merged


def pick_source_color(params: dict[str, str], output_param: str) -> str:
    aliases = SOURCE_COLOR_ALIASES.get(output_param, (output_param,))
    for a in aliases:
        v = clean(params.get(a, ""))
        if v:
            return v
    return ""


def collect_rows() -> list[dict[str, str]]:
    by_code, by_offer_id = load_source_data()
    items, out_cat_names = load_output_items()
    exact_map = load_exact_color_mappings()
    color_canons = load_color_canons()

    groups: dict[tuple[str, str, str, str], dict] = {}
    for item in items:
        item_id = clean(item.findtext("id"))
        item_code = clean(item.findtext("code"))
        target_cat = out_cat_names.get(clean(item.findtext("categoryId")), clean(item.findtext("categoryId")))
        if not target_cat:
            continue

        source_payload = None
        if item_code:
            source_payload = by_code.get(normalize(item_code))
        if source_payload is None and item_id:
            source_payload = by_offer_id.get(normalize(item_id))
        if source_payload is None:
            continue

        source_params = source_payload["params"]
        source_section = source_payload["source_category_name"]
        item_name_ru = clean(item.findtext("name_ru"))

        out_params = parse_offer_params(item)
        for output_param in COLOR_FIELDS:
            output_value = clean(out_params.get(output_param, ""))
            if not output_value:
                continue
            source_value = pick_source_color(source_params, output_param)
            if not source_value:
                continue

            s_key = normalize(source_value)
            o_key = normalize(output_value)
            if s_key == o_key:
                continue
            if (s_key, o_key) in KNOWN_GOOD_NORMALIZATIONS:
                continue

            exact = exact_map.get((normalize(target_cat), normalize(output_param), s_key), "")
            if exact and normalize(exact) == o_key:
                continue

            gkey = (target_cat, output_param, source_value, output_value)
            if gkey not in groups:
                groups[gkey] = {
                    "Раздел_ALLO": target_cat,
                    "Параметр_ALLO": output_param,
                    "Было_в_исходнике": source_value,
                    "Автоподстановка_сейчас": output_value,
                    "Подтвердить_канон_ALLO": output_value,
                    "Сколько_канонов_в_ALLO": str(len(color_canons.get((normalize(target_cat), normalize(output_param)), []))),
                    "Разделы_исходника_set": set(),
                    "Примеры_артикулов_list": [],
                }

            g = groups[gkey]
            if source_section:
                g["Разделы_исходника_set"].add(source_section)
            if item_code and item_code not in g["Примеры_артикулов_list"] and len(g["Примеры_артикулов_list"]) < 10:
                g["Примеры_артикулов_list"].append(item_code)
            if item_id and not item_code and item_id not in g["Примеры_артикулов_list"] and len(g["Примеры_артикулов_list"]) < 10:
                g["Примеры_артикулов_list"].append(item_id)
            if not g["Примеры_артикулов_list"] and item_name_ru:
                g["Примеры_артикулов_list"].append(item_name_ru[:50])

    rows: list[dict[str, str]] = []
    for (_cat, _param, _src, _out), g in sorted(groups.items(), key=lambda kv: (normalize(kv[0][0]), normalize(kv[0][1]), normalize(kv[0][2]), normalize(kv[0][3]))):
        rows.append(
            {
                "Раздел_ALLO": g["Раздел_ALLO"],
                "Параметр_ALLO": g["Параметр_ALLO"],
                "Было_в_исходнике": g["Было_в_исходнике"],
                "Автоподстановка_сейчас": g["Автоподстановка_сейчас"],
                "Подтвердить_канон_ALLO": g["Подтвердить_канон_ALLO"],
                "Сколько_канонов_в_ALLO": g["Сколько_канонов_в_ALLO"],
                "Разделы_исходника": ", ".join(sorted(g["Разделы_исходника_set"], key=normalize)),
                "Примеры_артикулов": ", ".join(g["Примеры_артикулов_list"]),
            }
        )
    return rows


def write_workbook(rows: list[dict[str, str]]) -> None:
    color_canons = load_color_canons()

    wb = Workbook()
    ws = wb.active
    ws.title = "Цвета_на_проверку"
    headers = [
        "Раздел_ALLO",
        "Параметр_ALLO",
        "Было_в_исходнике",
        "Автоподстановка_сейчас",
        "Подтвердить_канон_ALLO",
        "Сколько_канонов_в_ALLO",
        "Разделы_исходника",
        "Примеры_артикулов",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    green = PatternFill(fill_type="solid", start_color="E2F0D9", end_color="E2F0D9")
    yellow = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).fill = green
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=5).fill = yellow

    lists = wb.create_sheet("Lists")
    col = 1
    validation_ranges: dict[tuple[str, str], str] = {}
    for key, vals in sorted(color_canons.items(), key=lambda x: (x[0][0], x[0][1])):
        if not vals:
            continue
        cat_key, param_key = key
        lists.cell(row=1, column=col, value=f"{cat_key}::{param_key}")
        for i, v in enumerate(vals, start=2):
            lists.cell(row=i, column=col, value=v)
        end_row = len(vals) + 1
        letter = lists.cell(row=1, column=col).column_letter
        validation_ranges[key] = f"'Lists'!${letter}$2:${letter}${end_row}"
        col += 1
    lists.sheet_state = "hidden"

    for r in range(2, ws.max_row + 1):
        cat = clean(ws.cell(row=r, column=1).value)
        param = clean(ws.cell(row=r, column=2).value)
        rng = validation_ranges.get((normalize(cat), normalize(param)))
        if not rng:
            continue
        dv = DataValidation(type="list", formula1=f"={rng}", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=r, column=5))

    widths = {1: 30, 2: 18, 3: 24, 4: 24, 5: 24, 6: 18, 7: 42, 8: 30}
    for cidx, w in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=cidx).column_letter].width = w

    wb.save(OUT_WORKBOOK)


def main() -> int:
    rows = collect_rows()
    write_workbook(rows)
    print(f"✅ Создан файл: {OUT_WORKBOOK}")
    print(f"📌 Строк на проверку: {len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
