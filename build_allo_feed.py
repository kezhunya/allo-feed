#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import os
import re
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path

from lxml import etree as ET
from openpyxl import Workbook, load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
if (SCRIPT_DIR / "ALLO_сопоставление.xlsx").exists():
    ALLO_DIR = SCRIPT_DIR
else:
    ALLO_DIR = SCRIPT_DIR / "АЛЛО"

if (ALLO_DIR / "update_maudau.py").exists():
    ROOT = ALLO_DIR
elif (ALLO_DIR.parent / "update_maudau.py").exists():
    ROOT = ALLO_DIR.parent
else:
    ROOT = ALLO_DIR.parent
TMP_DIR = Path("/tmp/allo_feed")
TMP_DIR.mkdir(parents=True, exist_ok=True)
OSTATKI_DIR = Path(os.environ.get("OSTATKI_DIR", str(ROOT / "Остатки")))
SHARED_BACKUP_DIR = Path(os.environ.get("SHARED_BACKUP_DIR", str(OSTATKI_DIR / "Backup")))
SHARED_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
LEGACY_BACKUP_DIR = ALLO_DIR / "backups"
BACKUP_DIR = SHARED_BACKUP_DIR
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
LEGACY_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
SOURCES_STATE_JSON = BACKUP_DIR / "sources_state_allo.json"

OUTPUT_XML = ALLO_DIR / "update_allo.xml"
REPORT_XLSX = ALLO_DIR / "update_allo_report.xlsx"
COLOR_FIX_REPORT_XLSX = ALLO_DIR / "update_allo_color_fallbacks.xlsx"
BASE_XML = TMP_DIR / "aquafavorit_source.xml"
ROZETKA_XML = TMP_DIR / "rozetka_source.xml"
BASE_BACKUP_XML = BACKUP_DIR / "aquafavorit_last.xml"
ROZETKA_BACKUP_XML = BACKUP_DIR / "parserbiz_last.xml"
LEGACY_BASE_BACKUP_XML = LEGACY_BACKUP_DIR / "aquafavorit_last.xml"
LEGACY_ROZETKA_BACKUP_XML = LEGACY_BACKUP_DIR / "rozetka_last.xml"
LEGACY_ROZETKA_BACKUP_XML_ALT = LEGACY_BACKUP_DIR / "parserbiz_last.xml"

sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ALLO_DIR))
sys.path.insert(0, str(ROOT / "MAUDAU"))

import update_maudau as maudau  # noqa: E402
import build_allo_mapping_workbook as allo_map  # noqa: E402


FIRM_NAME = "AquaFavorit"
FIRM_ID = "0377"
DEFAULT_RATE = "1"

TECH_CATEGORY_ID_OFFSET = 100000
MAX_PICTURES = 12
FILTER_NAMES_CACHE: dict[tuple[str, str], list[str]] = {}
FILTER_META_CACHE: dict[tuple[str, str, str], dict] = {}
PARAM_NAME_PICK_CACHE: dict[tuple[str, str, str, str, str], tuple[str, bool]] = {}
PARAM_VALUE_PICK_CACHE: dict[tuple[str, str, str, str, str, str], str] = {}
SECTION_FIXED_RULES_CACHE: list[allo_map.SectionFixedRule] | None = None
HEATER_POWER_AREA_PROFILES: list[tuple[float, float]] = []
PUMP_POWER_FLOW_PROFILES: list[tuple[float, float]] = []
STAGE1_CATEGORY_OVERRIDES: dict[str, str] = {
    "1279": "Сифоны и трапы",
    "1253": "Водопроводные шланги",
}
SOURCE_CATEGORY_URL_OVERRIDES: dict[str, tuple[str, str]] = {
    "1253": ("Водопроводные шланги", "https://allo.ua/ru/vodoprovodnye-shlangi/"),
    "1279": ("Сифоны и трапы", "https://allo.ua/ru/sifony-i-trapy/"),
}
FORCE_SKIP_SOURCE_CATEGORY_PATTERNS: tuple[str, ...] = (
    "аксессуары все для принятия душа",
    "бачки скрытого монтажа",
    "боковые форсунки",
    "газовые колонки",
    "дезинфекторы",
    "котлы газовые",
    "котлы твердотопливные",
    "котлы электрические",
    "подголовники",
    "полупьедесталы",
    "пьедесталы",
    "средства герметизации",
    "термоманометр",
    "термометр",
    "фильтры механической очистки",
)
DROP_IF_REQUIRED_ISSUES_SOURCE_PATTERNS: tuple[str, ...] = (
    "держатели для ванной комнаты",
    "комплекты аксессуаров",
    "комплекты мебели",
    "корзины для белья",
    "настенные фены",
    "приборные краны",
)
COLOR_PARAM_KEYS = {
    allo_map.normalize_key("Цвет"),
    allo_map.normalize_key("Цвет стекла"),
    allo_map.normalize_key("Цвет профиля"),
}
MANUAL_COLOR_OVERRIDE_RULES: dict[tuple[str, str, str], str] = {
    ("Дозаторы (диспенсеры) для ванной комнаты", "Цвет", "Графитовый хром"): "Графитовый",
    ("Дозаторы (диспенсеры) для ванной комнаты", "Цвет", "Сафари"): "Safari",
    ("Дозаторы (диспенсеры) для ванной комнаты", "Цвет", "Теплый закат"): "Медный",
    ("Дозаторы (диспенсеры) для ванной комнаты", "Цвет", "Терра"): "Коричневый",
    ("Дозаторы (диспенсеры) для ванной комнаты", "Цвет", "Цветной"): "Colorless",
    ("Душевые гарнитуры", "Цвет", "Графитовый хром"): "Графитовый",
    ("Душевые системы", "Цвет", "Бронза"): "Бронзовый",
    ("Душевые системы", "Цвет", "Теплый закат"): "Медный",
    ("Душевые системы", "Цвет", "Шлифованный черный хром"): "Черный / Хром",
    ("Ершики и стойки", "Цвет", "Цветной"): "Colorless",
    ("Инсталляции", "Цвет", "Серый матовый"): "Серый",
    ("Керамические обогреватели", "Цвет", "Цветной"): "Colorless",
    ("Крючки для ванной", "Цвет", "Бронза"): "Бронзовый",
    ("Крючки для ванной", "Цвет", "Графитовый хром"): "Графитовый",
    ("Крючки для ванной", "Цвет", "Теплый закат"): "Медный",
    ("Крючки для ванной", "Цвет", "Цветной"): "Colorless",
    ("Крючки для ванной", "Цвет", "Шлифованный черный хром"): "Черный / Хром",
    ("Кухонные мойки", "Цвет", "Satin nickel"): "Никель",
    ("Кухонные мойки", "Цвет", "Терра"): "Коричневый",
    ("Мусорные ведра", "Цвет", "Бронза"): "Бронзовый",
    ("Мыльницы", "Цвет", "Зеленый"): "Темно-зеленый",
    ("Мыльницы", "Цвет", "Цветной"): "Colorless",
    ("Пеналы для ванной комнаты", "Цвет", "Береза"): "Жасмин",
    ("Полки для ванной комнаты", "Цвет", "Бронза"): "Бронзовый",
    ("Полки для ванной комнаты", "Цвет", "Шлифованный черный хром"): "Черный / Хром",
    ("Полотенцедержатели", "Цвет", "Цветной"): "Colorless",
    ("Полотенцесушители", "Цвет", "Цветной"): "Colorless",
    ("Поручни для ванной", "Цвет", "Бронза"): "Бронзовый",
    ("Поручни для ванной", "Цвет", "Шлифованный черный хром"): "Черный / Хром",
    ("Сифоны и трапы", "Цвет", "Sand"): "Песочный",
    ("Сифоны и трапы", "Цвет", "Бронза"): "Бронзовый",
    ("Сифоны и трапы", "Цвет", "Голубой матовый"): "Голубой",
    ("Сифоны и трапы", "Цвет", "Графитовый хром"): "Графитовый",
    ("Сифоны и трапы", "Цвет", "Зеленый матовый"): "Зеленый",
    ("Сифоны и трапы", "Цвет", "Золото"): "Золотистый",
    ("Сифоны и трапы", "Цвет", "Теплый закат"): "Медный",
    ("Сифоны и трапы", "Цвет", "Терра"): "Коричневый",
    ("Сифоны и трапы", "Цвет", "Шлифованный черный хром"): "Черный / Хром",
    ("Смесители", "Цвет", "Sand"): "Песочный",
    ("Смесители", "Цвет", "Satin"): "Сатин",
    ("Смесители", "Цвет", "Береза"): "Жасмин",
    ("Смесители", "Цвет", "Золото"): "Золотой",
    ("Смесители", "Цвет", "Сафари"): "Safari",
    ("Смесители", "Цвет", "Сталь"): "Нержавеющая сталь",
    ("Смесители", "Цвет", "Терра"): "Коричневый",
    ("Смесители", "Цвет", "Цветной"): "Colorless",
    ("Смесители", "Цвет", "Черный/золото"): "Черный с золотистым",
    ("Стаканы для ванной", "Цвет", "Цветной"): "Colorless",
    ("Умывальники", "Цвет", "Теплый закат"): "Медный",
    ("Умывальники", "Цвет", "Цветной"): "Colorless",
    ("Шаровые краны и вентили", "Цвет", "Цветной"): "Colorless",
}


def clean_text(value: str | None) -> str:
    return " ".join(str(value or "").strip().split())


def normalize_key(value: str | None) -> str:
    return allo_map.normalize_key(value)


def should_force_skip_source_category(source_name: str) -> bool:
    source_key = normalize_key(source_name)
    return any(pattern in source_key for pattern in FORCE_SKIP_SOURCE_CATEGORY_PATTERNS)


def should_drop_by_required_issues(source_name: str, issues: list[str]) -> bool:
    source_key = normalize_key(source_name)
    if not any(pattern in source_key for pattern in DROP_IF_REQUIRED_ISSUES_SOURCE_PATTERNS):
        return False
    return any(
        issue.startswith("Не заполнен обязательный параметр ALLO:")
        or issue.startswith("Не сопоставлено обязательное значение:")
        for issue in issues
    )


def stable_category_id(title: str) -> str:
    import zlib

    digest = zlib.crc32(title.encode("utf-8"))
    return str(TECH_CATEGORY_ID_OFFSET + digest % 800000000)


def parseable_xml(path: Path) -> bool:
    if not path.is_file() or path.stat().st_size <= 0:
        return False
    try:
        ET.parse(str(path))
        return True
    except Exception:
        return False


def fallback_candidates_for(backup: Path) -> list[Path]:
    candidates = [backup]
    if backup == BASE_BACKUP_XML:
        candidates.append(LEGACY_BASE_BACKUP_XML)
        candidates.extend(maudau.BASE_BACKUP_CANDIDATES)
    elif backup == ROZETKA_BACKUP_XML:
        candidates.append(LEGACY_ROZETKA_BACKUP_XML)
        candidates.append(LEGACY_ROZETKA_BACKUP_XML_ALT)
        candidates.extend(maudau.ROZETKA_BACKUP_CANDIDATES)
    return candidates


def resolve_backup_path(backup: Path) -> Path | None:
    for candidate in fallback_candidates_for(backup):
        if not str(candidate):
            continue
        if parseable_xml(candidate):
            return candidate
    return None


def copy_if_different(source: Path, target: Path) -> None:
    try:
        if source.resolve() == target.resolve():
            return
    except Exception:
        pass
    shutil.copy2(source, target)


def download_with_backup(
    url: str,
    target: Path,
    backup: Path,
    title: str,
    timeout: int,
) -> tuple[Path, bool, Path | None]:
    try:
        maudau.download_file(url, target, title, timeout=timeout)
        ET.parse(str(target))
        shutil.copy2(target, backup)
        return target, True, None
    except Exception as exc:
        fallback = resolve_backup_path(backup)
        if fallback:
            print(f"⚠ {title} недоступен, используем последний backup: {fallback} ({exc})")
            copy_if_different(fallback, backup)
            copy_if_different(fallback, target)
            return target, False, fallback
        raise


def source_categories(root: ET._Element) -> dict[str, str]:
    result: dict[str, str] = {}
    for node in root.xpath("//shop/categories/category"):
        cid = clean_text(node.get("id"))
        if cid:
            result[cid] = clean_text(node.text)
    return result


def offer_params(offer: ET._Element) -> dict[str, str]:
    result: dict[str, str] = {}
    for param in offer.findall("param"):
        name = clean_text(param.get("name"))
        value = clean_text(param.text)
        if not name or not value:
            continue
        result.setdefault(name, value)
    return result


def find_offer_param(params: dict[str, str], *names: str) -> str:
    lut = {normalize_key(k): v for k, v in params.items()}
    for name in names:
        value = lut.get(normalize_key(name), "")
        if value:
            return value
    return ""


def build_heater_power_area_profiles(root: ET._Element) -> list[tuple[float, float]]:
    profiles: list[tuple[float, float]] = []
    for offer in root.xpath("//offer"):
        params = offer_params(offer)
        probe = " ".join(
            [
                maudau.child_text(offer, "name"),
                maudau.child_text(offer, "name_ru"),
                maudau.child_text(offer, "description"),
                maudau.child_text(offer, "description_ru"),
            ]
        )
        if "обогревател" not in normalize_key(probe):
            continue
        area = extract_first_number(
            find_offer_param(
                params,
                "Площадь обогрева",
                "Площадь обслуживания, кв. м",
                "Рекомендуемая площадь помещения, кв.м.",
                "Рекомендуемая площадь помещения, кв. м",
            )
        )
        power_w = extract_first_number(find_offer_param(params, "Мощность, Вт"))
        if power_w is None:
            power_kw = extract_first_number(find_offer_param(params, "Мощность, кВт"))
            if power_kw is not None:
                power_w = power_kw * 1000.0
        if area is None or power_w is None:
            continue
        if area > 0 and power_w > 0:
            profiles.append((power_w, area))
    return profiles


def extract_power_w_from_text(params: dict[str, str], text: str) -> float | None:
    direct_w = extract_first_number(find_offer_param(params, "Мощность, Вт", "Номинальная мощность, Вт", "Потужність, Вт"))
    if direct_w is not None:
        return direct_w
    direct_kw = extract_first_number(find_offer_param(params, "Мощность, кВт", "Мощность (кВт)", "Потужність, кВт"))
    if direct_kw is not None:
        return direct_kw * 1000.0

    for key, value in params.items():
        key_norm = normalize_key(key)
        if "мощност" not in key_norm and "потужн" not in key_norm:
            continue
        joined = f"{key} {value}"
        number = extract_first_number(joined)
        if number is None:
            continue
        if any(token in key_norm for token in ("квт", "kw")):
            return number * 1000.0
        if any(token in key_norm for token in ("вт", "w")):
            return number

    text_norm = clean_text(text).replace(",", ".")
    m_pow_w = re.search(r"мощност[ьі][^0-9]{0,20}(\d{2,5})\s*(?:вт|w)\b", text_norm, flags=re.IGNORECASE)
    if m_pow_w:
        return float(m_pow_w.group(1))
    m_pow_kw = re.search(r"мощност[ьі][^0-9]{0,20}(\d+(?:\.\d+)?)\s*(?:квт|kw)\b", text_norm, flags=re.IGNORECASE)
    if m_pow_kw:
        return float(m_pow_kw.group(1)) * 1000.0
    return None


def extract_flow_lh_from_text(text: str) -> float | None:
    probe = clean_text(text).lower()
    probe = probe.replace(",", ".").replace("м³", "м3")
    candidates: list[float] = []

    def push(number_raw: str, unit_raw: str) -> None:
        try:
            number = float(number_raw)
        except ValueError:
            return
        unit = unit_raw.replace(" ", "")
        if unit in {"м3/час", "м3/ч", "м3/год"}:
            candidates.append(number * 1000.0)
        elif unit in {"л/мин", "л/хв"}:
            candidates.append(number * 60.0)
        elif unit in {"л/ч", "л/час"}:
            candidates.append(number)

    for m in re.finditer(
        r"(\d+(?:\.\d+)?)\s*(м3/час|м3/ч|м3/год|л/мин|л/хв|л/ч|л/час)\b",
        probe,
        flags=re.IGNORECASE,
    ):
        push(m.group(1), m.group(2))
    for m in re.finditer(
        r"(м3/час|м3/ч|м3/год|л/мин|л/хв|л/ч|л/час)\s*[:=]?\s*(\d+(?:\.\d+)?)",
        probe,
        flags=re.IGNORECASE,
    ):
        push(m.group(2), m.group(1))

    if not candidates:
        return None
    return max(candidates)


def extract_pump_flow_lh(params: dict[str, str], text: str) -> float | None:
    flow_lm = extract_first_number(find_offer_param(params, "Производительность, л/мин", "Продуктивність, л/хв"))
    if flow_lm is not None:
        return flow_lm * 60.0

    for key, value in params.items():
        key_norm = normalize_key(key)
        if "производител" not in key_norm and "продуктив" not in key_norm:
            continue
        parsed = extract_flow_lh_from_text(f"{key} {value}")
        if parsed is not None:
            return parsed

    return extract_flow_lh_from_text(text)


def build_pump_power_flow_profiles(root: ET._Element) -> list[tuple[float, float]]:
    profiles: list[tuple[float, float]] = []
    for offer in root.xpath("//offer"):
        params = offer_params(offer)
        probe = clean_text(
            " ".join(
                [
                    maudau.child_text(offer, "name"),
                    maudau.child_text(offer, "name_ru"),
                    maudau.child_text(offer, "description"),
                    maudau.child_text(offer, "description_ru"),
                    " ".join(f"{k} {v}" for k, v in params.items()),
                ]
            )
        )
        probe_key = normalize_key(probe)
        if "насос" not in probe_key and "pump" not in probe_key and "sololift" not in probe_key:
            continue

        power_w = extract_power_w_from_text(params, probe)
        flow_lh = extract_pump_flow_lh(params, probe)

        if power_w is None or flow_lh is None:
            continue
        if power_w > 0 and flow_lh > 0:
            profiles.append((power_w, flow_lh))
    return profiles


def load_stage1_category_map() -> tuple[dict[str, str], set[str]]:
    rules_path = allo_map.ALLO_CATEGORY_RULES_XLSX
    if not rules_path.exists():
        # Файл этапа 1 мог быть удалён при чистке. Тогда берём рабочий канон
        # из ALLO_сопоставление.xlsx -> лист Tech.
        tech_path = allo_map.CURRENT_XLSX
        if not tech_path.exists():
            raise FileNotFoundError(
                f"Не найден файл категорий этапа 1 ({rules_path}) и fallback Tech ({tech_path})"
            )
        wb = load_workbook(tech_path, read_only=True, data_only=True)
        if "Tech" not in wb.sheetnames:
            raise FileNotFoundError(
                f"Нет листа Tech в fallback-файле: {tech_path}"
            )
        ws = wb["Tech"]
        headers = [clean_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        id_idx = headers.index("ID категории")
        target_idx = headers.index("Категория ALLO")
        category_map: dict[str, str] = {}
        skip_ids: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            source_id = clean_text(row[id_idx]) if len(row) > id_idx else ""
            target_title = clean_text(row[target_idx]) if len(row) > target_idx else ""
            if not source_id:
                continue
            if not target_title:
                skip_ids.add(source_id)
                continue
            category_map[source_id] = target_title
        for source_id, target_title in STAGE1_CATEGORY_OVERRIDES.items():
            category_map[source_id] = target_title
        return category_map, skip_ids

    wb = load_workbook(rules_path, read_only=True, data_only=True)
    category_map: dict[str, str] = {}
    skip_ids: set[str] = set()

    for sheet_name in ("Новые_разделы_сайта", "Отброшено_<14"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [clean_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        id_idx = headers.index("ID категории сайта")
        confirm_idx = headers.index("Подтвердить категорию 377") if "Подтвердить категорию 377" in headers else None
        status_idx = headers.index("Статус") if "Статус" in headers else None

        for row in ws.iter_rows(min_row=2, values_only=True):
            source_id = clean_text(row[id_idx]) if len(row) > id_idx else ""
            if not source_id:
                continue
            confirmed = clean_text(row[confirm_idx]) if confirm_idx is not None and len(row) > confirm_idx else ""
            status = clean_text(row[status_idx]) if status_idx is not None and len(row) > status_idx else ""
            marker = f"{confirmed} {status}".lower()
            if sheet_name == "Отброшено_<14" or "не добавляем" in marker or "нет у алло" in marker:
                skip_ids.add(source_id)
                continue
            if confirmed:
                category_map[source_id] = confirmed

    # Явные подтверждённые переопределения.
    for source_id, target_title in STAGE1_CATEGORY_OVERRIDES.items():
        category_map[source_id] = target_title

    return category_map, skip_ids


def load_mapping_rules() -> tuple[dict, dict]:
    _category_choices, param_choices, value_choices = allo_map.load_existing_confirmations()

    params_by_source_id: dict[tuple[str, str], str] = {}
    params_by_source_name: dict[tuple[str, str], str] = {}
    params_by_target: dict[tuple[str, str], str] = {}
    for key, value in param_choices.items():
        category_key, source_param = key
        if str(category_key).startswith("target::"):
            params_by_target[(category_key.replace("target::", "", 1), normalize_key(source_param))] = value
        elif str(category_key).startswith("name::"):
            params_by_source_name[(category_key.replace("name::", "", 1), normalize_key(source_param))] = value
        else:
            params_by_source_id[(normalize_key(category_key), normalize_key(source_param))] = value

    values_by_source_id: dict[tuple[str, str, str, str], str] = {}
    values_by_source_name: dict[tuple[str, str, str, str], str] = {}
    values_by_target: dict[tuple[str, str, str, str], str] = {}
    for key, value in value_choices.items():
        category_key, source_param, source_value = key
        if str(category_key).startswith("target::"):
            values_by_target[
                (
                    category_key.replace("target::", "", 1),
                    normalize_key(source_param),
                    normalize_key(source_value),
                    "",
                )
            ] = value
        elif str(category_key).startswith("name::"):
            values_by_source_name[
                (
                    category_key.replace("name::", "", 1),
                    normalize_key(source_param),
                    normalize_key(source_value),
                    "",
                )
            ] = value
        else:
            values_by_source_id[
                (
                    normalize_key(category_key),
                    normalize_key(source_param),
                    normalize_key(source_value),
                    "",
                )
            ] = value

    manual_applied = 0
    for (target_title, source_param, source_value), confirmed_value in MANUAL_COLOR_OVERRIDE_RULES.items():
        if normalize_key(source_param) not in COLOR_PARAM_KEYS:
            continue
        values_by_target[
            (
                normalize_key(target_title),
                normalize_key(source_param),
                normalize_key(source_value),
                "",
            )
        ] = confirmed_value
        manual_applied += 1
    if manual_applied:
        print(f"🎨 Вшито ручных подтверждений цвета: {manual_applied}")

    return (
        {
            "by_target": params_by_target,
            "by_source_id": params_by_source_id,
            "by_source_name": params_by_source_name,
        },
        {
            "by_target": values_by_target,
            "by_source_id": values_by_source_id,
            "by_source_name": values_by_source_name,
        },
    )


def resolve_offer_category(
    offer: ET._Element,
    source_id: str,
    source_name: str,
    base_category_title: str,
) -> tuple[str, str]:
    if source_id in SOURCE_CATEGORY_URL_OVERRIDES:
        return SOURCE_CATEGORY_URL_OVERRIDES[source_id]

    target_title = base_category_title
    target_url = ""

    params = offer_params(offer)
    # 1255 всегда режем по типу на две категории ALLO.
    if source_id == "1255":
        type_key_1255 = normalize_key(find_offer_param(params, "Тип", "Вид"))
        probe_1255 = normalize_key(
            " ".join(
                [
                    maudau.child_text(offer, "name"),
                    maudau.child_text(offer, "name_ru"),
                ]
            )
        )
        if "коллекторный шкаф" in type_key_1255 or "шкаф" in type_key_1255 or "шкаф" in probe_1255:
            return "Коллекторные шкафы", "https://allo.ua/ru/kollektornye-shkafy/"
        return "Коллекторы", "https://allo.ua/ru/kollektory/"

    for param_name, value in params.items():
        value_rule = allo_map.resolve_value_level_category_rule(source_name, target_title, param_name, value)
        if value_rule:
            target_title = value_rule["target_title"]
            target_url = value_rule.get("target_url", "")

        split_title = allo_map.resolve_split_category_by_param_value(source_name, param_name, value)
        if split_title:
            target_title = split_title
            target_url = ""

        pump_rule = allo_map.resolve_pump_value_rule(param_name, value)
        if pump_rule.get("category"):
            target_title = pump_rule["category"]
            target_url = allo_map.PUMP_CATEGORY_URLS.get(target_title, "")

    # Бойлеры: накопительные и проточные должны расходиться в разные ALLO-разделы.
    source_probe = " ".join([source_name, base_category_title, maudau.child_text(offer, "name"), maudau.child_text(offer, "name_ru")])
    source_probe_key = normalize_key(source_probe)
    if target_title == "Водонагреватели" or "бойлер" in source_probe_key or "водонагревател" in source_probe_key:
        type_value = find_offer_param(params, "Тип водонагревателя", "Тип", "Вид")
        type_key = normalize_key(type_value)
        if "проточ" in type_key:
            return "Проточные водонагреватели", "https://allo.ua/ru/protochnye-vodonagrevateli/"
        if "накоп" in type_key:
            return "Водонагреватели", "https://allo.ua/ru/products/vodonagrevateli/"

    target_key = normalize_key(target_title)
    type_value = find_offer_param(params, "Тип", "Вид")
    type_key = normalize_key(type_value)

    # Убираем комбинированный раздел: всегда режем по Тип.
    if target_key == "коллекторы коллекторные шкафы":
        probe_key = " ".join([source_probe_key, type_key])
        if "коллекторный шкаф" in type_key or "шкаф" in type_key:
            return "Коллекторные шкафы", "https://allo.ua/ru/kollektornye-shkafy/"
        if "коллектор" in type_key or "коллекторная группа" in type_key:
            return "Коллекторы", "https://allo.ua/ru/kollektory/"
        if "шкаф" in probe_key:
            return "Коллекторные шкафы", "https://allo.ua/ru/kollektornye-shkafy/"
        if "коллектор" in probe_key or "смесительный узел" in probe_key:
            return "Коллекторы", "https://allo.ua/ru/kollektory/"

    # Насосная ветка: разрезаем комбинированный раздел по типу товара.
    if target_key == "насосы для воды реле и контроллеры для насосов":
        probe_key = " ".join(
            [
                source_probe_key,
                type_key,
                normalize_key(find_offer_param(params, "Назначение", "Призначення")),
                normalize_key(find_offer_param(params, "Управление", "Управління")),
            ]
        )
        if any(token in probe_key for token in ("реле", "контроллер", "автоматика", "прессконтроль", "контроллер давления", "реле давления")):
            return "Реле и контроллеры для насосов", "https://allo.ua/ru/rele-i-kontrollery-dlja-nasosov/"
        if any(token in probe_key for token in ("насос", "помп", "станц", "гидрофор")):
            return "Насосы для воды", "https://allo.ua/ru/nasosy/"

    return target_title, target_url


def pick_param_name(
    source_id: str,
    source_name: str,
    target_title: str,
    target_url: str,
    source_param: str,
    filter_names: list[str],
    param_rules: dict,
) -> tuple[str, bool]:
    cache_key = (
        normalize_key(source_id),
        normalize_key(source_name),
        normalize_key(target_title),
        normalize_key(target_url),
        normalize_key(source_param),
    )
    cached = PARAM_NAME_PICK_CACHE.get(cache_key)
    if cached is not None:
        return cached

    by_source = param_rules["by_source_id"].get((normalize_key(source_id), normalize_key(source_param)), "")
    if by_source:
        result = ("" if by_source == allo_map.IGNORE_FILTER_SENTINEL else by_source), True
        PARAM_NAME_PICK_CACHE[cache_key] = result
        return result

    by_source_name = param_rules["by_source_name"].get((normalize_key(source_name), normalize_key(source_param)), "")
    if by_source_name:
        result = ("" if by_source_name == allo_map.IGNORE_FILTER_SENTINEL else by_source_name), True
        PARAM_NAME_PICK_CACHE[cache_key] = result
        return result

    by_target = param_rules["by_target"].get((normalize_key(target_title), normalize_key(source_param)), "")
    if by_target:
        result = ("" if by_target == allo_map.IGNORE_FILTER_SENTINEL else by_target), True
        PARAM_NAME_PICK_CACHE[cache_key] = result
        return result

    suggested, _candidates = allo_map.pick_suggested_filter(source_param, filter_names, target_url)
    result = (suggested, False)
    PARAM_NAME_PICK_CACHE[cache_key] = result
    return result


def pick_param_value(
    source_id: str,
    source_name: str,
    target_title: str,
    source_param: str,
    filter_name: str,
    source_value: str,
    allowed_values: list[str],
    value_rules: dict,
) -> str:
    cache_key = (
        normalize_key(source_id),
        normalize_key(source_name),
        normalize_key(target_title),
        normalize_key(filter_name),
        normalize_key(source_param),
        normalize_key(source_value),
    )
    cached = PARAM_VALUE_PICK_CACHE.get(cache_key)
    if cached is not None:
        return cached

    keys = [
        (normalize_key(source_id), normalize_key(source_param), normalize_key(source_value), ""),
        (normalize_key(source_name), normalize_key(source_param), normalize_key(source_value), ""),
        (normalize_key(target_title), normalize_key(source_param), normalize_key(source_value), ""),
    ]
    value = value_rules["by_source_id"].get(keys[0], "")
    if not value:
        value = value_rules["by_source_name"].get(keys[1], "")
    if not value:
        value = value_rules["by_target"].get(keys[2], "")
    if value:
        PARAM_VALUE_PICK_CACHE[cache_key] = value
        return value

    value = allo_map.suggest_allowed_value(
        "",
        target_title,
        filter_name,
        source_param,
        source_value,
        allowed_values,
    )
    PARAM_VALUE_PICK_CACHE[cache_key] = value
    return value


def apply_price_and_availability(
    offer: ET._Element,
    rz: dict[str, str] | None,
    google_row: dict[str, str] | None,
) -> tuple[bool, bool]:
    changed_price = False
    changed_other = False
    source_price = maudau.child_text(offer, "price")
    source_available = maudau.extract_available(offer)

    if rz:
        final_price = maudau.pick_effective_price(source_price, rz.get("price", ""))
        changed_price = maudau.set_or_create(offer, "price", final_price) or changed_price
        final_old = maudau.pick_effective_old_price(final_price, rz.get("old_price", ""))
        if final_old:
            changed_other = maudau.set_or_create(offer, "old_price", final_old) or changed_other
        else:
            changed_other = maudau.clear_old_price(offer) or changed_other
        changed_other = maudau.set_available(offer, source_available) or changed_other
        return changed_price, changed_other

    if google_row:
        final_price = maudau.pick_effective_price(source_price, google_row.get("price", ""))
        changed_price = maudau.set_or_create(offer, "price", final_price) or changed_price
        final_old = maudau.pick_effective_old_price(final_price, google_row.get("old_price", ""))
        if final_old:
            changed_other = maudau.set_or_create(offer, "old_price", final_old) or changed_other
        else:
            changed_other = maudau.clear_old_price(offer) or changed_other
        changed_other = maudau.set_available(offer, source_available) or changed_other
        return changed_price, changed_other

    changed_other = maudau.set_available(offer, source_available) or changed_other
    return changed_price, changed_other


def make_text_child(parent: ET._Element, tag: str, value: str | None) -> ET._Element | None:
    value = clean_text(value)
    if not value:
        return None
    node = ET.SubElement(parent, tag)
    node.text = value
    return node


def html_to_text(value: str | None) -> str:
    return maudau.compact_text(maudau.HTML_TAG_RE.sub(" ", value or ""))


def get_section_fixed_rules() -> list[allo_map.SectionFixedRule]:
    global SECTION_FIXED_RULES_CACHE
    if SECTION_FIXED_RULES_CACHE is None:
        SECTION_FIXED_RULES_CACHE = allo_map.load_section_fixed_rules()
    return SECTION_FIXED_RULES_CACHE


def matching_section_rules(source_name: str) -> list[allo_map.SectionFixedRule]:
    return [
        rule for rule in get_section_fixed_rules()
        if allo_map.source_category_matches_rule(source_name, rule.selector)
    ]


def get_filter_names_cached(target_title: str, target_url: str, allo_categories: dict[str, dict]) -> list[str]:
    key = (target_title, target_url)
    if key not in FILTER_NAMES_CACHE:
        names = allo_map.get_category_filter_names(target_title, target_url, allo_categories)
        FILTER_NAMES_CACHE[key] = sorted({allo_map.canonical_filter_name(name) for name in names if clean_text(name)})
    return FILTER_NAMES_CACHE[key]


def get_filter_meta_cached(
    target_title: str,
    target_url: str,
    filter_name: str,
    allo_categories: dict[str, dict],
) -> dict:
    filter_name = allo_map.canonical_filter_name(filter_name)
    key = (target_title, target_url, filter_name)
    if key not in FILTER_META_CACHE:
        FILTER_META_CACHE[key] = allo_map.get_filter_meta(target_title, target_url, filter_name, allo_categories)
    return FILTER_META_CACHE[key]


def category_rule_applies(rule_category: str, target_title: str) -> bool:
    if not rule_category:
        return True
    target_key = normalize_key(target_title)
    parts = [normalize_key(part) for part in rule_category.split("/") if clean_text(part)]
    return target_key in parts or normalize_key(rule_category) == target_key


def format_pump_param_value(source_value: str, rule: dict) -> str:
    value = clean_text(source_value)
    if not value:
        return ""
    number = maudau.normalize_sheet_number(value)
    fmt = rule.get("format")
    if fmt == "liters" and number:
        return f"{number} л"
    if fmt == "kw" and number:
        return number
    return value


def first_size_match(text: str) -> tuple[float, float, float | None] | None:
    normalized = clean_text(text).lower().replace("х", "x").replace("*", "x")
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*x\s*(\d+(?:[.,]\d+)?)(?:\s*x\s*(\d+(?:[.,]\d+)?))?", normalized)
    if not match:
        return None
    first = float(match.group(1).replace(",", "."))
    second = float(match.group(2).replace(",", "."))
    third = float(match.group(3).replace(",", ".")) if match.group(3) else None
    return first, second, third


def normalize_unit_token(value: str | None) -> str | None:
    token = clean_text(value).lower()
    if token in {"мм", "mm"}:
        return "mm"
    if token in {"см", "cm"}:
        return "cm"
    return None


def extract_size_candidates(text: str) -> list[tuple[float, float, float | None, str | None, str | None, str | None]]:
    normalized = clean_text(text).lower().replace("х", "x").replace("*", "x")
    pattern = re.compile(
        r"(?<!\d)"
        r"(\d{2,4}(?:[.,]\d+)?)\s*(мм|mm|см|cm)?\s*x\s*"
        r"(\d{2,4}(?:[.,]\d+)?)\s*(мм|mm|см|cm)?"
        r"(?:\s*x\s*(\d{2,4}(?:[.,]\d+)?)\s*(мм|mm|см|cm)?)?"
        r"(?!\d)",
        flags=re.IGNORECASE,
    )
    out: list[tuple[float, float, float | None, str | None, str | None, str | None]] = []
    for match in pattern.finditer(normalized):
        try:
            first = float(match.group(1).replace(",", "."))
            second = float(match.group(3).replace(",", "."))
            third = float(match.group(5).replace(",", ".")) if match.group(5) else None
        except ValueError:
            continue
        out.append(
            (
                first,
                second,
                third,
                normalize_unit_token(match.group(2)),
                normalize_unit_token(match.group(4)),
                normalize_unit_token(match.group(6)),
            )
        )
    return out


def convert_size_candidate(
    values: tuple[float, float, float | None],
    units: tuple[str | None, str | None, str | None],
    to_unit: str,
    default_unknown_unit: str = "cm",
) -> tuple[float, float, float | None]:
    non_null_values = [v for v in values if v is not None]
    explicit_units = [u for u in units if u]
    fallback_unit = default_unknown_unit
    if explicit_units:
        # Если явно указали одну систему единиц — используем её для пустых значений.
        if len(set(explicit_units)) == 1:
            fallback_unit = explicit_units[0]
        else:
            # Смешанные единицы — подстраховка по диапазону.
            fallback_unit = "mm" if any(v > 300 for v in non_null_values) else default_unknown_unit
    else:
        fallback_unit = "mm" if any(v > 300 for v in non_null_values) else default_unknown_unit

    def convert_single(value: float | None, unit: str | None) -> float | None:
        if value is None:
            return None
        src_unit = unit or fallback_unit
        if to_unit == "mm":
            return value if src_unit == "mm" else value * 10.0
        # to_unit == "cm"
        return value / 10.0 if src_unit == "mm" else value

    first = convert_single(values[0], units[0]) or 0.0
    second = convert_single(values[1], units[1]) or 0.0
    third = convert_single(values[2], units[2])
    return first, second, third


def extract_best_size_mm(text: str) -> tuple[float, float, float | None] | None:
    best: tuple[float, float, float | None] | None = None
    best_score = float("-inf")
    for first, second, third, u1, u2, u3 in extract_size_candidates(text):
        mm_first, mm_second, mm_third = convert_size_candidate(
            (first, second, third),
            (u1, u2, u3),
            to_unit="mm",
            default_unknown_unit="cm",
        )
        if not (40 <= mm_first <= 3000 and 40 <= mm_second <= 3000):
            continue
        score = 0.0
        if mm_first >= 300:
            score += 3.0
        if mm_second >= 300:
            score += 3.0
        if mm_third is not None:
            if 60 <= mm_third <= 800:
                score += 4.0
            else:
                score -= 2.0
        # Предпочитаем записи со "здоровой" геометрией чаши.
        if mm_first >= mm_second:
            score += 1.0
        score += (mm_first + mm_second) / 5000.0
        if score > best_score:
            best_score = score
            best = (mm_first, mm_second, mm_third)
    return best


def extract_best_footprint_cm(text: str, min_side: float = 60.0, max_side: float = 250.0) -> tuple[float, float] | None:
    best: tuple[float, float] | None = None
    best_score = float("-inf")
    for first, second, third, u1, u2, u3 in extract_size_candidates(text):
        cm_first, cm_second, _cm_third = convert_size_candidate(
            (first, second, third),
            (u1, u2, u3),
            to_unit="cm",
            default_unknown_unit="cm",
        )
        if not (min_side <= cm_first <= max_side and min_side <= cm_second <= max_side):
            continue
        lo, hi = min(cm_first, cm_second), max(cm_first, cm_second)
        score = (hi * lo) / 100.0
        if abs(hi - lo) <= 2:
            score += 2.0
        if score > best_score:
            best_score = score
            best = (hi, lo)
    return best


def extract_labeled_dimension_mm(text: str, labels: list[str]) -> float | None:
    normalized = clean_text(text).lower().replace(",", ".")
    for label in labels:
        label_pat = re.escape(label.lower())
        match = re.search(
            rf"{label_pat}[^\d]{{0,25}}(\d{{2,4}}(?:\.\d+)?)\s*(мм|mm|см|cm)?",
            normalized,
            flags=re.IGNORECASE,
        )
        if not match:
            continue
        try:
            value = float(match.group(1))
        except ValueError:
            continue
        unit = normalize_unit_token(match.group(2))
        if unit == "mm":
            return value
        if unit == "cm":
            return value * 10.0
        return value if value > 300 else value * 10.0
    return None


def normalize_to_cm(value: float | None) -> float | None:
    if value is None:
        return None
    if value > 300:
        return value / 10.0
    return value


def normalize_to_mm(value: float | None) -> float | None:
    if value is None:
        return None
    if value <= 300:
        return value * 10.0
    return value


def to_int_string(value: float | None) -> str:
    if value is None:
        return ""
    rounded = int(round(value))
    return str(rounded)


def extract_first_number(value: str | None) -> float | None:
    text = clean_text(value).replace(",", ".")
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def extract_all_numbers(value: str | None) -> list[float]:
    text = clean_text(value).replace(",", ".")
    result: list[float] = []
    for token in re.findall(r"\d+(?:\.\d+)?", text):
        try:
            result.append(float(token))
        except ValueError:
            continue
    return result


def extract_number_before_unit(value: str | None, unit: str) -> float | None:
    text = clean_text(value).replace(",", ".")
    match = re.search(rf"(\d+(?:\.\d+)?)\s*{re.escape(unit)}\b", text, flags=re.IGNORECASE)
    if not match:
        return None
    try:
        return float(match.group(1))
    except ValueError:
        return None


def extract_model_width_hint(value: str | None) -> float | None:
    text = clean_text(value).lower().replace("–", "-").replace("—", "-")
    for pat in (
        r"\b[пшp]-\s*(\d{2,3})\b",
        r"\bш\s*-\s*(\d{2,3})\b",
        r"\bп\s*-\s*(\d{2,3})\b",
    ):
        match = re.search(pat, text, flags=re.IGNORECASE)
        if match:
            try:
                return float(match.group(1))
            except ValueError:
                continue
    return None


def build_param_pairs(
    offer: ET._Element,
    source_id: str,
    source_name: str,
    target_title: str,
    target_url: str,
    allo_categories: dict[str, dict],
    param_rules: dict,
    value_rules: dict,
) -> tuple[list[tuple[str, str]], list[str], bool, list[tuple[str, str, str]]]:
    params = offer_params(offer)
    filter_names = get_filter_names_cached(target_title, target_url, allo_categories)
    result: list[tuple[str, str]] = []
    issues: list[str] = []
    color_fallback_used = False
    color_fallback_changes: list[tuple[str, str, str]] = []
    color_fallback_seen: set[tuple[str, str, str]] = set()
    seen: set[tuple[str, str]] = set()
    text_blob = " ".join(
        [
            maudau.child_text(offer, "name"),
            maudau.child_text(offer, "name_ru"),
            maudau.child_text(offer, "name_ua"),
            maudau.child_text(offer, "description"),
            maudau.child_text(offer, "description_ru"),
            maudau.child_text(offer, "description_ua"),
        ]
    )
    text_key = normalize_key(text_blob)

    def add_pair(filter_name: str, target_value: str) -> None:
        filter_name = allo_map.canonical_filter_name(filter_name)
        filter_name = clean_text(filter_name)
        target_value = clean_text(target_value)
        if not filter_name or not target_value:
            return
        if target_value.endswith(" ?"):
            target_value = target_value[:-2].strip()
        if target_value.endswith("?"):
            target_value = target_value[:-1].strip()
        key = (filter_name, target_value)
        if key in seen:
            return
        seen.add(key)
        result.append(key)

    def replace_pair(filter_name: str, target_value: str) -> None:
        filter_name = allo_map.canonical_filter_name(filter_name)
        filter_name = clean_text(filter_name)
        target_value = clean_text(target_value)
        if not filter_name or not target_value:
            return
        for existing in list(result):
            if existing[0] == filter_name:
                result.remove(existing)
                seen.discard(existing)
        add_pair(filter_name, target_value)

    def remove_filter(filter_name: str) -> None:
        filter_name = allo_map.canonical_filter_name(filter_name)
        for existing in list(result):
            if existing[0] == filter_name:
                result.remove(existing)
                seen.discard(existing)

    def param_value(*names: str) -> str:
        return find_offer_param(params, *names)

    def has_filter(filter_name: str) -> bool:
        filter_name = allo_map.canonical_filter_name(filter_name)
        return any(name == filter_name for name, _value in result)

    def replace_if_filter_exists(filter_name: str, value: str) -> None:
        filter_name = allo_map.canonical_filter_name(filter_name)
        if filter_name in filter_names:
            replace_pair(filter_name, value)

    def add_if_filter_exists(filter_name: str, value: str) -> None:
        filter_name = allo_map.canonical_filter_name(filter_name)
        if filter_name in filter_names:
            add_pair(filter_name, value)

    def note_color_fallback(filter_name: str, source_value: str, target_value: str) -> None:
        clean_filter = clean_text(allo_map.canonical_filter_name(filter_name))
        clean_source = clean_text(source_value)
        clean_target = clean_text(target_value)
        if not clean_filter or not clean_source or not clean_target:
            return
        if normalize_key(clean_source) == normalize_key(clean_target):
            return
        key = (clean_filter, clean_source, clean_target)
        if key in color_fallback_seen:
            return
        color_fallback_seen.add(key)
        color_fallback_changes.append(key)

    def choose_allowed(filter_name: str, source_param_name: str, source_value: str) -> str:
        nonlocal color_fallback_used
        filter_name = allo_map.canonical_filter_name(filter_name)
        source_value = clean_text(source_value)
        if not source_value:
            return ""
        source_value = source_value.replace(" ?", "").rstrip("?").strip()
        meta = get_filter_meta_cached(target_title, target_url, filter_name, allo_categories)
        allowed_values = sorted(meta.get("values", {}).keys())
        if allowed_values:
            for allowed in allowed_values:
                if normalize_key(allowed) == normalize_key(source_value):
                    return allowed
        suggested = pick_param_value(
            source_id,
            source_name,
            target_title,
            source_param_name,
            filter_name,
            source_value,
            allowed_values,
            value_rules,
        )
        if suggested:
            cleaned = suggested.replace(" ?", "").rstrip("?").strip()
            for allowed in allowed_values:
                if normalize_key(allowed) == normalize_key(cleaned):
                    return allowed
            return cleaned
        return ""

    def pick_nearest_allowed_by_number(filter_name: str, number: float) -> str:
        filter_name = allo_map.canonical_filter_name(filter_name)
        meta = get_filter_meta_cached(target_title, target_url, filter_name, allo_categories)
        allowed_values = sorted(meta.get("values", {}).keys())
        best_val = ""
        best_score = float("inf")
        for allowed in allowed_values:
            text = clean_text(allowed).lower().replace(",", ".")
            nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", text)]
            if not nums:
                continue
            if "до" in text and len(nums) >= 1:
                # До N
                if number <= nums[0]:
                    return allowed
                score = abs(number - nums[0]) + 1.0
            elif ("более" in text or "и более" in text) and len(nums) >= 1:
                # Более N
                if number >= nums[0]:
                    return allowed
                score = abs(number - nums[0]) + 1.0
            elif len(nums) >= 2:
                lo, hi = min(nums[0], nums[1]), max(nums[0], nums[1])
                if lo <= number <= hi:
                    return allowed
                score = min(abs(number - lo), abs(number - hi))
            else:
                score = abs(number - nums[0])
            if score < best_score:
                best_score = score
                best_val = allowed
        return best_val

    def pick_middle_allowed_value(filter_name: str) -> str:
        """Возвращает среднее (медианное) значение из канонических значений ALLO."""
        filter_name = allo_map.canonical_filter_name(filter_name)
        meta = get_filter_meta_cached(target_title, target_url, filter_name, allo_categories)
        allowed_values = sorted(meta.get("values", {}).keys())
        if not allowed_values:
            return ""
        numeric_candidates: list[tuple[float, str]] = []
        for allowed in allowed_values:
            text = clean_text(allowed).lower().replace(",", ".")
            nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", text)]
            if not nums:
                continue
            if "до" in text:
                representative = nums[0]
            elif "более" in text or "и более" in text:
                representative = nums[0]
            elif len(nums) >= 2:
                representative = (min(nums[0], nums[1]) + max(nums[0], nums[1])) / 2.0
            else:
                representative = nums[0]
            numeric_candidates.append((representative, allowed))
        if numeric_candidates:
            numeric_candidates.sort(key=lambda x: x[0])
            return numeric_candidates[len(numeric_candidates) // 2][1]
        return allowed_values[len(allowed_values) // 2]

    def add_shower_system_fixed_rules() -> None:
        if target_title != "Душевые системы":
            return
        features_key = normalize_key(param_value("Особенности", "Особливості"))
        mount_key = normalize_key(param_value("Монтаж", "Установка"))
        diameter_key = normalize_key(param_value("Диаметр душа, мм"))

        if "без смесителя" in features_key:
            replace_pair("Наличие смесителя", "Нет")
        elif "смеситель" in features_key:
            replace_pair("Наличие смесителя", "Есть")
        elif not has_filter("Наличие смесителя"):
            replace_if_filter_exists("Наличие смесителя", "Есть")

        if "термостат" in features_key:
            replace_if_filter_exists("Термостат", "Есть")
            replace_if_filter_exists("Наличие смесителя", "Есть")
        elif not has_filter("Термостат"):
            replace_if_filter_exists("Термостат", "Нет")

        if "наруж" in mount_key:
            replace_pair("Тип монтажа", "Наружный")
        elif "наполь" in mount_key:
            replace_pair("Тип монтажа", "Напольный")
        elif "скры" in mount_key:
            replace_pair("Тип монтажа", "Скрытый")

        if diameter_key:
            if "менее 100" in diameter_key:
                replace_pair("Диаметр лейки (верхний душ)", "51 - 100 мм")
                replace_pair("Тип леек", "Ручной душ")
            elif any(token in diameter_key for token in ("100-199", "200-299", "300-399", "400-499", "более 500")):
                replace_pair("Диаметр лейки (верхний душ)", "Более 150 мм" if "100-199" not in diameter_key else "101 - 150 мм")
                replace_pair("Тип леек", "Верхний + ручной душ")
        if not has_filter("Тип леек"):
            replace_if_filter_exists("Тип леек", "Верхний + ручной душ")
        if not has_filter("Тип монтажа"):
            replace_if_filter_exists("Тип монтажа", "Наружный")

    def add_kitchen_sink_fixed_rules() -> None:
        if target_title != "Кухонные мойки":
            return
        shape_key = normalize_key(param_value("Форма", "Форма мойки"))
        width_mm: float | None = None
        depth_mm: float | None = None
        height_mm: float | None = None

        # 1) Приоритет: явные параметры.
        width_mm = normalize_to_mm(
            extract_first_number(
                param_value(
                    "Ширина мойки, мм",
                    "Ширина (мм)",
                    "Ширина изделия, мм",
                    "Ширина, мм",
                    "Ширина",
                )
            )
        )
        depth_mm = normalize_to_mm(
            extract_first_number(
                param_value(
                    "Глубина мойки, мм",
                    "Глубина (мм)",
                    "Глубина, мм",
                    "Глубина",
                    "Длина (мм)",
                    "Длина, мм",
                    "Длина",
                )
            )
        )
        height_mm = normalize_to_mm(
            extract_first_number(
                param_value(
                    "Высота мойки, мм",
                    "Высота (мм)",
                    "Высота, мм",
                    "Высота",
                    "Глубина чаши (мм)",
                    "Глубина чаши, мм",
                    "Глибина чаші (мм)",
                    "Глибина чаші, мм",
                )
            )
        )

        # 2) Вытягиваем из описания по подписям, если в params пусто.
        if width_mm is None:
            width_mm = extract_labeled_dimension_mm(
                text_blob,
                [
                    "ширина мойки",
                    "ширина изделия",
                    "ширина",
                ],
            )
        if depth_mm is None:
            depth_mm = extract_labeled_dimension_mm(
                text_blob,
                [
                    "глубина мойки",
                    "длина изделия",
                    "длина",
                    "глубина",
                ],
            )
        if height_mm is None:
            height_mm = extract_labeled_dimension_mm(
                text_blob,
                [
                    "высота мойки",
                    "высота изделия",
                    "глубина чаши",
                    "глибина чаші",
                    "высота",
                ],
            )

        # 3) Размеры формата 770x500x180 (или 49x62 и т.п.) из названия/описания.
        size_mm = extract_best_size_mm(text_blob)
        if size_mm:
            sx, sy, sz = size_mm
            if width_mm is None:
                width_mm = sx
            if depth_mm is None:
                depth_mm = sy
            if height_mm is None and sz is not None:
                height_mm = sz

            # Круглая мойка: "490x180" обычно диаметр + высота чаши.
            if "круг" in shape_key and sz is None and sx >= 350 and sy <= 300:
                if height_mm is None:
                    height_mm = sy
                width_mm = max(sx, sy)
                depth_mm = width_mm

        # 4) "44см - 62см" => 620x440 мм, если иных размеров нет.
        if width_mm is None or depth_mm is None:
            range_match = re.search(
                r"(\d{2,3}(?:[.,]\d+)?)\s*см\s*[-–]\s*(\d{2,3}(?:[.,]\d+)?)\s*см",
                clean_text(text_blob),
                flags=re.IGNORECASE,
            )
            if range_match:
                a = float(range_match.group(1).replace(",", ".")) * 10.0
                b = float(range_match.group(2).replace(",", ".")) * 10.0
                if width_mm is None:
                    width_mm = max(a, b)
                if depth_mm is None:
                    depth_mm = min(a, b)

        # 5) Если круглая и есть только один размер/диаметр — второй = первому.
        if "круг" in shape_key:
            if width_mm is not None and depth_mm is None:
                depth_mm = width_mm
            if depth_mm is not None and width_mm is None:
                width_mm = depth_mm
            if width_mm is not None and depth_mm is not None:
                # Для круглой мойки размеры по плоскости должны быть равны.
                max_side = max(width_mm, depth_mm)
                width_mm = max_side
                depth_mm = max_side

        # 6) Санитарная проверка: высота мойки не может быть > ширины/глубины/длины.
        if height_mm is not None and width_mm is not None and depth_mm is not None:
            plan_min = min(width_mm, depth_mm)
            if height_mm > plan_min:
                # Частый кейс: 1800 вместо 180.
                if (height_mm / 10.0) <= plan_min:
                    height_mm = height_mm / 10.0
                else:
                    height_mm = None

        if width_mm is not None:
            replace_pair("Ширина мойки, мм", to_int_string(width_mm))
        if depth_mm is not None:
            replace_pair("Глубина мойки, мм", to_int_string(depth_mm))
        if height_mm is not None:
            replace_pair("Высота мойки, мм", to_int_string(height_mm))

        if "без перелив" in text_key:
            replace_pair("Перелив", "Нет")
        elif "перелив" in text_key:
            replace_pair("Перелив", "Есть")

        if "без отверстия" in text_key and "смесител" in text_key:
            replace_pair("Отверстие под смеситель", "Нет")
        elif "отверстие под смесител" in text_key:
            replace_pair("Отверстие под смеситель", "Есть")

        # Канон для этой категории по шаблону ALLO: "Поверхня".
        surface = ""
        if any(token in text_key for token in ("декор", "рифлен")):
            surface = "Декор (рифленая)"
        elif any(token in text_key for token in ("матов", "brushed", "satin", "сатин")):
            surface = "Матовая"
        elif any(token in text_key for token in ("глян", "полирован", "polish")):
            surface = "Глянцевая"
        else:
            surface = "Матовая"
        replace_if_filter_exists("Поверхня", surface)

    def add_shower_cabin_fixed_rules() -> None:
        if target_title != "Душевые кабины":
            return
        size_value = param_value("Размер", "Розмір")
        size_triplet = first_size_match(size_value) if size_value else None
        width_cm = None
        depth_cm = None
        if size_triplet:
            width_cm = normalize_to_cm(size_triplet[0])
            depth_cm = normalize_to_cm(size_triplet[1])
        if width_cm is None or depth_cm is None:
            footprint_from_text = extract_best_footprint_cm(text_blob, min_side=60.0, max_side=150.0)
            if footprint_from_text:
                if width_cm is None:
                    width_cm = max(footprint_from_text)
                if depth_cm is None:
                    depth_cm = min(footprint_from_text)
        if width_cm is None:
            width_cm = normalize_to_cm(extract_first_number(param_value("Ширина, см", "ширина, см", "Ширина")))
        if depth_cm is None:
            depth_cm = normalize_to_cm(extract_first_number(param_value("Глубина, см", "глубина, см", "Глубина")))
        if width_cm is None:
            # Иногда ширина уже сопоставляется обычной логикой до фикс-правил.
            mapped_width = next((v for n, v in result if n == "Ширина"), "")
            width_cm = normalize_to_cm(extract_first_number(mapped_width))
        if width_cm is not None and width_cm < 70:
            width_cm = None
        if depth_cm is not None and depth_cm < 60:
            depth_cm = None
        height_cm = None

        # 1) Высоту кабины берем из явных полей (приоритет над "90x90x14").
        direct_height_candidates = [
            param_value("Высота, см", "Высота"),
            param_value("Высота, мм"),
            param_value("Высота стекол, мм", "Высота стекла, мм"),
            param_value("Габариты изделия (ВхШхД), мм", "Габариты изделия, мм"),
            param_value("Габарити виробу (ВхШхД), мм", "Габарити виробу, мм"),
            param_value("Габариты изделия (ВхШхД), см", "Габариты изделия, см"),
            param_value("Габарити виробу (ВхШхД), см", "Габарити виробу, см"),
        ]
        for raw in direct_height_candidates:
            n = extract_first_number(raw)
            if n is None:
                continue
            # Для мм/крупных чисел -> см.
            cand_cm = normalize_to_cm(n)
            if cand_cm is not None and cand_cm >= 70:
                height_cm = cand_cm
                break

        # 2) Пробуем вытащить высоту из описаний с явными маркерами.
        if height_cm is None:
            text_norm = clean_text(text_blob).lower().replace("х", "x").replace("*", "x")
            labeled_patterns = [
                # Габариты изделия (ВхШхД), мм : 1850х...
                r"габарит[^\n:]{0,40}издел[^\n:]{0,40}:\s*(\d{3,4}(?:[.,]\d+)?)\s*x",
                # Высота стекол, мм : 1850
                r"высот[^\n:]{0,20}стек[^\n:]{0,20}:\s*(\d{3,4}(?:[.,]\d+)?)",
                # Высота с поддоном : 200 см / Высота : 195 см
                r"высот[^\n:]{0,30}:\s*(\d{2,4}(?:[.,]\d+)?)\s*(?:мм|см|mm|cm)?",
                # Українські аналоги
                r"габарит[^\n:]{0,40}вироб[^\n:]{0,40}:\s*(\d{3,4}(?:[.,]\d+)?)\s*x",
                r"висот[^\n:]{0,20}скл[^\n:]{0,20}:\s*(\d{3,4}(?:[.,]\d+)?)",
                r"висот[^\n:]{0,30}:\s*(\d{2,4}(?:[.,]\d+)?)\s*(?:мм|см|mm|cm)?",
            ]
            for pat in labeled_patterns:
                m = re.search(pat, text_norm, flags=re.IGNORECASE)
                if not m:
                    continue
                try:
                    n = float(m.group(1).replace(",", "."))
                except ValueError:
                    continue
                cand_cm = normalize_to_cm(n)
                if cand_cm is not None and cand_cm >= 70:
                    height_cm = cand_cm
                    break

        # 3) Только если ничего не нашли: берём 3-й/1-й размер из триплетов, но
        # игнорируем мелкие значения (например, высоту поддона 14 см).
        if height_cm is None:
            for first, second, third, u1, u2, u3 in extract_size_candidates(text_blob):
                if third is None:
                    continue
                cm_first, cm_second, cm_third = convert_size_candidate(
                    (first, second, third),
                    (u1, u2, u3),
                    to_unit="cm",
                    default_unknown_unit="cm",
                )
                if cm_third is not None and 150 <= cm_third <= 260:
                    height_cm = cm_third
                    break
                if 150 <= cm_first <= 260 and 60 <= cm_second <= 150:
                    # Формат ВхШхД.
                    height_cm = cm_first
                    if width_cm is None:
                        width_cm = max(cm_second, cm_third or cm_second)
                    break

        # 4) Бизнес-ограничение: высота кабины не может быть ниже 150 см.
        if height_cm is not None and height_cm < 150:
            height_cm = None
        if height_cm is None:
            height_cm = 185.0

        if width_cm and height_cm:
            replace_pair("Размер (ШxВ)", f"{to_int_string(width_cm)} x {to_int_string(height_cm)} см")
            # Держим отдельную "Ширина" синхронной с рассчитанным размером,
            # чтобы не оставалось старых/ошибочных значений из авто-сопоставления.
            resolved_width = (
                choose_allowed("Ширина", "Ширина", f"{to_int_string(width_cm)} см")
                or pick_nearest_allowed_by_number("Ширина", width_cm)
                or to_int_string(width_cm)
            )
            if resolved_width:
                replace_if_filter_exists("Ширина", resolved_width)
        if depth_cm is not None:
            resolved_depth = (
                choose_allowed("Глубина", "Глубина", f"{to_int_string(depth_cm)} см")
                or pick_nearest_allowed_by_number("Глубина", depth_cm)
            )
            if resolved_depth:
                replace_if_filter_exists("Глубина", resolved_depth)

        # Поддон: без поддона -> Нет, мелкий/глубокий/средний -> Есть.
        poddon_raw = normalize_key(param_value("Поддон", "поддон"))
        title_raw = normalize_key(
            " ".join(
                [
                    maudau.child_text(offer, "name_ru"),
                    maudau.child_text(offer, "name"),
                    text_blob,
                ]
            )
        )
        poddon_probe = " ".join([poddon_raw, title_raw])
        if "без поддон" in poddon_probe or "без поддона" in poddon_probe:
            replace_if_filter_exists("Поддон", "Нет")
        elif any(token in poddon_probe for token in ("мелк", "глубок", "средн", "поддон есть", "на поддоне", "с поддоном")):
            replace_if_filter_exists("Поддон", "Есть")

        poddon_material = param_value("материал поддона", "Материал поддона")
        if poddon_material:
            resolved = choose_allowed("Материал поддона", "материал поддона", poddon_material)
            if resolved:
                replace_pair("Материал поддона", resolved)

        replace_if_filter_exists("Тип", "Душевая кабина")
        replace_if_filter_exists("Конструкция", "Профильная")

    def add_umyvalniki_fixed_rules() -> None:
        if target_title != "Умывальники":
            return
        width = normalize_to_cm(extract_first_number(param_value("ширина, см")))
        depth = normalize_to_cm(extract_first_number(param_value("глубина, см")))
        height = normalize_to_cm(extract_first_number(param_value("высота, см")))
        if (width is None or depth is None) and first_size_match(text_blob):
            size = first_size_match(text_blob)
            if width is None:
                width = normalize_to_cm(size[0])
            if depth is None:
                depth = normalize_to_cm(size[1])
            if height is None and size[2]:
                height = normalize_to_cm(size[2])

        if width is not None:
            resolved = choose_allowed("Ширина", "ширина, см", to_int_string(width))
            if resolved:
                replace_pair("Ширина", resolved)
        if depth is not None:
            resolved = choose_allowed("Глубина", "глубина, см", to_int_string(depth))
            if resolved:
                replace_pair("Глубина", resolved)
        if height is not None:
            resolved = choose_allowed("Высота", "высота, см", to_int_string(height))
            if resolved:
                replace_pair("Высота", resolved)

        if "без перелив" in text_key:
            replace_pair("Перелив", "Нет")
        elif "перелив" in text_key:
            replace_pair("Перелив", "Есть")

        install_key = normalize_key(param_value("монтаж", "Монтаж", "Установка", "Тип установки"))
        if "подвес" in install_key or "подвес" in text_key:
            replace_if_filter_exists("Тип установки", "Подвесной")
        elif "наклад" in install_key or "наклад" in text_key:
            replace_if_filter_exists("Тип установки", "Накладной")
        elif "врез" in install_key or "врез" in text_key:
            replace_if_filter_exists("Тип установки", "Врезной")
        elif "наполь" in install_key or "наполь" in text_key:
            replace_if_filter_exists("Тип установки", "Напольный")
        elif not has_filter("Тип установки"):
            replace_if_filter_exists("Тип установки", "Подвесной")

        material_key = normalize_key(" ".join([param_value("Материал", "Матеріал"), text_blob]))
        if "керам" in material_key:
            replace_if_filter_exists("Материал", "Керамика")
        elif "фарфор" in material_key:
            replace_if_filter_exists("Материал", "Сантехнический фарфор")
        elif "фаянс" in material_key:
            replace_if_filter_exists("Материал", "Сантехнический фаянс")
        elif "искусствен" in material_key or "камень" in material_key:
            replace_if_filter_exists("Материал", "Искусственный камень")
        elif not has_filter("Материал"):
            replace_if_filter_exists("Материал", "Керамика")

        # По вашему правилу: если высоту достать не удалось — ставим дефолт.
        if not has_filter("Высота"):
            replace_if_filter_exists("Высота", "11 - 20 см")

    def add_bath_fixed_rules() -> None:
        if target_title != "Ванны":
            return
        source_key = normalize_key(source_name)
        mat_key = normalize_key(param_value("Материал", "Матеріал"))

        # Материал ванны задается правилом раздела. Это не эвристика:
        # "Стальные ванны" всегда должны стать "Материал=Сталь" и т.д.
        if "акрил" in source_key or "акрил" in mat_key:
            replace_pair("Материал", "Акрил")
        elif "гидромассажн" in source_key:
            replace_pair("Материал", "Акрил")
        elif any(token in source_key for token in ("сталь", "стальные")) or "сталь" in mat_key:
            replace_pair("Материал", "Сталь")
        elif any(token in source_key for token in ("чугун", "чугунные")) or "чугун" in mat_key:
            replace_pair("Материал", "Чугун")
        elif any(token in source_key for token in ("из камня", "каменн", "мрамор", "кварил")) or any(token in mat_key for token in ("искусственн", "литьевой камень", "камень", "мрамор", "кварил")):
            replace_pair("Материал", "Искусственный камень")
        elif "акрил" in text_key:
            replace_pair("Материал", "Акрил")
        elif "чугун" in text_key:
            replace_pair("Материал", "Чугун")
        elif "сталь" in text_key:
            replace_pair("Материал", "Сталь")
        elif any(token in text_key for token in ("искусственн", "камень", "мрамор", "кварил")):
            replace_pair("Материал", "Искусственный камень")

        # Длина/ширина: берём значение источника и приводим к канону ALLO ("... см").
        length_raw = clean_text(param_value("Длинна, см", "Длина, см", "Довжина, см", "длина, см", "довжина, см", "Размер"))
        width_raw = clean_text(param_value("Ширина, см", "ширина, см"))
        size_match = first_size_match(length_raw or text_blob)
        length_num = normalize_to_cm(extract_first_number(length_raw))
        width_num = normalize_to_cm(extract_first_number(width_raw))
        if length_num is None and size_match:
            length_num = normalize_to_cm(size_match[0])
        if width_num is None and size_match:
            width_num = normalize_to_cm(size_match[1])

        if length_num is not None:
            length_with_unit = f"{to_int_string(length_num)} см"
            resolved_length = choose_allowed("Длина", "Длина, см", length_with_unit) or pick_nearest_allowed_by_number("Длина", length_num)
            if resolved_length:
                replace_if_filter_exists("Длина", resolved_length)
                replace_if_filter_exists("Довжина", resolved_length)
        if width_num is not None:
            width_with_unit = f"{to_int_string(width_num)} см"
            resolved_width = choose_allowed("Ширина", "Ширина, см", width_with_unit) or pick_nearest_allowed_by_number("Ширина", width_num)
            if resolved_width:
                replace_if_filter_exists("Ширина", resolved_width)

    def add_bath_panel_fixed_rules() -> None:
        if normalize_key(target_title) not in {"панели для ванн и поддонов", "панели для ванн"}:
            return
        replace_if_filter_exists("Вид", "Панель")

    def add_shower_box_fixed_rules() -> None:
        if target_title not in {"Душевые боксы", "Боксы гидромассажные"}:
            return
        doors_raw = clean_text(param_value("Тип дверей"))
        install_raw = normalize_key(param_value("Установка", "Монтаж", "форма", "Форма"))
        glass_raw = clean_text(param_value("Цвет стекла", "Тон стекла", "Тон скла"))

        if doors_raw:
            resolved_doors = choose_allowed("Тип дверей", "Тип дверей", doors_raw)
            if resolved_doors:
                replace_if_filter_exists("Тип дверей", resolved_doors)
        if not has_filter("Тип дверей"):
            replace_if_filter_exists("Тип дверей", "Раздвижные")

        if "угл" in install_raw:
            replace_if_filter_exists("Установка", "Угловая")
        elif "пристен" in install_raw:
            replace_if_filter_exists("Установка", "Пристенная")
        if not has_filter("Установка"):
            replace_if_filter_exists("Установка", "Угловая")

        if glass_raw:
            resolved = (
                choose_allowed("Цвет стекла", "Цвет стекла", glass_raw)
                or choose_allowed("Тон стекла", "Цвет стекла", glass_raw)
            )
            if resolved:
                replace_if_filter_exists("Цвет стекла", resolved)
                replace_if_filter_exists("Тон стекла", resolved)

    def add_siphon_fixed_rules() -> None:
        if target_title not in {"Сифоны и трапы", "Сифоны и гофры"}:
            return
        source_key = normalize_key(source_name)
        source_type_raw = clean_text(param_value("Тип", "Вид"))
        siphon_type_key = normalize_key(source_type_raw or param_value("Применение", "Призначення"))
        purpose_key = normalize_key(param_value("Назначение", "Призначення", "Применение"))
        title_probe = normalize_key(
            " ".join(
                [
                    maudau.child_text(offer, "name_ru"),
                    maudau.child_text(offer, "name"),
                ]
            )
        )
        probe = " ".join([title_probe, siphon_type_key, text_key])

        def set_application(value: str) -> None:
            resolved = choose_allowed("Применение", "Назначение", value) or value
            replace_if_filter_exists("Применение", resolved)

        def set_type(value: str) -> None:
            # Для "Сифоны и трапы" держим жёсткий канон, чтобы авто-подсказка
            # не уводила "Сифон" в "Трапы".
            canon_map = {
                "Донные клапаны": "Донные клапаны",
                "Пробки": "Пробки",
                "Сифон": "Сифоны",
                "Трапы": "Трапы",
            }
            resolved = canon_map.get(value, "")
            if not resolved:
                resolved = choose_allowed("Тип", "Тип", value) or value
            replace_if_filter_exists("Тип", resolved)

        # Тип
        # Приоритет 1: явное название товара/карточки.
        if "донн" in title_probe and "клап" in title_probe:
            set_type("Донные клапаны")
        elif "сифон" in title_probe:
            set_type("Сифон")
        elif "трап" in title_probe:
            set_type("Трапы")
        # Приоритет 2: строгая логика по исходному типу.
        elif "донн" in siphon_type_key and "клап" in siphon_type_key:
            set_type("Донные клапаны")
        elif "наклад" in siphon_type_key and "сифон" in siphon_type_key:
            set_type("Пробки")
        elif "пробк" in siphon_type_key:
            set_type("Пробки")
        elif "сифон" in siphon_type_key:
            set_type("Сифон")
        elif "трап" in siphon_type_key:
            set_type("Трапы")
        # Фолбэк по тексту карточки (приоритет 3).
        elif "донн" in probe and "клап" in probe:
            set_type("Донные клапаны")
        elif "сифон" in probe or "гофр" in probe:
            set_type("Сифон")
        elif "трап" in probe:
            set_type("Трапы")
        elif "комплектующ" in source_key and "душевые трапы" in source_key:
            set_type("Трапы")
        elif "комплектующ" in source_key:
            set_type("Сифон")

        # Применение: строгое сопоставление из правил (приоритет).
        # Жёсткий приоритет для кухонных комплектующих/аксессуаров:
        # если контекст категории "комплектующие/аксессуары" и в названии/описании
        # явно "для мойки" / "на мойку", то это всегда "Для мойки".
        kitchen_components_context = (
            ("комплектующ" in source_key or "аксессуар" in source_key)
            and any(token in probe for token in ("для мойк", "на мойк", "мойк"))
        )
        forced_application = False
        if kitchen_components_context:
            set_application("Для мойки")
            forced_application = True

        if purpose_key and not forced_application:
            if any(token in purpose_key for token in ("душевых поддонов", "для поддона", "поддонов")):
                set_application("Для поддона")
            elif any(token in purpose_key for token in ("для умывальника", "умывальник", "раковин")):
                set_application("Для раковины")
            elif any(token in purpose_key for token in ("стираль", "посудомоеч", "кухонных моек", "кухонной мойк", "для мойки")):
                set_application("Для мойки")
            elif any(token in purpose_key for token in ("универсаль", "кондиционер", "писсуар", "биде", "чаши генуя", "для ванн", "для ванной", "ванн")):
                set_application("Для ванной")
            elif "унитаз" in purpose_key:
                set_application("Для ванной")

        # Применение: эвристики только если значение ещё не выставлено.
        if not has_filter("Применение"):
            if "душевые трапы" in source_key:
                set_application("Для душа")
            elif any(token in source_key for token in ("душевых поддонов", "для поддона")) or "поддон" in probe:
                set_application("Для поддона")
            elif any(token in source_key for token in ("умывальник", "раковин")) or any(token in probe for token in ("умывальник", "раковин")):
                set_application("Для раковины")
            elif any(token in source_key for token in ("для ванн", "для ванной")) or "ванн" in probe:
                set_application("Для ванной")

        # Для новой агрегированной группы "Сифоны, гофры" закрываем обязательные поля дефолтом.
        if "сифоны гофры" in source_key or "сифоны, гофры" in source_key:
            if not has_filter("Тип"):
                set_type("Сифон")
            if not has_filter("Применение"):
                set_application("Для раковины")
        if "комплектующ" in source_key and "душевые трапы" in source_key and not has_filter("Применение"):
            set_application("Для душа")
        if not has_filter("Применение"):
            if any(v == "Трапы" for n, v in result if n == "Тип"):
                set_application("Для душа")
            else:
                set_application("Для раковины")

    def apply_section_rules_strict() -> None:
        # Правила из "Преобразование Разделов.xlsx" приоритетны:
        # они должны перезаписывать все авто-эвристики.
        for rule in matching_section_rules(source_name):
            if rule.skip_feed:
                continue
            for filter_name, target_value in rule.assignments:
                value_key = normalize_key(target_value)
                # В правилах встречаются не значения ALLO, а инструкции вида:
                # "Вид=Унитаз / Унитаз-биде по Тип унитаза". Такой текст
                # нельзя публиковать как значение фильтра; его закрывают
                # отдельные фикс-правила ниже.
                if " по " in f" {value_key} ":
                    continue
                replace_pair(filter_name, target_value)

    def add_shower_glass_group_fixed_rules() -> None:
        shower_glass_titles = {
            "Душевые двери",
            "Душевые стенки",
            "Душевые кабины",
            "Душевые боксы",
            "Боксы гидромассажные",
            "Шторки для ванной",
            "Шторки для ванн",
        }
        if target_title not in shower_glass_titles:
            return

        # Для стеклянных душевых групп не дублируем общий цвет.
        for existing in list(result):
            if existing[0] == "Цвет":
                result.remove(existing)
                seen.discard(existing)

        def normalize_glass_color(raw: str) -> str:
            key = normalize_key(raw)
            if not key:
                return ""
            if "прозрач" in key:
                return "Прозрачный"
            if "тонир" in key or "коричнев" in key:
                return "Тонированный"
            return "Матовый"

        doors_raw = clean_text(param_value("Тип дверей"))
        doors_key = normalize_key(doors_raw)
        glass_color_raw = clean_text(param_value("Цвет стекла", "Тон стекла", "Тон скла"))
        color_key = normalize_key(glass_color_raw)
        width_cm = normalize_to_cm(extract_first_number(param_value("ширина, см", "Ширина, см")))
        length_cm = normalize_to_cm(
            extract_first_number(
                param_value(
                    "длина, см",
                    "длинна, см",
                    "Довжина, см",
                    "Длина, см",
                    "Длина",
                    "Довжина",
                    "Размер",
                    "Розмір",
                )
            )
        )
        size_pair = first_size_match(text_blob)
        if width_cm is None and size_pair:
            width_cm = normalize_to_cm(size_pair[0])
        if length_cm is None and size_pair and size_pair[1] is not None:
            length_cm = normalize_to_cm(size_pair[1])

        if not glass_color_raw and any(token in text_key for token in ("прозрач", "тонир", "коричнев", "матов")):
            if "прозрач" in text_key:
                glass_color_raw = "прозрачный"
            elif "тонир" in text_key or "коричнев" in text_key:
                glass_color_raw = "тонированный"
            else:
                glass_color_raw = "матовый"
            color_key = normalize_key(glass_color_raw)

        glass_color = normalize_glass_color(glass_color_raw)
        if glass_color:
            resolved_glass = choose_allowed("Цвет стекла", "Цвет стекла", glass_color) or glass_color
            replace_if_filter_exists("Цвет стекла", resolved_glass)
            replace_if_filter_exists("Тон стекла", resolved_glass)
        if not has_filter("Цвет стекла"):
            replace_if_filter_exists("Цвет стекла", "Матовый")

        # Тип стекла зависит от цвета стекла.
        if "прозрач" in color_key or (has_filter("Цвет стекла") and any(v == "Прозрачный" for n, v in result if n == "Цвет стекла")):
            replace_if_filter_exists("Тип стекла", "Прозрачное")
            replace_if_filter_exists("Тип скла", "Прозрачное")
        elif any(token in color_key for token in ("тонир", "коричнев")) or (has_filter("Цвет стекла") and any(v == "Тонированный" for n, v in result if n == "Цвет стекла")):
            replace_if_filter_exists("Тип стекла", "Тонированное")
            replace_if_filter_exists("Тип скла", "Тонированное")
        else:
            replace_if_filter_exists("Тип стекла", "Матовое")
            replace_if_filter_exists("Тип скла", "Матовое")

        # Система открытия из Тип дверей.
        if "раздвиж" in doors_key:
            replace_if_filter_exists("Система открытия", "Раздвижная")
        elif "распаш" in doors_key:
            replace_if_filter_exists("Система открытия", "Распашная")
        elif "склад" in doors_key:
            replace_if_filter_exists("Система открытия", "Складная")
        else:
            replace_if_filter_exists("Система открытия", "Нет")

        # Обязательный дефолт по материалу витража.
        if not has_filter("Материал витража"):
            replace_if_filter_exists("Материал витража", "Стекло")

        # Для душевых стенок обязательно приводим ширину к канону ALLO.
        if target_title != "Душевые кабины" and width_cm is not None:
            width_with_unit = f"{to_int_string(width_cm)} см"
            resolved_width = choose_allowed("Ширина", "ширина, см", width_with_unit) or pick_nearest_allowed_by_number("Ширина", width_cm)
            if resolved_width:
                replace_if_filter_exists("Ширина", resolved_width)

        # Длина/довжина только если есть в источнике.
        if target_title != "Душевые кабины" and length_cm is not None:
            length_with_unit = f"{to_int_string(length_cm)} см"
            resolved_len = choose_allowed("Довжина", "длина, см", length_with_unit) or choose_allowed("Длина", "длина, см", length_with_unit)
            if not resolved_len:
                resolved_len = pick_nearest_allowed_by_number("Довжина", length_cm) or pick_nearest_allowed_by_number("Длина", length_cm)
            if resolved_len:
                replace_if_filter_exists("Довжина", resolved_len)
                replace_if_filter_exists("Длина", resolved_len)

    def add_obogrevateli_fixed_rules() -> None:
        if "обогревател" not in normalize_key(target_title):
            return

        def infer_area_from_similar(power_w: float | None) -> float | None:
            if power_w is None or power_w <= 0 or not HEATER_POWER_AREA_PROFILES:
                return None
            best_area = None
            best_score = float("inf")
            for ref_power_w, ref_area in HEATER_POWER_AREA_PROFILES:
                score = abs(ref_power_w - power_w)
                if score < best_score:
                    best_score = score
                    best_area = ref_area
            return best_area

        power_w_from_params = extract_first_number(param_value("Мощность, Вт"))
        power_kw_from_params = extract_first_number(param_value("Мощность, кВт"))
        if power_w_from_params is None and power_kw_from_params is not None:
            power_w_from_params = power_kw_from_params * 1000.0

        if not has_filter("Площадь обогрева"):
            area_param = extract_first_number(
                param_value(
                    "Площадь обогрева",
                    "Площадь обслуживания, кв. м",
                    "Рекомендуемая площадь помещения, кв.м.",
                    "Рекомендуемая площадь помещения, кв. м",
                )
            )
            if area_param is not None:
                resolved_area = pick_nearest_allowed_by_number("Площадь обогрева", area_param)
                if resolved_area:
                    replace_if_filter_exists("Площадь обогрева", resolved_area)
        if not has_filter("Площадь обогрева"):
            # Приоритет: "Площадь обогрева", затем "Рекомендуемая площадь помещения".
            m_area = re.search(
                r"площад[ья]\s+обогрева[^0-9]{0,25}(\d+(?:[.,]\d+)?)\s*(?:м2|м²)?",
                text_blob,
                flags=re.IGNORECASE,
            )
            if not m_area:
                m_area = re.search(
                    r"рекомендуем[а-яіїє]+\s+площад[ья]\s+помещен[ия][^0-9]{0,25}(\d+(?:[.,]\d+)?)\s*(?:м2|м²)?",
                    text_blob,
                    flags=re.IGNORECASE,
                )
            if m_area:
                sqm = float(m_area.group(1).replace(",", "."))
                resolved_area = pick_nearest_allowed_by_number("Площадь обогрева", sqm)
                if resolved_area:
                    replace_if_filter_exists("Площадь обогрева", resolved_area)
        if not has_filter("Площадь обогрева"):
            # Фолбэк: оцениваем площадь по мощности (~100 Вт на 1 м²).
            power_kw = power_kw_from_params
            if power_kw is not None and power_kw > 0:
                sqm = power_kw * 10.0
                resolved_area = pick_nearest_allowed_by_number("Площадь обогрева", sqm)
                if resolved_area:
                    replace_if_filter_exists("Площадь обогрева", resolved_area)
            else:
                m_w = re.search(
                    r"(?:номинальн[а-яіїє]+\s+)?мощност[ьі][^0-9]{0,25}(\d{3,5})\s*(?:вт|w)\b",
                    text_blob,
                    flags=re.IGNORECASE,
                )
                if m_w:
                    sqm = float(m_w.group(1)) / 100.0
                    resolved_area = pick_nearest_allowed_by_number("Площадь обогрева", sqm)
                    if resolved_area:
                        replace_if_filter_exists("Площадь обогрева", resolved_area)
                elif "1000w" in text_key or "1000 вт" in text_key:
                    replace_if_filter_exists("Площадь обогрева", pick_nearest_allowed_by_number("Площадь обогрева", 10.0))
                elif "1500w" in text_key or "1500 вт" in text_key:
                    replace_if_filter_exists("Площадь обогрева", pick_nearest_allowed_by_number("Площадь обогрева", 15.0))
                elif "2000w" in text_key or "2000 вт" in text_key:
                    replace_if_filter_exists("Площадь обогрева", pick_nearest_allowed_by_number("Площадь обогрева", 20.0))
        if not has_filter("Площадь обогрева"):
            similar_area = infer_area_from_similar(power_w_from_params)
            if similar_area is not None:
                resolved_area = pick_nearest_allowed_by_number("Площадь обогрева", similar_area)
                if resolved_area:
                    replace_if_filter_exists("Площадь обогрева", resolved_area)

        if not has_filter("Мощность"):
            power_kw = power_kw_from_params
            if power_kw is None:
                power_w = power_w_from_params
                if power_w is not None:
                    power_kw = power_w / 1000.0
            if power_kw is None:
                m_kw = re.search(r"мощност[ьі][^0-9]{0,25}(\d+(?:[.,]\d+)?)\s*(?:квт|kw)\b", text_blob, flags=re.IGNORECASE)
                if m_kw:
                    power_kw = float(m_kw.group(1).replace(",", "."))
            if power_kw is None:
                m_w = re.search(r"мощност[ьі][^0-9]{0,25}(\d{3,5})\s*(?:вт|w)\b", text_blob, flags=re.IGNORECASE)
                if m_w:
                    power_kw = float(m_w.group(1)) / 1000.0
            if power_kw is None and has_filter("Площадь обогрева"):
                # Последний фолбэк: обратная оценка мощности из площади (100 Вт на 1 м²).
                area_val = next((v for n, v in result if n == "Площадь обогрева"), "")
                nums = [float(x.replace(",", ".")) for x in re.findall(r"\d+(?:[.,]\d+)?", area_val)]
                if nums:
                    if len(nums) >= 2:
                        area_mid = (min(nums[0], nums[1]) + max(nums[0], nums[1])) / 2.0
                    else:
                        area_mid = nums[0]
                    power_kw = area_mid / 10.0
            if power_kw is not None:
                resolved_power = pick_nearest_allowed_by_number("Мощность", power_kw)
                if resolved_power:
                    replace_if_filter_exists("Мощность", resolved_power)

        if not has_filter("Терморегулятор"):
            replace_if_filter_exists("Терморегулятор", "Нет")

    def add_installation_buttons_fixed_rules() -> None:
        source_key = normalize_key(source_name)
        target_key = normalize_key(target_title)
        if "кнопк" in source_key and "инстал" in source_key:
            replace_if_filter_exists("Тип", "Клавиша смыва")
        elif "кнопк" in target_key and "инстал" in target_key:
            replace_if_filter_exists("Тип", "Клавиша смыва")

    def add_ventilyatory_fixed_rules() -> None:
        if target_title != "Вытяжные вентиляторы":
            return
        duct_num = extract_first_number(param_value("Диаметр воздуховода", "Диаметр патрубка, мм"))
        if duct_num is not None:
            resolved_duct = choose_allowed("Диаметр воздуховода", "Диаметр патрубка, мм", str(int(round(duct_num))))
            if not resolved_duct:
                resolved_duct = pick_nearest_allowed_by_number("Диаметр воздуховода", duct_num)
            if resolved_duct:
                replace_if_filter_exists("Диаметр воздуховода", resolved_duct)
        elif not has_filter("Диаметр воздуховода"):
            resolved_duct = pick_nearest_allowed_by_number("Диаметр воздуховода", 100.0)
            if resolved_duct:
                replace_if_filter_exists("Диаметр воздуховода", resolved_duct)

        # Если ничего не подхватили в оснащении — ставим безопасный дефолт.
        ru_val = next((v for n, v in result if n == "Оснащение"), "")
        ua_val = next((v for n, v in result if n == "Оснащення"), "")
        if not ua_val and ru_val:
            replace_if_filter_exists("Оснащення", ru_val)
        if not ru_val and ua_val:
            replace_if_filter_exists("Оснащение", ua_val)
        if not ru_val and not ua_val:
            replace_if_filter_exists("Оснащение", "Световой индикатор")
            replace_if_filter_exists("Оснащення", "Световой индикатор")

    def add_ball_valves_fixed_rules() -> None:
        if target_title != "Шаровые краны и вентили":
            return
        source_key = normalize_key(source_name)
        if "приборные краны" in source_key or "краны шаровые и вентили" in source_key:
            replace_if_filter_exists("Назначение", "Для водопровода")

    def add_mixer_accessories_fixed_rules() -> None:
        source_key = normalize_key(source_name)
        if "аксессуары смесители" not in source_key:
            return
        probe = text_key
        if "картридж" in probe:
            replace_if_filter_exists("Тип детали", "Картридж")
            replace_if_filter_exists("Тип деталі", "Картридж")
        elif "аэратор" in probe or "аератор" in probe:
            replace_if_filter_exists("Тип детали", "Аэратор")
            replace_if_filter_exists("Тип деталі", "Аэратор")

    def add_collectors_fixed_rules() -> None:
        if target_title not in {"Коллекторы", "Коллекторные шкафы"}:
            return
        diameter = param_value("Диаметр резьбы")
        type_key = normalize_key(param_value("Тип", "Вид"))
        title_key = normalize_key(maudau.child_text(offer, "name") or maudau.child_text(offer, "name_ru"))
        collector_probe = " ".join([type_key, title_key, text_key])

        if target_title == "Коллекторы":
            replace_if_filter_exists("Тип", "Разделительные")
            if "вентил" in collector_probe:
                replace_if_filter_exists("Вид", "С вентилями")
            elif "расходомер" in collector_probe:
                replace_if_filter_exists("Вид", "С расходомерами")
            elif "термостат" in collector_probe or "термоклап" in collector_probe:
                replace_if_filter_exists("Вид", "С термоклапанами")
            else:
                replace_if_filter_exists("Вид", "Простые")

            if "тепл" in collector_probe and "пол" in collector_probe:
                replace_if_filter_exists("Назначение", "Для теплого пола")
            elif "радиатор" in collector_probe:
                replace_if_filter_exists("Назначение", "Для радиаторов")
            elif "водоснаб" in collector_probe or "водопровод" in collector_probe:
                replace_if_filter_exists("Назначение", "Для водоснабжения")
            elif "отоплен" in collector_probe:
                replace_if_filter_exists("Назначение", "Для отопления")
            elif not has_filter("Назначение"):
                stable_seed = maudau.resolve_offer_id_key(offer) or maudau.child_text(offer, "vendorCode") or "0"
                variants = [
                    "Для водоснабжения",
                    "Для отопления",
                    "Для радиаторов",
                    "Для теплого пола",
                ]
                idx = int(hashlib.md5(stable_seed.encode("utf-8")).hexdigest(), 16) % len(variants)
                replace_if_filter_exists("Назначение", variants[idx])

            if diameter:
                resolved = choose_allowed("Диаметр", "Диаметр резьбы", diameter)
                if resolved:
                    replace_if_filter_exists("Диаметр", resolved)

        if target_title == "Коллекторные шкафы":
            title_key = normalize_key(maudau.child_text(offer, "name") or maudau.child_text(offer, "name_ru"))
            if "внутрен" in title_key:
                replace_pair("Монтаж", "Встраиваемый")
            elif "наруж" in title_key:
                replace_pair("Монтаж", "Наружный")
            if "замок" in text_key:
                replace_pair("Особенности", "Замок на двери")

    def add_hose_fixed_rules() -> None:
        if target_title != "Водопроводные шланги":
            return
        kind_key = normalize_key(param_value("Вид", "Тип"))
        conn_key = normalize_key(param_value("Тип соединения"))
        length_raw = clean_text(param_value("Длина, см", "Довжина", "Длина"))
        thread_raw = clean_text(param_value("Диаметр резьбы"))
        name_key = normalize_key(maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name"))
        hose_probe = " ".join([kind_key, name_key, text_key])

        length_num = extract_first_number(length_raw)
        if length_num is not None:
            cm = normalize_to_cm(length_num)
            if cm is not None:
                length_with_unit = f"{to_int_string(cm)} см"
                resolved_len = (
                    choose_allowed("Длина", "Длина, см", length_with_unit)
                    or choose_allowed("Довжина", "Длина, см", length_with_unit)
                    or pick_nearest_allowed_by_number("Длина", cm)
                    or pick_nearest_allowed_by_number("Довжина", cm)
                    or length_with_unit
                )
                replace_if_filter_exists("Длина", resolved_len)
                replace_if_filter_exists("Довжина", resolved_len)

        if "стираль" in kind_key:
            replace_if_filter_exists("Назначение", "Для стиральной машины")
        elif "смесител" in kind_key:
            replace_if_filter_exists("Назначение", "Для смесителя")
        elif any(token in kind_key for token in ("антивибрац", "водяной", "гофр")):
            replace_if_filter_exists("Назначение", "Для бойлера")
        elif "заливн" in hose_probe or "сливн" in hose_probe:
            replace_if_filter_exists("Назначение", "Для стиральной машины")
        elif not has_filter("Назначение"):
            replace_if_filter_exists("Назначение", "Для смесителя")

        if "внутрен" in conn_key and "наруж" in conn_key:
            replace_if_filter_exists("Тип резьбы", "Внутренняя / Наружная")
        elif "наруж" in conn_key:
            replace_if_filter_exists("Тип резьбы", "Внутренняя / Наружная")
        elif "внутрен" in conn_key:
            replace_if_filter_exists("Тип резьбы", "Внутренняя / Внутренняя")
        elif " заливн" in f" {hose_probe} " or "вв" in hose_probe:
            replace_if_filter_exists("Тип резьбы", "Внутренняя / Внутренняя")
        elif " сливн" in f" {hose_probe} ":
            replace_if_filter_exists("Тип резьбы", "Внутренняя / Внутренняя")
        elif not has_filter("Тип резьбы"):
            replace_if_filter_exists("Тип резьбы", "Внутренняя / Наружная")

        if thread_raw:
            resolved_thread = choose_allowed("Диаметр резьбы", "Диаметр резьбы", thread_raw)
            if not resolved_thread:
                thread_nums = extract_all_numbers(thread_raw)
                if thread_nums:
                    resolved_thread = pick_nearest_allowed_by_number("Диаметр резьбы", thread_nums[0])
            if not resolved_thread and "3/4" in hose_probe:
                resolved_thread = choose_allowed("Диаметр резьбы", "Диаметр резьбы", '3/4"')
            if not resolved_thread and "1/2" in hose_probe:
                resolved_thread = choose_allowed("Диаметр резьбы", "Диаметр резьбы", '1/2"')
            if resolved_thread:
                replace_if_filter_exists("Диаметр резьбы", resolved_thread)
        elif "3/4" in hose_probe:
            resolved_thread = choose_allowed("Диаметр резьбы", "Диаметр резьбы", '3/4"')
            if resolved_thread:
                replace_if_filter_exists("Диаметр резьбы", resolved_thread)
        elif "1/2" in hose_probe:
            resolved_thread = choose_allowed("Диаметр резьбы", "Диаметр резьбы", '1/2"')
            if resolved_thread:
                replace_if_filter_exists("Диаметр резьбы", resolved_thread)

        if not has_filter("Длина") and not has_filter("Довжина"):
            mid = pick_middle_allowed_value("Длина") or pick_middle_allowed_value("Довжина")
            if mid:
                replace_if_filter_exists("Длина", mid)
                replace_if_filter_exists("Довжина", mid)

    def add_fittings_water_fixed_rules() -> None:
        if target_title != "Фитинги для водопроводных труб":
            return
        type_key = normalize_key(param_value("Тип", "Вид"))
        if "вентил" in type_key:
            replace_if_filter_exists("Вид", "Вентиль")
        elif "муфта разбор" in type_key:
            replace_if_filter_exists("Вид", "Муфта")
        elif "муфт" in type_key:
            replace_if_filter_exists("Вид", "Муфта")
        elif "переходник" in type_key:
            replace_if_filter_exists("Вид", "Переходник")
        elif "заглуш" in type_key:
            replace_if_filter_exists("Вид", "Заглушка")
        elif "колено" in type_key:
            replace_if_filter_exists("Вид", "Колено")
        elif any(token in type_key for token in ("кран шаров", "кран радиатор", "кран вентиль", "кран термостат")):
            replace_if_filter_exists("Вид", "Шаровой кран")
        elif "фланец" in type_key:
            replace_if_filter_exists("Вид", "Фланец")
        elif "фильтр" in type_key:
            replace_if_filter_exists("Вид", "Фильтр")
        elif "тройник" in type_key:
            replace_if_filter_exists("Вид", "Тройник")
        elif "крестов" in type_key:
            replace_if_filter_exists("Вид", "Крестовина")
        elif "клапан" in type_key:
            replace_if_filter_exists("Вид", "Клапан")
        elif any(token in type_key for token in ("врезка", "гарпун", "планка", "пятерник", "сгон американка", "цанга", "фитинг")):
            replace_if_filter_exists("Вид", "Фитинг")

    def resolve_fittings_water_diameter_conflict() -> None:
        if target_title != "Фитинги для водопроводных труб":
            return

        diameter_values = [value for name, value in result if name == "Диаметр"]
        if len(diameter_values) <= 1:
            return

        meta = get_filter_meta_cached(target_title, target_url, "Диаметр", allo_categories)
        allowed_values = sorted(meta.get("values", {}).keys())
        allowed_keys = {normalize_key(v) for v in allowed_values}

        def is_canonical(value: str) -> bool:
            return normalize_key(value) in allowed_keys

        thread_raw = clean_text(param_value("Диаметр резьбы"))
        metric_raw = clean_text(param_value("Диаметр, мм", "Диаметр"))
        thread_candidate = choose_allowed("Диаметр", "Диаметр резьбы", thread_raw) if thread_raw else ""
        metric_candidate = choose_allowed("Диаметр", "Диаметр, мм", metric_raw) if metric_raw else ""

        chosen = ""

        # Приоритет пользователя: при конфликте берём значение из "Диаметр резьбы".
        if thread_candidate and any(normalize_key(v) == normalize_key(thread_candidate) for v in diameter_values):
            chosen = next(v for v in diameter_values if normalize_key(v) == normalize_key(thread_candidate))

        # Если только одно из конфликтующих значений каноническое — берём каноническое.
        if not chosen:
            canonical_values = [v for v in diameter_values if is_canonical(v)]
            if len(canonical_values) == 1:
                chosen = canonical_values[0]

        # Если прямого попадания из резьбы не было, но резьбовой кандидат канонический — берём его.
        if not chosen and thread_candidate and is_canonical(thread_candidate):
            chosen = thread_candidate

        # Если у резьбы неканон, а у "Диаметр, мм" канон — берём канон.
        if not chosen and metric_candidate and is_canonical(metric_candidate) and (not thread_candidate or not is_canonical(thread_candidate)):
            chosen = metric_candidate

        # Фолбэки.
        if not chosen and thread_candidate:
            chosen = thread_candidate
        if not chosen:
            chosen = next((v for v in diameter_values if is_canonical(v)), "")
        if not chosen:
            chosen = diameter_values[0]

        remove_filter("Диаметр")
        add_if_filter_exists("Диаметр", chosen)

    def dedupe_single_select_filters() -> None:
        # Убираем конфликты вида один фильтр -> несколько значений
        # только для single-select полей ALLO. Для multiple-select оставляем как есть.
        grouped: dict[str, list[str]] = defaultdict(list)
        for name, value in result:
            grouped[name].append(value)

        for filter_name, values in grouped.items():
            unique_values = []
            seen_norm: set[str] = set()
            for value in values:
                key = normalize_key(value)
                if key in seen_norm:
                    continue
                seen_norm.add(key)
                unique_values.append(value)

            if len(unique_values) <= 1:
                continue

            meta = get_filter_meta_cached(target_title, target_url, filter_name, allo_categories)
            if (meta.get("selection_mode") or "").strip().lower() == "multiple":
                continue

            allowed_values = sorted((meta.get("values") or {}).keys())
            allowed_norm = {normalize_key(v) for v in allowed_values}
            canonical_values = [v for v in unique_values if normalize_key(v) in allowed_norm]

            chosen = ""
            if len(canonical_values) == 1:
                chosen = canonical_values[0]
            elif len(canonical_values) > 1:
                # Если канонических несколько, приоритет у последнего добавленного.
                for value in reversed(unique_values):
                    if normalize_key(value) in allowed_norm:
                        chosen = value
                        break
            else:
                # Если канона нет, оставляем последнее добавленное значение.
                chosen = unique_values[-1]

            if not chosen:
                chosen = unique_values[-1]

            remove_filter(filter_name)
            add_pair(filter_name, chosen)

    def add_fittings_sewer_fixed_rules() -> None:
        if target_title != "Фитинги для канализационных труб":
            return
        source_key_local = normalize_key(source_name)
        type_key = normalize_key(param_value("Тип", "Вид"))
        dia_key = clean_text(param_value("Диаметр, мм", "Диаметр"))
        replace_if_filter_exists("Назначение", "Для ППР труб")

        if "трубы канализационные" in source_key_local:
            replace_if_filter_exists("Вид", "Труба")
        elif "заглуш" in type_key:
            replace_if_filter_exists("Вид", "Заглушка")
        elif "клапан" in type_key:
            replace_if_filter_exists("Вид", "Клапан")
        elif "колено" in type_key:
            replace_if_filter_exists("Вид", "Колено")
        elif "крестов" in type_key:
            replace_if_filter_exists("Вид", "Крестовина")
        elif "редукц" in type_key:
            replace_if_filter_exists("Вид", "Переходник")
        elif "муфт" in type_key:
            replace_if_filter_exists("Вид", "Муфта")
        elif "тройник" in type_key:
            replace_if_filter_exists("Вид", "Тройник")
        elif any(token in type_key for token in ("грибок", "патрубок", "ревизия", "розетка", "врезка", "гарпун", "планка", "пятерник", "сгон американка", "цанга")):
            replace_if_filter_exists("Вид", "Фитинг")

        dia_num = extract_first_number(dia_key)
        if dia_num is not None:
            if int(round(dia_num)) == 50:
                replace_if_filter_exists("Диаметр", "50 мм")
            elif int(round(dia_num)) == 110:
                replace_if_filter_exists("Диаметр", "110 мм")
            elif int(round(dia_num)) == 160:
                replace_if_filter_exists("Диаметр", "150 мм")
            elif int(round(dia_num)) == 200:
                replace_if_filter_exists("Диаметр", "Нет")
            else:
                replace_if_filter_exists("Диаметр", f"{to_int_string(dia_num)} мм")

    def add_pipe_fixed_rules() -> None:
        source_key_local = normalize_key(source_name)
        if target_title == "Водопроводные трубы" or "трубы водопроводные" in source_key_local:
            if not has_filter("Назначение"):
                replace_if_filter_exists("Назначение", "Для водопровода")

    def add_glass_cup_fixed_rules() -> None:
        if target_title != "Стаканы для ванной":
            return
        name_key = normalize_key(maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name"))
        replace_if_filter_exists("Назначение", "Для зубных щёток")

        if "настольн" in name_key or "отдельностоящ" in name_key:
            replace_if_filter_exists("Тип установки", "Настольный")
        else:
            replace_if_filter_exists("Тип установки", "Настенный (подвесной)")

        glasses_count = "1"
        if any(token in name_key for token in ("двойн", "2 стак", "2-стак", "на 2 стак", "на два стак")):
            glasses_count = "2"
        replace_if_filter_exists("Количество стаканов", glasses_count)
        replace_if_filter_exists("Кількість стаканів", glasses_count)
        replace_if_filter_exists("Кількість склянок", glasses_count)

    def add_handrail_fixed_rules() -> None:
        if target_title == "Поручни для ванной" or normalize_key(target_title) == "поручни":
            replace_if_filter_exists("Крепление", "На дюбель")

    def add_bath_shelf_fixed_rules() -> None:
        if target_title != "Полки для ванной комнаты":
            return
        replace_if_filter_exists("Расположение", "Подвесные")
        replace_if_filter_exists("Розташування", "Подвесные")
        replace_if_filter_exists("Тип", "Открытые")

    def add_soap_dish_fixed_rules() -> None:
        if target_title != "Мыльницы":
            return
        name_key = normalize_key(maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name"))
        replace_if_filter_exists("Способ монтажа", "Отсутствует")
        replace_if_filter_exists("Спосіб монтажу", "Отсутствует")
        if "настольн" in name_key:
            replace_if_filter_exists("Тип установки", "Настольный")
        else:
            replace_if_filter_exists("Тип установки", "Настенный (подвесной)")

    def add_hand_dryer_fixed_rules() -> None:
        if target_title != "Сушилки для рук":
            return

        replace_if_filter_exists("Тип включения", "Автоматический")
        replace_if_filter_exists("Тип включення", "Автоматический")

        body_material = normalize_key(param_value("Материал корпуса"))
        if "металл" in body_material:
            replace_if_filter_exists("Материал корпуса", "Металл")
        elif "abs" in body_material or "аbs" in body_material or "пластик" in body_material:
            replace_if_filter_exists("Материал корпуса", "Пластик")

        power_num = extract_first_number(param_value("Общая мощность", "Мощность, Вт", "Номинальная мощность, Вт", "Мощность"))
        if power_num is not None:
            power_val = ""
            if 700 <= power_num <= 1000:
                power_val = "До 1000 Вт"
            elif 1000 < power_num <= 1599:
                power_val = "1000 - 1500 Вт"
            elif 1600 <= power_num <= 2199:
                power_val = "1501 - 2000 Вт"
            elif 2200 <= power_num <= 2499:
                power_val = "2001 - 2500 Вт"
            elif power_num >= 2500:
                power_val = "Более 2500 Вт"
            if power_val:
                replace_if_filter_exists("Мощность", power_val)

    def add_urinal_fixed_rules() -> None:
        if target_title != "Писсуары":
            return
        mount_key = normalize_key(param_value("Монтаж"))
        if "наполь" in mount_key:
            replace_if_filter_exists("Тип монтажу", "Наружный")
            replace_if_filter_exists("Тип монтажа", "Наружный")
        elif "наруж" in mount_key or "под инстал" in mount_key:
            replace_if_filter_exists("Тип монтажу", "Подвесной")
            replace_if_filter_exists("Тип монтажа", "Подвесной")

    def add_shower_set_fixed_rules() -> None:
        if target_title != "Душевые гарнитуры":
            return
        mount_ru = clean_text(param_value("Тип монтажа", "Тип монтажу"))
        if mount_ru:
            replace_if_filter_exists("Тип монтажа", mount_ru)
            replace_if_filter_exists("Тип монтажу", mount_ru)
        if not has_filter("Тип монтажа"):
            replace_if_filter_exists("Тип монтажа", "Наружный")
        if not has_filter("Тип монтажу"):
            replace_if_filter_exists("Тип монтажу", "Наружный")

        kit_val = clean_text(param_value("В комплекте", "У комплекті"))
        if not kit_val:
            kit_val = "Ручная лейка , Душевой шланг , Держатель для душа"
        replace_if_filter_exists("В комплекте", kit_val)
        replace_if_filter_exists("У комплекті", kit_val)

    def add_penal_fixed_rules() -> None:
        if target_title != "Пеналы для ванной комнаты":
            return
        width_num = normalize_to_cm(extract_first_number(param_value("ширина, см", "Ширина, см", "Ширина")))
        height_num = normalize_to_cm(extract_first_number(param_value("Высота, см", "высота, см", "Высота")))
        mount_key = normalize_key(param_value("монтаж", "Монтаж"))
        name_text = clean_text(maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name"))

        if width_num is None:
            width_num = normalize_to_cm(extract_number_before_unit(name_text, "см"))
        if width_num is None:
            width_num = normalize_to_cm(extract_model_width_hint(name_text))
        if width_num is not None:
            width_with_unit = f"{to_int_string(width_num)} см"
            resolved_w = (
                choose_allowed("Ширина", "ширина, см", width_with_unit)
                or pick_nearest_allowed_by_number("Ширина", width_num)
            )
            if resolved_w:
                replace_if_filter_exists("Ширина", resolved_w)
        if not has_filter("Ширина"):
            fallback_w = pick_middle_allowed_value("Ширина")
            if fallback_w:
                replace_if_filter_exists("Ширина", fallback_w)

        if height_num is None:
            size_triplet = first_size_match(name_text)
            if size_triplet and size_triplet[2] is not None:
                height_num = normalize_to_cm(size_triplet[2])
        if height_num is not None:
            height_with_unit = f"{to_int_string(height_num)} см"
            resolved_h = (
                choose_allowed("Высота", "Высота, см", height_with_unit)
                or pick_nearest_allowed_by_number("Высота", height_num)
            )
            if resolved_h:
                replace_if_filter_exists("Высота", resolved_h)
        if not has_filter("Высота"):
            fallback_h = pick_middle_allowed_value("Высота")
            if fallback_h:
                replace_if_filter_exists("Высота", fallback_h)

        if "подвес" in mount_key or "навес" in text_key:
            replace_if_filter_exists("Монтаж", "Подвесной")
        elif "наполь" in mount_key:
            replace_if_filter_exists("Монтаж", "Напольный")
        elif not has_filter("Монтаж"):
            replace_if_filter_exists("Монтаж", "Подвесной")

    def add_dispenser_fixed_rules() -> None:
        if target_title != "Дозаторы (диспенсеры) для ванной комнаты":
            return
        replace_if_filter_exists("Назначение", "Для жидкого мыла")
        replace_if_filter_exists("Тип", "Кнопочные")

    def add_trash_bin_fixed_rules() -> None:
        if target_title != "Мусорные ведра":
            return
        source_volume = clean_text(param_value("Объем", "Об'єм"))
        if not source_volume:
            vol_match = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:л|литр)", clean_text(maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name")), flags=re.IGNORECASE)
            if vol_match:
                source_volume = f"{vol_match.group(1).replace(',', '.')} л"
        if source_volume:
            nums = extract_all_numbers(source_volume)
            resolved = choose_allowed("Объем", "Об'єм", source_volume)
            if not resolved and nums:
                resolved = pick_nearest_allowed_by_number("Объем", nums[0]) or f"{to_int_string(nums[0])} л"
            if resolved:
                replace_if_filter_exists("Объем", resolved)
                replace_if_filter_exists("Об'єм", resolved)

        replace_if_filter_exists("Тип", "Ведра")

        bin_text = normalize_key(
            " ".join(
                [
                    maudau.child_text(offer, "name"),
                    maudau.child_text(offer, "name_ru"),
                    maudau.child_text(offer, "description"),
                    maudau.child_text(offer, "description_ru"),
                ]
            )
        )
        if any(token in bin_text for token in ("нажимной на педаль", "плавным закрытием", "педаль", "педалью")):
            replace_if_filter_exists("Конструкция", "С педалью , С крышкой")
        else:
            replace_if_filter_exists("Конструкция", "С крышкой")

    def add_engineering_valve_fixed_rules() -> None:
        if normalize_key(target_title) != "комплектующая запорная арматура":
            return
        replace_if_filter_exists("Назначение", "Для водопровода")
        type_key = normalize_key(param_value("Тип"))
        if "сгон американка" in type_key:
            replace_if_filter_exists("Вид", "Фитинг")
        if not has_filter("Вид"):
            replace_if_filter_exists("Вид", "Фитинг")

    def add_radiator_accessories_fixed_rules() -> None:
        if target_title != "Комплектующие к радиаторам":
            return
        source_key = normalize_key(source_name)
        type_key = normalize_key(param_value("Тип", "Вид"))

        if "комплектующие для радиаторов" not in source_key:
            return
        if "клапан под термоголовку" in type_key:
            replace_if_filter_exists("Тип", "Запорные клапаны")
        elif "комплект кранов с термоголовкой" in type_key:
            replace_if_filter_exists("Тип", "Комплекты RTL")
        elif "комплект радиаторный" in type_key:
            replace_if_filter_exists("Тип", "Комплекты радиаторные")
        elif "кран радиаторный" in type_key:
            replace_if_filter_exists("Тип", "Краны радиаторные")
        elif "крепление" in type_key:
            replace_if_filter_exists("Тип", "Крепления для радиатора")
        elif "термостатическая головка" in type_key:
            replace_if_filter_exists("Тип", "Термостатические головки")
        elif "трубка" in type_key:
            replace_if_filter_exists("Тип", "Адаптеры")
        elif "узел нижнего подключения радиатора" in type_key:
            replace_if_filter_exists("Тип", "Узлы нижнего подключения радиатора")

    def add_generic_bathroom_accessories_fixed_rules() -> None:
        source_key = normalize_key(source_name)
        name_key = normalize_key(maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name"))

        if target_title == "Полотенцедержатели":
            replace_if_filter_exists("Тип установки", "Настенный")
        if target_title == "Держатели для туалетной бумаги":
            replace_if_filter_exists("Тип установки", "Настенная (подвесная)")
        if target_title == "Ершики и стойки":
            if "наполь" in name_key or "на пол" in name_key:
                replace_if_filter_exists("Установка", "Напольная")
            else:
                replace_if_filter_exists("Установка", "Настенная (подвесная)")
        if target_title == "Косметические зеркала":
            replace_if_filter_exists("Тип", "Универсальное")
        if target_title == "Крючки для ванной":
            if "планк" in name_key:
                replace_if_filter_exists("Тип", "Планки с крючками")
            else:
                replace_if_filter_exists("Тип", "Крючки")
            replace_if_filter_exists("Установка", "Настенная (подвесная)")
            hooks_value = "1"
            if any(token in name_key for token in ("двойн", "с 2-мя", "2 подвижных")):
                hooks_value = "2"
            elif any(token in name_key for token in ("тройн", "с тремя", "с 3-мя")):
                hooks_value = "3"
            elif any(token in name_key for token in ("четыре", "с 4-мя", "на 4", "4 крючка")):
                hooks_value = "4"
            elif any(token in name_key for token in ("на 5", "с 5-ю", "5 подвижных", "5 крючков")):
                hooks_value = "5"
            elif any(token in name_key for token in ("6 подвижных", "с 6-ю")):
                hooks_value = "6"
            replace_if_filter_exists("Количество крючков", hooks_value)
            replace_if_filter_exists("Кількість гачків", hooks_value)

    def add_hose_connections_fixed_rules() -> None:
        source_key = normalize_key(source_name)
        if target_title == "Комплектующие для душа" and "шланговые подсоединения" in source_key:
            replace_if_filter_exists("Тип детали", "Шланговое подключение")
            replace_if_filter_exists("Тип деталі", "Шланговое подключение")

    def add_kitchen_accessories_fixed_rules() -> None:
        source_key = normalize_key(source_name)
        if target_title != "Аксессуары к кухонным мойкам":
            return
        if "кухонные аксессуары кухня" not in source_key:
            return
        kind_key = normalize_key(param_value("Вид", "Тип"))
        if "моечная корзина" in kind_key:
            replace_if_filter_exists("Тип", "Корзины")
        elif "сушка" in kind_key:
            replace_if_filter_exists("Тип", "Сушилка")
        elif "коландер" in kind_key:
            replace_if_filter_exists("Тип", "Коландеры")
        elif "рабочая доска" in kind_key:
            replace_if_filter_exists("Тип", "Коврики")

    def add_shower_tray_fixed_rules() -> None:
        if target_title != "Душевые поддоны":
            return
        shape_key = normalize_key(param_value("Форма", "форма"))
        mapped_material_key = normalize_key(next((v for n, v in result if n == "Материал"), ""))
        material_key = normalize_key(param_value("Материал", "Матеріал"))
        material_probe = " ".join([material_key, mapped_material_key, text_key])
        tray_type_key = normalize_key(param_value("Поддон"))
        if "мелк" in tray_type_key:
            replace_if_filter_exists("Тип поддонов", "Мелкие")
        elif "глубок" in tray_type_key:
            replace_if_filter_exists("Тип поддонов", "Глубокие")
        elif "средн" in tray_type_key:
            replace_if_filter_exists("Тип поддонов", "Средние")
        elif not has_filter("Тип поддонов"):
            replace_if_filter_exists("Тип поддонов", "Мелкие")

        # Материал: приводим к канонам.
        if "акрил" in material_probe:
            replace_if_filter_exists("Материал", "Акрил")
        elif any(token in material_probe for token in ("искусствен", "камень", "кварил", "мрамор")):
            replace_if_filter_exists("Материал", "Искусственный камень")
        elif "сталь" in material_probe:
            replace_if_filter_exists("Материал", "Сталь")

        # Длина/ширина должны передаваться, если размер есть в источнике.
        length_cm = normalize_to_cm(
            extract_first_number(param_value("Длина, см", "Довжина, см", "Длина", "Довжина"))
        )
        width_cm = normalize_to_cm(extract_first_number(param_value("Ширина, см", "Ширина")))

        size_raw = clean_text(param_value("Размер", "Розмір"))
        footprint = extract_best_footprint_cm(size_raw, min_side=60.0, max_side=220.0)
        if footprint is None:
            footprint = extract_best_footprint_cm(text_blob, min_side=60.0, max_side=220.0)
        if footprint:
            if length_cm is None:
                length_cm = max(footprint)
            if width_cm is None:
                width_cm = min(footprint)

        # Кейсы вида "Angela-80" для квадратных поддонов.
        if (length_cm is None or width_cm is None) and "квадрат" in shape_key:
            name_text = clean_text(maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name"))
            single = re.search(r"(?:-|–|—)([6-9]\d|1\d{2}|2[0-2]\d)(?:\D|$)", name_text)
            if single:
                side = float(single.group(1))
                if length_cm is None:
                    length_cm = side
                if width_cm is None:
                    width_cm = side

        if length_cm is not None:
            length_with_unit = f"{to_int_string(length_cm)} см"
            resolved_len = (
                choose_allowed("Длина", "Длина, см", length_with_unit)
                or choose_allowed("Довжина", "Длина, см", length_with_unit)
                or choose_allowed("Длина, см", "Длина, см", length_with_unit)
                or pick_nearest_allowed_by_number("Длина", length_cm)
                or pick_nearest_allowed_by_number("Довжина", length_cm)
                or pick_nearest_allowed_by_number("Длина, см", length_cm)
            )
            if resolved_len:
                replace_if_filter_exists("Длина", resolved_len)
                replace_if_filter_exists("Довжина", resolved_len)
                replace_if_filter_exists("Длина, см", resolved_len)

        if width_cm is not None:
            width_with_unit = f"{to_int_string(width_cm)} см"
            resolved_w = (
                choose_allowed("Ширина", "Ширина, см", width_with_unit)
                or choose_allowed("Ширина, см", "Ширина, см", width_with_unit)
                or pick_nearest_allowed_by_number("Ширина", width_cm)
                or pick_nearest_allowed_by_number("Ширина, см", width_cm)
            )
            if resolved_w:
                replace_if_filter_exists("Ширина", resolved_w)
                replace_if_filter_exists("Ширина, см", resolved_w)

    def add_polotencesushiteli_fixed_rules() -> None:
        if target_title != "Полотенцесушители":
            return
        source_key = normalize_key(source_name)
        form_key = normalize_key(param_value("форма", "Форма"))
        power_key = normalize_key(param_value("Оснащение"))

        if "электрические" in source_key:
            replace_if_filter_exists("Тип", "Электрический")
        elif "водяные" in source_key:
            replace_if_filter_exists("Тип", "Водяной")

        if "змейка" in form_key:
            replace_if_filter_exists("Форма", "Змейка")
        elif "каскад" in form_key or "лесенка" in form_key:
            replace_if_filter_exists("Форма", "Лесенка")
        elif "нестандарт" in form_key:
            replace_if_filter_exists("Форма", "Дизайнерская")
        elif "поворот" in form_key:
            replace_if_filter_exists("Форма", "Дизайнерская")
        elif "с полкой" in form_key:
            replace_if_filter_exists("Форма", "Лесенка")
        elif "прямоуголь" in form_key:
            replace_if_filter_exists("Форма", "Панель")

        if "с полкой" in form_key or "с полкой" in power_key:
            add_if_filter_exists("Особенности", "С полкой")

        if "поворот" in form_key:
            replace_if_filter_exists("Тип крепления", "Поворотный")
        elif not has_filter("Тип крепления"):
            replace_if_filter_exists("Тип крепления", "Стационарный")

        if "водяные полотенцесушители" in source_key and not has_filter("Форма"):
            replace_if_filter_exists("Форма", "Дизайнерська")

    def add_radiator_fixed_rules() -> None:
        if target_title != "Радиаторы отопления":
            return
        rad_kind = normalize_key(param_value("Вид радиатора"))
        rad_type = normalize_key(param_value("Тип радиатора"))
        conn_type = normalize_key(param_value("Тип подключения"))
        conn_diam = normalize_key(param_value("Диаметр подключения"))
        material_key = normalize_key(param_value("Материал"))
        height_num = normalize_to_cm(extract_first_number(param_value("Высота, см", "высота, см")))

        if "панель" in rad_kind:
            replace_if_filter_exists("Тип радиатора", "Панельный")
        elif "секцион" in rad_kind:
            replace_if_filter_exists("Тип радиатора", "Секционный")
        elif "трубн" in rad_kind:
            replace_if_filter_exists("Тип радиатора", "Трубчатый")

        if rad_type in {"11", "12", "13", "22", "33"}:
            replace_if_filter_exists("Тип панельного радиатора", rad_type)
            replace_if_filter_exists("Тип радиатора", "Панельный")

        if "нижн" in conn_type:
            replace_if_filter_exists("Подключение", "Нижнее")
        elif "боков" in conn_type:
            replace_if_filter_exists("Подключение", "Боковое")
        elif "диагон" in conn_type or "универс" in conn_type:
            replace_if_filter_exists("Подключение", "Универсальное")
        elif conn_diam in {"1/2\"", "1\"", "1/2", "1"} and not has_filter("Подключение"):
            replace_if_filter_exists("Подключение", "Универсальное")

        if "алюмин" in material_key:
            replace_if_filter_exists("Материал", "Алюминиевые")
        elif "бимет" in material_key:
            replace_if_filter_exists("Материал", "Биметаллические")
        elif "чугун" in material_key:
            replace_if_filter_exists("Материал", "Чугунные")
        elif "стал" in material_key:
            replace_if_filter_exists("Материал", "Стальные")
        elif "алюмин" in text_key:
            replace_if_filter_exists("Материал", "Алюминиевые")
        elif "бимет" in text_key:
            replace_if_filter_exists("Материал", "Биметаллические")
        elif "чугун" in text_key:
            replace_if_filter_exists("Материал", "Чугунные")
        elif "панель" in rad_kind or rad_type in {"10", "11", "12", "13", "19", "20", "21", "22", "30", "33"}:
            replace_if_filter_exists("Материал", "Стальные")

        if height_num is not None and not has_filter("Высота"):
            resolved_h = choose_allowed("Высота", "Высота, см", f"{to_int_string(height_num)} см")
            if not resolved_h:
                resolved_h = pick_nearest_allowed_by_number("Высота", height_num)
            if resolved_h:
                replace_if_filter_exists("Высота", resolved_h)

        # Если площадь не пришла параметром, пробуем вытащить из описания.
        if not has_filter("Площадь обогрева"):
            m = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:м2|м²)", text_blob.lower())
            if m:
                val = float(m.group(1).replace(",", "."))
                if val <= 10:
                    replace_if_filter_exists("Площадь обогрева", "До 10 м²")
                elif val <= 15:
                    replace_if_filter_exists("Площадь обогрева", "10.1 - 15 м²")
                elif val <= 20:
                    replace_if_filter_exists("Площадь обогрева", "15.1 - 20 м²")
                else:
                    replace_if_filter_exists("Площадь обогрева", "Более 20 м²")
        # Если в описании нет явной площади — считаем приблизительно из размеров и типа.
        if not has_filter("Площадь обогрева"):
            h_cm = normalize_to_cm(extract_first_number(param_value("Высота, см", "высота, см")))
            w_cm = normalize_to_cm(extract_first_number(param_value("ширина, см", "Ширина, см")))
            panel_num_key = normalize_key(param_value("Тип радиатора"))
            rad_kind_key = normalize_key(param_value("Вид радиатора"))
            if h_cm is not None and w_cm is not None:
                coef = 0.000033  # тип 22 по умолчанию (как в вашем примере 500x1000 ~= 15-18 м²)
                if panel_num_key in {"11", "10", "12", "13"}:
                    coef = 0.000022
                elif panel_num_key in {"33", "30"}:
                    coef = 0.000045
                elif "панель" in rad_kind_key:
                    coef = 0.000033
                elif "секцион" in rad_kind_key:
                    coef = 0.000030
                elif "трубн" in rad_kind_key:
                    coef = 0.000036
                area = (h_cm * 10.0) * (w_cm * 10.0) * coef
                if area <= 10:
                    replace_if_filter_exists("Площадь обогрева", "До 10 м²")
                elif area <= 15:
                    replace_if_filter_exists("Площадь обогрева", "10.1 - 15 м²")
                elif area <= 20:
                    replace_if_filter_exists("Площадь обогрева", "15.1 - 20 м²")
                else:
                    replace_if_filter_exists("Площадь обогрева", "Более 20 м²")

    def add_bojlery_fixed_rules() -> None:
        if target_title not in {"Водонагреватели", "Проточные водонагреватели"}:
            return
        type_key = normalize_key(param_value("Тип водонагревателя"))
        mount_key = normalize_key(param_value("монтаж", "Монтаж", "Установка"))
        ten_type_key = normalize_key(param_value("Тип ТЭНа", "Тип ТЕНа"))
        form_key = normalize_key(param_value("форма", "Форма"))
        qty_ten = extract_first_number(param_value("Количество ТЭНов"))
        power_raw = clean_text(param_value("Мощность, кВт"))
        volume_raw = clean_text(param_value("Объем бака, л"))
        name_key = normalize_key(maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name"))
        name_text = clean_text(maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name"))

        def infer_power_kw_from_text() -> float | None:
            scan_text = clean_text(" ".join([name_text, text_blob]))
            if not scan_text:
                return None
            # 1) Точная запись из требования: "Мощность (кВт): 3,60"
            m_kw = re.search(
                r"мощност[ьі]\s*\(\s*квт\s*\)\s*[:=]\s*(\d+(?:[.,]\d+)?)",
                scan_text,
                flags=re.IGNORECASE,
            )
            if m_kw:
                return float(m_kw.group(1).replace(",", "."))
            # 2) "Номинальная мощность, Вт: 3000"
            m_w_nom = re.search(
                r"номинальн[а-яіїє]+\s+мощност[ьі][^0-9]{0,25}(\d{3,5})\s*вт",
                scan_text,
                flags=re.IGNORECASE,
            )
            if m_w_nom:
                return float(m_w_nom.group(1)) / 1000.0
            # 3) Обратный порядок: "Номинальная мощность, Вт : 3000"
            m_w_nom_rev = re.search(
                r"номинальн[а-яіїє]+\s+мощност[ьі]\s*,?\s*вт\s*[:=]\s*(\d{3,5})",
                scan_text,
                flags=re.IGNORECASE,
            )
            if m_w_nom_rev:
                return float(m_w_nom_rev.group(1)) / 1000.0
            # 4) Общие фолбэки.
            m_kw_generic = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:квт|kw)\b", scan_text, flags=re.IGNORECASE)
            if m_kw_generic:
                return float(m_kw_generic.group(1).replace(",", "."))
            m_w_generic = re.search(r"(\d{3,5})\s*(?:вт|w)\b", scan_text, flags=re.IGNORECASE)
            if m_w_generic:
                return float(m_w_generic.group(1)) / 1000.0
            return None

        if "вертик" in mount_key:
            replace_if_filter_exists("Установка", "Вертикальная")
        elif "горизонт" in mount_key:
            replace_if_filter_exists("Установка", "Горизонтальная")
        elif "универс" in mount_key:
            replace_if_filter_exists("Установка", "Вертикальная, Горизонтальная")
        elif "над мойкой" in mount_key:
            replace_if_filter_exists("Установка", "Над раковиной")
        elif "под мойкой" in mount_key:
            replace_if_filter_exists("Установка", "Под раковиной")
        elif "наполь" in mount_key:
            replace_if_filter_exists("Установка", "Напольная")

        if "сух" in ten_type_key:
            replace_if_filter_exists("Тип ТЭНа", "Скрытый (Сухой)")
        elif "мокр" in ten_type_key:
            replace_if_filter_exists("Тип ТЭНа", "Открытый (Мокрый)")
        elif "сух" in name_key:
            replace_if_filter_exists("Тип ТЭНа", "Скрытый (Сухой)")
        elif "мокр" in name_key:
            replace_if_filter_exists("Тип ТЭНа", "Открытый (Мокрый)")

        if qty_ten is not None:
            q = str(int(round(qty_ten)))
            if q in {"1", "2", "3"}:
                replace_if_filter_exists("Количество ТЭНов", q)

        if "прямоуголь" in form_key:
            replace_if_filter_exists("Форма", "Прямоугольная")
        elif "круг" in form_key:
            replace_if_filter_exists("Форма", "Цилиндрическая")
        elif "плоск" in form_key:
            replace_if_filter_exists("Форма", "Плоская (Прямоугольная)")
        elif "узк" in form_key:
            replace_if_filter_exists("Форма", "Узкая (Slim)")

        if power_raw:
            resolved_power = choose_allowed("Мощность", "Мощность, кВт", power_raw)
            if not resolved_power:
                nums = [float(x.replace(",", ".")) for x in re.findall(r"\d+(?:[.,]\d+)?", power_raw)]
                if nums:
                    avg = sum(nums) / len(nums)
                    resolved_power = pick_nearest_allowed_by_number("Мощность", avg)
            if resolved_power:
                replace_if_filter_exists("Мощность", resolved_power)
        elif not has_filter("Мощность"):
            inferred = infer_power_kw_from_text()
            if inferred is not None:
                resolved_power = pick_nearest_allowed_by_number("Мощность", inferred)
                if resolved_power:
                    replace_if_filter_exists("Мощность", resolved_power)
            else:
                # По проточным тянем мощность из описания; если не нашли,
                # и по остальным бойлерам тоже, ставим среднее каноническое значение ALLO.
                fallback_power = pick_middle_allowed_value("Мощность")
                if fallback_power:
                    replace_if_filter_exists("Мощность", fallback_power)

        if volume_raw:
            resolved_volume = choose_allowed("Объем", "Объем бака, л", volume_raw)
            if not resolved_volume:
                if "менее" in normalize_key(volume_raw):
                    resolved_volume = pick_nearest_allowed_by_number("Объем", 20.0)
                elif "более" in normalize_key(volume_raw):
                    resolved_volume = pick_nearest_allowed_by_number("Объем", 150.0)
                else:
                    nums = [float(x.replace(",", ".")) for x in re.findall(r"\d+(?:[.,]\d+)?", volume_raw)]
                    if nums:
                        avg = sum(nums) / len(nums)
                        resolved_volume = pick_nearest_allowed_by_number("Объем", avg)
            if resolved_volume:
                replace_if_filter_exists("Объем", resolved_volume)
        elif not has_filter("Объем"):
            m_l = re.search(r"(\d+(?:[.,]\d+)?)\s*л\b", name_text, flags=re.IGNORECASE)
            if m_l:
                inferred_l = float(m_l.group(1).replace(",", "."))
                resolved_volume = pick_nearest_allowed_by_number("Объем", inferred_l)
                if resolved_volume:
                    replace_if_filter_exists("Объем", resolved_volume)
            else:
                m_model_l = re.search(r"\b0?(\d{2,3})-\d\b", name_text)
                if m_model_l:
                    inferred_l = float(m_model_l.group(1))
                    if 10 <= inferred_l <= 300:
                        resolved_volume = pick_nearest_allowed_by_number("Объем", inferred_l)
                        if resolved_volume:
                            replace_if_filter_exists("Объем", resolved_volume)
                elif target_title == "Водонагреватели":
                    # Последний фолбэк для единичных карточек без объёма в исходнике.
                    replace_if_filter_exists("Объем", "150 л")

        # В шаблонах ALLO встречаются оба столбца: Оснащение и Оснащення.
        # Заполняем оба, чтобы не было обязательных пустот.
        if "душ" in name_key:
            replace_if_filter_exists("Оснащение", "Душ")
            replace_if_filter_exists("Оснащення", "Душ")
        elif "кран" in name_key:
            replace_if_filter_exists("Оснащение", "Кран")
            replace_if_filter_exists("Оснащення", "Кран")
        elif "смесител" in name_key:
            replace_if_filter_exists("Оснащение", "Смеситель")
            replace_if_filter_exists("Оснащення", "Смеситель")
        elif target_title == "Проточные водонагреватели" or "проточ" in type_key:
            replace_if_filter_exists("Оснащение", "Кран")
            replace_if_filter_exists("Оснащення", "Кран")
        else:
            replace_if_filter_exists("Оснащение", "Нет")
            replace_if_filter_exists("Оснащення", "Нет")

        # Фолбэки для редких пустых карточек.
        if not has_filter("Установка"):
            replace_if_filter_exists("Установка", "Вертикальная")
        if not has_filter("Форма"):
            replace_if_filter_exists("Форма", "Цилиндрическая")
        if not has_filter("Тип ТЭНа") and target_title == "Водонагреватели":
            replace_if_filter_exists("Тип ТЭНа", "Открытый (Мокрый)")

    def add_shower_accessories_fixed_rules() -> None:
        source_key = normalize_key(source_name)
        forma_leiki = normalize_key(param_value("Форма лейки"))

        if target_title == "Душевые лейки":
            if "ручные души" in source_key:
                replace_if_filter_exists("Вид лейки", "Ручные")
                if "трубк" in forma_leiki:
                    replace_if_filter_exists("Форма", "Карандаш")
            elif "верхние души" in source_key:
                replace_if_filter_exists("Вид лейки", "Верхний душ")
            if "нестандарт" in forma_leiki and not has_filter("Форма"):
                replace_if_filter_exists("Форма", "Дизайнерская")
            if "ручные души" in source_key and not has_filter("Форма"):
                replace_if_filter_exists("Форма", "Круглая")

        if target_title == "Душевые штанги и держатели":
            if "держатели для душа" in source_key:
                replace_if_filter_exists("Тип", "Держатель душа")
            elif "кронштейны для верхнего душа" in source_key:
                replace_if_filter_exists("Тип", "Душевой кронштейн")

        if target_title == "Душевые шланги":
            replace_if_filter_exists("Диаметр подключения", '1/2"')
            length_raw = clean_text(param_value("Длинна шланга, мм", "Длина шланга, мм"))
            length_num = extract_first_number(length_raw)
            if length_num is not None:
                cm = normalize_to_cm(length_num)
                if cm is not None:
                    length_val = f"{to_int_string(cm)} см"
                    replace_if_filter_exists("Длина", length_val)
                    replace_if_filter_exists("Довжина", length_val)
            elif "растяг" in normalize_key(length_raw):
                replace_if_filter_exists("Длина", "150 см")
                replace_if_filter_exists("Довжина", "150 см")

    def add_global_color_fixed_rules() -> None:
        # Явное правило из задачи: "Цветной" -> "Colorless".
        source_color = normalize_key(param_value("Цвет", "Колір"))
        if source_color in {"цветной", "кольоровий"}:
            replace_if_filter_exists("Цвет", "Colorless")
        for filter_name, current_value in list(result):
            if filter_name != "Цвет":
                continue
            if normalize_key(current_value) in {"цветной", "кольоровий"}:
                replace_pair("Цвет", "Colorless")

    def add_global_country_guarantee_rules() -> None:
        # Строго по правилу:
        # исходник "Страна регистрации бренда" -> ALLO "Страна производитель"
        # исходник "Гарантийный срок, мес." -> ALLO "Гарантия" (канон из шаблонов)
        country_src = clean_text(param_value("Страна регистрации бренда", "Країна реєстрації бренду"))
        if country_src:
            resolved_country = choose_allowed("Страна производитель", "Страна регистрации бренда", country_src)
            if not resolved_country:
                resolved_country = country_src
            replace_pair("Страна производитель", resolved_country)
            remove_filter("Страна регистрации бренда")

        guarantee_src = clean_text(param_value("Гарантийный срок, мес.", "Гарантийный срок, мес", "Гарантийный срок, мес "))
        if guarantee_src:
            candidates: list[str] = [guarantee_src]
            months_num = extract_first_number(guarantee_src)
            if months_num is not None:
                months = int(round(months_num))
                candidates.extend([f"{months} мес.", f"{months} мес", str(months)])
                if months > 0 and months % 12 == 0:
                    years = months // 12
                    candidates.extend(
                        [
                            f"{years} год",
                            f"{years} года",
                            f"{years} лет",
                            f"{years} рік",
                            f"{years} роки",
                            f"{years} років",
                        ]
                    )

            resolved_guarantee = ""
            for candidate in candidates:
                resolved_guarantee = choose_allowed("Гарантия", "Гарантийный срок, мес.", candidate)
                if resolved_guarantee:
                    break
            if not resolved_guarantee and months_num is not None:
                resolved_guarantee = (
                    pick_nearest_allowed_by_number("Гарантия", float(months_num))
                    or pick_nearest_allowed_by_number("Гарантия", float(months_num) / 12.0)
                )
            if not resolved_guarantee and months_num is not None:
                months = int(round(months_num))
                if months > 0 and months % 12 == 0:
                    years = months // 12
                    if years == 1:
                        resolved_guarantee = "1 год"
                    elif 2 <= years <= 4:
                        resolved_guarantee = f"{years} года"
                    else:
                        resolved_guarantee = f"{years} лет"
                else:
                    resolved_guarantee = f"{months} мес."
            if not resolved_guarantee and guarantee_src:
                resolved_guarantee = guarantee_src
            if resolved_guarantee:
                replace_pair("Гарантия", resolved_guarantee)

    def add_mirror_fixed_rules() -> None:
        if target_title != "Зеркала для ванной комнаты":
            return
        shape_key = normalize_key(param_value("форма", "Форма"))
        kit_key = normalize_key(param_value("Комплектация"))
        mirror_text = " ".join([shape_key, kit_key, text_key])

        replace_if_filter_exists("Тип", "Настенные")

        if "прямоуголь" in shape_key:
            replace_if_filter_exists("Форма", "Прямоугольная")
        elif "фигур" in shape_key:
            replace_if_filter_exists("Форма", "Фигурная")
        elif "квадрат" in shape_key:
            replace_if_filter_exists("Форма", "Квадратная")
        elif "овал" in shape_key:
            replace_if_filter_exists("Форма", "Овальная")
        elif "круг" in shape_key:
            replace_if_filter_exists("Форма", "Круглая")
        elif "прямоуголь" in mirror_text:
            replace_if_filter_exists("Форма", "Прямоугольная")
        elif "фигур" in mirror_text:
            replace_if_filter_exists("Форма", "Фигурная")
        elif "квадрат" in mirror_text:
            replace_if_filter_exists("Форма", "Квадратная")
        elif "овал" in mirror_text:
            replace_if_filter_exists("Форма", "Овальная")
        elif "круг" in mirror_text:
            replace_if_filter_exists("Форма", "Круглая")

        if "подогрев" in mirror_text:
            add_if_filter_exists("Оснащение", "Подогрев")
            add_if_filter_exists("Оснащення", "Подогрев")
        if "подсвет" in mirror_text:
            add_if_filter_exists("Оснащение", "Подсветка")
            add_if_filter_exists("Оснащення", "Подсветка")
        if "полка" in mirror_text:
            add_if_filter_exists("Оснащение", "Полка")
            add_if_filter_exists("Оснащення", "Полка")
        if any(token in text_key for token in ("touch", "диммер", "сенсорный", "сенсорн", "тип выключателя сенсор")):
            add_if_filter_exists("Оснащение", "Сенсорный выключатель")
            add_if_filter_exists("Оснащення", "Сенсорный выключатель")

        if not has_filter("Форма"):
            replace_if_filter_exists("Форма", "Прямоугольная")
        if not has_filter("Оснащение") and not has_filter("Оснащення"):
            replace_if_filter_exists("Оснащение", "Крючки")
            replace_if_filter_exists("Оснащення", "Крючки")

    def add_toilet_fixed_rules() -> None:
        if target_title != "Унитазы":
            return
        source_key = normalize_key(source_name)
        toilet_type_key = normalize_key(param_value("Тип унитаза"))

        if "подвесные" in source_key:
            replace_if_filter_exists("Тип установки", "Подвесной")
        else:
            replace_if_filter_exists("Тип установки", "Напольный")

        if "компак" in source_key:
            replace_if_filter_exists("Вид", "Компакт")
        elif "унитаз биде" in toilet_type_key or "унитаз-биде" in toilet_type_key:
            replace_if_filter_exists("Вид", "Унитаз-биде")
        else:
            replace_if_filter_exists("Вид", "Унитаз")

    def add_toilet_armature_fixed_rules() -> None:
        if target_title != "Арматура для унитазов":
            return
        source_type = normalize_key(param_value("Тип", "Вид"))
        if "комплект арматуры для бачка" in source_type:
            replace_if_filter_exists("Тип", "Впускной и смывной механизмы")
        elif "впускной клапан" in source_type:
            replace_if_filter_exists("Тип", "Впускной механизм")
        elif "сливной клапан" in source_type:
            replace_if_filter_exists("Тип", "Сливные механизмы")

    def add_installation_fixed_rules() -> None:
        if target_title != "Инсталляции":
            return
        source_key = normalize_key(source_name)
        if "инсталяц" in source_key or "инсталляц" in source_key or "унитаз" in source_key:
            replace_if_filter_exists("Тип", "Для унитаза")
            replace_if_filter_exists("Крепление", "С каркасом")

    def add_tumba_fixed_rules() -> None:
        if target_title != "Тумбы для ванной комнаты":
            return
        source_key = normalize_key(source_name)
        mount_key = normalize_key(param_value("монтаж", "Монтаж", "Установка"))
        width = normalize_to_cm(extract_first_number(param_value("ширина, см")))
        depth = normalize_to_cm(extract_first_number(param_value("глубина, см")))
        height = normalize_to_cm(extract_first_number(param_value("Высота, см", "высота, см")))
        name_text = clean_text(maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name"))
        if (width is None or depth is None or height is None) and first_size_match(text_blob):
            size = first_size_match(text_blob)
            if width is None:
                width = normalize_to_cm(size[0])
            if depth is None:
                depth = normalize_to_cm(size[1])
            if height is None and size[2] is not None:
                height = normalize_to_cm(size[2])
        if width is None:
            width = normalize_to_cm(extract_number_before_unit(name_text, "см"))
        if width is None:
            width = normalize_to_cm(extract_model_width_hint(name_text))

        if "тумба с умывальником" in source_key:
            replace_if_filter_exists("Вид", "С умывальником")
        elif "тумбы со столешницей" in source_key or "тумбы под умывальник" in source_key:
            replace_if_filter_exists("Вид", "Без умывальника")

        if "подвес" in mount_key or "универс" in mount_key:
            replace_if_filter_exists("Установка", "Подвесная")
            replace_if_filter_exists("Установлення", "Подвесная")
        elif "наполь" in mount_key:
            replace_if_filter_exists("Установка", "Напольная")
            replace_if_filter_exists("Установлення", "Напольная")
        elif "подвес" in text_key:
            replace_if_filter_exists("Установка", "Подвесная")
            replace_if_filter_exists("Установлення", "Подвесная")
        elif "наполь" in text_key:
            replace_if_filter_exists("Установка", "Напольная")
            replace_if_filter_exists("Установлення", "Напольная")
        if not has_filter("Установка"):
            replace_if_filter_exists("Установка", "Подвесная")
        if not has_filter("Установлення"):
            replace_if_filter_exists("Установлення", "Подвесная")

        if width is not None:
            resolved = choose_allowed("Ширина", "ширина, см", to_int_string(width))
            if resolved:
                replace_if_filter_exists("Ширина", resolved)
        if depth is not None:
            resolved = choose_allowed("Глубина", "глубина, см", to_int_string(depth))
            if resolved:
                replace_if_filter_exists("Глубина", resolved)
        if height is not None:
            resolved = choose_allowed("Высота", "высота, см", to_int_string(height))
            if resolved:
                replace_if_filter_exists("Высота", resolved)
        if not has_filter("Высота"):
            replace_if_filter_exists("Высота", "51 - 100 см")

    def add_pump_fixed_rules() -> None:
        if target_title != "Насосы для воды":
            return
        pump_type = normalize_key(param_value("Тип"))
        pump_name = normalize_key(maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name"))
        pump_params_blob = " ".join(f"{k} {v}" for k, v in params.items())
        pump_probe = " ".join([pump_type, pump_name, text_key, normalize_key(pump_params_blob)])

        def assign_flow_lh(flow_lh: float) -> None:
            if flow_lh <= 0:
                return
            resolved = pick_nearest_allowed_by_number("Производительность", flow_lh)
            if resolved:
                replace_if_filter_exists("Производительность", resolved)
                return
            if flow_lh <= 1000:
                replace_if_filter_exists("Производительность", "До 1000 л/час")
            elif flow_lh <= 5000:
                replace_if_filter_exists("Производительность", "1001 - 5000 л/час")
            elif flow_lh <= 10000:
                replace_if_filter_exists("Производительность", "5001 - 10000 л/час")
            elif flow_lh <= 15000:
                replace_if_filter_exists("Производительность", "10001 - 15000 л/час")
            elif flow_lh <= 20000:
                replace_if_filter_exists("Производительность", "15001 - 20000 л/час")
            elif flow_lh <= 25000:
                replace_if_filter_exists("Производительность", "20001 - 25000 л/час")
            elif flow_lh <= 30000:
                replace_if_filter_exists("Производительность", "25001 - 30000 л/час")
            elif flow_lh <= 35000:
                replace_if_filter_exists("Производительность", "30001 - 35000 л/час")
            elif flow_lh <= 40000:
                replace_if_filter_exists("Производительность", "35001 - 40000 л/час")
            else:
                replace_if_filter_exists("Производительность", "Более 40000 л/час")

        if "дренажно фекаль" in pump_type:
            replace_if_filter_exists("Вид", "Дренажно-фекальный")
            replace_if_filter_exists("Тип установки", "Поверхностный")
            replace_if_filter_exists("Назначение", "Дренажный")
        elif "дренажн" in pump_type:
            replace_if_filter_exists("Вид", "Дренажный")
            replace_if_filter_exists("Тип установки", "Поверхностный")
            replace_if_filter_exists("Назначение", "Дренажный")
        elif "канализационные установки sololift" in pump_type:
            replace_if_filter_exists("Назначение", "Для унитаза")
            replace_if_filter_exists("Тип установки", "Поверхностный")
            replace_if_filter_exists("Вид", "Дренажно-фекальный")
        elif "насосные станции" in pump_type:
            replace_if_filter_exists("Назначение", "Станция водоснабжения")
            replace_if_filter_exists("Тип установки", "Поверхностный")
            replace_if_filter_exists("Вид", "Центробежный")
        elif "поверхностные насосы" in pump_type:
            replace_if_filter_exists("Тип установки", "Поверхностный")
            replace_if_filter_exists("Вид", "Центробежный")
            replace_if_filter_exists("Назначение", "Для откачки воды")
        elif "скважинные насосы" in pump_type:
            replace_if_filter_exists("Тип установки", "Погружной")
            replace_if_filter_exists("Назначение", "Для скважины")
            replace_if_filter_exists("Вид", "Центробежный")
        elif "фекальные насосы" in pump_type:
            replace_if_filter_exists("Вид", "Фекальный")
            replace_if_filter_exists("Тип установки", "Поверхностный")
            replace_if_filter_exists("Назначение", "Дренажный")
        elif "циркуляционные насосы" in pump_type:
            replace_if_filter_exists("Вид", "Циркуляционный")
            replace_if_filter_exists("Тип установки", "Поверхностный")
            replace_if_filter_exists("Назначение", "Для централизованного отопления")
        elif "автоматика" in pump_type or "комплектующие" in pump_type:
            replace_if_filter_exists("Тип установки", "Поверхностный")
            replace_if_filter_exists("Вид", "Центробежный")
            replace_if_filter_exists("Назначение", "Для повышения давления")
        elif not has_filter("Тип установки"):
            replace_if_filter_exists("Тип установки", "Поверхностный")

        if not has_filter("Вид"):
            if "дренажно" in pump_probe and "фекал" in pump_probe:
                replace_if_filter_exists("Вид", "Дренажно-фекальный")
            elif "фекал" in pump_probe:
                replace_if_filter_exists("Вид", "Фекальный")
            elif "дренаж" in pump_probe:
                replace_if_filter_exists("Вид", "Дренажный")
            elif "вихрев" in pump_probe:
                replace_if_filter_exists("Вид", "Вихревой")
            elif "вибрац" in pump_probe:
                replace_if_filter_exists("Вид", "Вибрационный")
            elif "шнек" in pump_probe:
                replace_if_filter_exists("Вид", "Шнековый")
            elif "циркуляц" in pump_probe:
                replace_if_filter_exists("Вид", "Циркуляционный")
            elif "центробеж" in pump_probe and "самовсас" in pump_probe:
                replace_if_filter_exists("Вид", "Центробежный самовсасывающий")
            elif "центробеж" in pump_probe and "многоступ" in pump_probe:
                replace_if_filter_exists("Вид", "Центробежный многоступенчатый")
            elif "центробеж" in pump_probe and "эжектор" in pump_probe:
                replace_if_filter_exists("Вид", "Центробежный с внешним эжектором")
            elif "центробеж" in pump_probe:
                replace_if_filter_exists("Вид", "Центробежный")

        if not has_filter("Назначение"):
            if "унитаз" in pump_probe or "канализац" in pump_probe or "sololift" in pump_probe:
                replace_if_filter_exists("Назначение", "Для унитаза")
            elif "циркуляц" in pump_probe or "отоплен" in pump_probe:
                replace_if_filter_exists("Назначение", "Для систем отопления")
            elif "скважин" in pump_probe:
                replace_if_filter_exists("Назначение", "Для скважины")
            elif "колод" in pump_probe:
                replace_if_filter_exists("Назначение", "Для колодца")
            elif "бассейн" in pump_probe:
                replace_if_filter_exists("Назначение", "Для бассейнов")
            elif "фонтан" in pump_probe:
                replace_if_filter_exists("Назначение", "Для фонтанов")
            elif "полив" in pump_probe:
                replace_if_filter_exists("Назначение", "Для полива огорода")
            elif "дренаж" in pump_probe or "фекал" in pump_probe:
                replace_if_filter_exists("Назначение", "Для откачки воды")
            elif "станц" in pump_probe:
                replace_if_filter_exists("Назначение", "Станция водоснабжения")
            elif "давлен" in pump_probe:
                replace_if_filter_exists("Назначение", "Для повышения давления")
            elif not has_filter("Назначение"):
                replace_if_filter_exists("Назначение", "Для откачки воды")

        if not has_filter("Производительность"):
            parsed_flow_lh = extract_pump_flow_lh(params, f"{text_blob} {pump_params_blob}")
            if parsed_flow_lh is not None:
                assign_flow_lh(parsed_flow_lh)
        if not has_filter("Производительность"):
            power_w = extract_power_w_from_text(params, f"{text_blob} {pump_params_blob}")
            if power_w is not None and PUMP_POWER_FLOW_PROFILES:
                best_flow = None
                best_diff = float("inf")
                for ref_power, ref_flow in PUMP_POWER_FLOW_PROFILES:
                    diff = abs(ref_power - power_w)
                    if diff < best_diff:
                        best_diff = diff
                        best_flow = ref_flow
                if best_flow is not None:
                    assign_flow_lh(best_flow)
        if not has_filter("Производительность"):
            fallback = pick_middle_allowed_value("Производительность")
            if fallback:
                replace_if_filter_exists("Производительность", fallback)

    def add_mixer_fixed_rules() -> None:
        if target_title != "Смесители":
            return
        source_key = normalize_key(source_name)
        purpose_key = normalize_key(param_value("Назначение", "Призначення"))
        features_key = normalize_key(param_value("Особенности", "Особливості"))
        control_key = normalize_key(param_value("Управление", "Управління"))
        mount_key = normalize_key(param_value("Монтаж", "Установка"))

        two_water = any(
            token in features_key
            for token in ("на две воды", "на дві води", "на две воды выдвижной излив", "на две воды гибкий излив")
        )

        if "комплекты смесителей" in source_key:
            add_pair("Возможное применение", "Для ванной")
            add_pair("Дополнительная комплектация", "Набор смесителей")
        elif "смесители для умывальника" in source_key:
            if "для накладных раковин" in purpose_key:
                add_pair("Возможное применение", "Для накладных раковин")
            else:
                add_pair("Возможное применение", "Для раковины / умывальника")
        elif "смесители для ванны" in source_key:
            add_pair("Возможное применение", "Для ванной")
        elif "смесители для душа" in source_key:
            add_pair("Возможное применение", "Для душа")
        elif "смесители для кухни" in source_key or "кухонные смесители" in source_key:
            add_pair("Возможное применение", "Для фильтра" if two_water else "Для кухни")
        elif "смесители для биде" in source_key:
            add_pair("Возможное применение", "Для биде")
        elif "гигиенические души" in source_key:
            add_pair("Возможное применение", "Гигиенический душ")
        elif "монокраны" in source_key:
            add_pair("Возможное применение", "Для кухни")
            replace_pair("Тип", "Кран")
        elif "скрытый монтаж" in source_key:
            replace_pair("Вид монтажа", "Скрытый")
            add_pair("Возможное применение", "Для ванной")

        if two_water:
            add_pair("Возможное применение", "Для фильтра")
        if "выдвижной излив" in features_key:
            replace_pair("Тип излива", "Выдвижной")
        if "гибкий излив" in features_key:
            replace_pair("Тип излива", "Гибкий")
        if features_key == "излив":
            replace_pair("Тип излива", "Длинный")
        if "каскад" in features_key:
            replace_pair("Тип излива", "Каскадный")
        if "термостат" in features_key or "центральный термостат" in purpose_key or "термостат с переключателем" in purpose_key:
            replace_pair("Вид смесителя", "Термостатический")
        if "без смесителя" in features_key:
            replace_pair("Вид смесителя", "Без смесителя")
        if "электронный" in features_key or "сенсор" in control_key:
            replace_pair("Вид смесителя", "Электронный (сенсорный)")
        if "кран дозатор" in features_key or "кран-дозатор" in features_key or "кнопочный" in control_key or "порционный" in control_key:
            replace_pair("Вид смесителя", "Нажимной")
        if "с бойлером" in features_key:
            replace_pair("Тип", "Кран с электронагревом")

        if "рычаг" in control_key or "рукоятка select" in control_key or "локтевой" in control_key:
            replace_pair("Вид смесителя", "Однорычажный")
        elif "джойстик" in control_key:
            replace_pair("Вид смесителя", "Джойстик")
        elif "вентиль" in control_key:
            replace_pair("Вид смесителя", "Одновентильный" if "монокраны" in source_key else "Двухвентильный")

        if "для накладных раковин" in purpose_key:
            add_pair("Возможное применение", "Для накладных раковин")
        if "излив для ванны" in purpose_key:
            replace_pair("Тип", "Излив")
        if "скрытая часть" in purpose_key:
            add_pair("Дополнительная комплектация", "Скрытая часть смесителя")
        if "подвижный корпус" in purpose_key:
            replace_pair("Тип излива", "Поворотный")
        if "с душевым гарнитуром" in purpose_key or "с душевой лейкой" in purpose_key:
            add_pair("Дополнительная комплектация", "Лейка")
        if "гигиеническая лейка" in purpose_key or "гигиенический гарнитур" in purpose_key:
            add_pair("Дополнительная комплектация", "Гигиенический душ")

        if "наружный" in mount_key:
            replace_pair("Вид монтажа", "Наружный")
        elif "напольный" in mount_key:
            replace_pair("Вид монтажа", "Напольный")
        elif "настенный" in mount_key:
            replace_pair("Вид монтажа", "Настенный")
        elif any(token in mount_key for token in ("на одно отверстие", "на два отверстия", "на три отверстия", "на борт ванны")):
            replace_pair("Вид монтажа", "Врезной")
        elif "скрытый" in mount_key:
            replace_pair("Вид монтажа", "Скрытый")
        elif "приставка для унитаза" in mount_key:
            replace_pair("Вид монтажа", "Горизонтальный")

        material_key = normalize_key(" ".join([param_value("Материал", "Матеріал"), text_blob]))
        if any(token in material_key for token in ("abs пластик", "пластик", "силумин", "алюмини")):
            replace_pair("Материал", "Пластик")
        elif "нержаве" in material_key:
            replace_pair("Материал", "Нержавеющая сталь")
        elif "латун" in material_key:
            replace_pair("Материал", "Латунь")
        elif not has_filter("Материал"):
            replace_pair("Материал", "Латунь")

        if not has_filter("Тип"):
            if any(token in source_key for token in ("смесители для душа", "гигиенические души", "скрытый монтаж")):
                replace_pair("Тип", "Без излива")
            else:
                replace_pair("Тип", "Излив")
        if not has_filter("Цвет"):
            replace_if_filter_exists("Цвет", "Хром")

    for source_param, source_value in params.items():
        if normalize_key(source_param) in allo_map.IGNORED_SOURCE_PARAM_KEYS:
            continue
        source_param_key = normalize_key(source_param)
        if target_title == "Душевые системы" and source_param_key in {
            "управление",
            "управління",
            "форма лейки",
            "тип струи",
            "тип струменя",
            "особенности",
            "особливості",
        }:
            continue
        if target_title in {"Душевые двери", "Душевые стенки", "Душевые кабины", "Душевые боксы", "Боксы гидромассажные", "Шторки для ванной", "Шторки для ванн"} and source_param_key in {
            "цвет стекла",
            "тон стекла",
            "тон скла",
            "цвет профиля",
            "цвет",
            "колір",
        }:
            continue
        filter_name, explicit_rule = pick_param_name(
            source_id,
            source_name,
            target_title,
            target_url,
            source_param,
            filter_names,
            param_rules,
        )
        if not filter_name:
            continue
        if filter_name not in filter_names:
            continue
        meta = get_filter_meta_cached(target_title, target_url, filter_name, allo_categories)
        allowed_values = sorted(meta.get("values", {}).keys())
        target_value = pick_param_value(
            source_id,
            source_name,
            target_title,
            source_param,
            filter_name,
            source_value,
            allowed_values,
            value_rules,
        )
        if not target_value:
            if meta.get("required"):
                issues.append(f"Не сопоставлено обязательное значение: {filter_name} <= {source_param}={source_value}")
            continue
        add_pair(filter_name, target_value)

    # Жёсткие правила уровня значения могут добавлять параметр, даже если исходный
    # параметр в основной категории был не тем ALLO-фильтром.
    for source_param, source_value in params.items():
        value_rule = allo_map.resolve_value_level_category_rule("", target_title, source_param, source_value)
        if value_rule and value_rule["target_title"] == target_title:
            add_pair(value_rule["filter"], value_rule["value"])

        pump_value_rule = allo_map.resolve_pump_value_rule(source_param, source_value)
        if pump_value_rule and category_rule_applies(pump_value_rule.get("category", ""), target_title):
            for filter_name, target_value in pump_value_rule.get("assignments", []):
                add_pair(filter_name, target_value)

        pump_param_rule = allo_map.resolve_pump_param_rule(source_param)
        if (
            pump_param_rule
            and not pump_param_rule.get("skip")
            and category_rule_applies(pump_param_rule.get("category", ""), target_title)
        ):
            add_pair(pump_param_rule.get("filter", ""), format_pump_param_value(source_value, pump_param_rule))

    for rule in matching_section_rules(source_name):
        if rule.skip_feed:
            continue
        for filter_name, target_value in rule.assignments:
            add_pair(filter_name, target_value)

    add_mixer_fixed_rules()
    add_shower_system_fixed_rules()
    add_kitchen_sink_fixed_rules()
    add_shower_cabin_fixed_rules()
    add_shower_box_fixed_rules()
    add_shower_glass_group_fixed_rules()
    add_shower_tray_fixed_rules()
    add_umyvalniki_fixed_rules()
    add_bath_fixed_rules()
    add_bath_panel_fixed_rules()
    add_siphon_fixed_rules()
    add_collectors_fixed_rules()
    add_hose_fixed_rules()
    add_fittings_water_fixed_rules()
    add_fittings_sewer_fixed_rules()
    add_pipe_fixed_rules()
    add_urinal_fixed_rules()
    add_shower_set_fixed_rules()
    add_dispenser_fixed_rules()
    add_penal_fixed_rules()
    add_trash_bin_fixed_rules()
    add_engineering_valve_fixed_rules()
    add_radiator_accessories_fixed_rules()
    add_generic_bathroom_accessories_fixed_rules()
    add_glass_cup_fixed_rules()
    add_handrail_fixed_rules()
    add_bath_shelf_fixed_rules()
    add_soap_dish_fixed_rules()
    add_hand_dryer_fixed_rules()
    add_hose_connections_fixed_rules()
    add_kitchen_accessories_fixed_rules()
    add_polotencesushiteli_fixed_rules()
    add_radiator_fixed_rules()
    add_mirror_fixed_rules()
    add_toilet_fixed_rules()
    add_toilet_armature_fixed_rules()
    add_installation_fixed_rules()
    add_tumba_fixed_rules()
    add_pump_fixed_rules()
    add_bojlery_fixed_rules()
    add_shower_accessories_fixed_rules()
    add_obogrevateli_fixed_rules()
    add_installation_buttons_fixed_rules()
    add_ventilyatory_fixed_rules()
    add_ball_valves_fixed_rules()
    add_mixer_accessories_fixed_rules()
    add_global_color_fixed_rules()
    add_global_country_guarantee_rules()
    apply_section_rules_strict()
    # После книжных правил повторно применяем условия, где в книге записана
    # трактовка, а не буквальное значение (например "Унитаз / Унитаз-биде по Тип унитаза").
    add_toilet_fixed_rules()
    # Старые правила из книги могут содержать "грязные" значения материала
    # (например, "Сталь покрытая порошковой краской"). Канон по разделам ванн
    # должен быть финальным: Сталь / Акрил / Чугун / Искусственный камень.
    add_bath_fixed_rules()
    resolve_fittings_water_diameter_conflict()
    dedupe_single_select_filters()

    # Синхронизация дублей RU/UA, чтобы не оставлять обязательные поля пустыми.
    mirror_pairs = [
        ("Тип монтажа", "Тип монтажу"),
        ("В комплекте", "У комплекті"),
        ("Оснащение", "Оснащення"),
        ("Объем", "Об'єм"),
        ("Установка", "Установлення"),
    ]
    for left, right in mirror_pairs:
        left_val = next((v for n, v in result if n == left), "")
        right_val = next((v for n, v in result if n == right), "")
        if left_val and not right_val:
            replace_if_filter_exists(right, left_val)
        if right_val and not left_val:
            replace_if_filter_exists(left, right_val)

    # Финальный канонический срез: оставляем только фильтры из ALLO-категории.
    canonical_result: list[tuple[str, str]] = []
    canonical_seen: set[tuple[str, str]] = set()
    force_pass_filters = {"Страна производитель", "Гарантия"}
    for name, value in result:
        name = allo_map.canonical_filter_name(name)
        if name not in filter_names and name not in force_pass_filters:
            continue
        key = (name, value)
        if key in canonical_seen:
            continue
        canonical_seen.add(key)
        canonical_result.append(key)
    result = canonical_result

    # Если обязательный фильтр в итоге заполнен фикс-правилом, убираем
    # промежуточные сообщения "не сопоставлено значение" по этому фильтру.
    filtered_issues: list[str] = []
    for issue in issues:
        if issue.startswith("Не сопоставлено обязательное значение:") and "<=" in issue:
            filter_name = allo_map.canonical_filter_name(issue.split(":", 1)[1].split("<=", 1)[0].strip())
            if has_filter(filter_name):
                continue
        filtered_issues.append(issue)
    issues = filtered_issues

    for filter_name in filter_names:
        meta = get_filter_meta_cached(target_title, target_url, filter_name, allo_categories)
        if target_title == "Душевые стенки" and filter_name in {"Довжина", "Длина"}:
            # По шаблону ALLO для душевых стенок обязательна только "Ширина",
            # а "Длина/Довжина" — необязательная.
            continue
        if meta.get("required") and not any(name == filter_name for name, _value in result):
            issues.append(f"Не заполнен обязательный параметр ALLO: {filter_name}")
    return result, issues, color_fallback_used, color_fallback_changes


def offer_to_allo_item(
    offer: ET._Element,
    category_id: str,
    source_name: str,
    target_title: str,
    target_url: str,
    allo_categories: dict[str, dict],
    param_rules: dict,
    value_rules: dict,
) -> tuple[ET._Element | None, list[str], bool, list[tuple[str, str, str]]]:
    issues: list[str] = []
    color_fallback_used = False
    color_fallback_changes: list[tuple[str, str, str]] = []
    source_id = maudau.child_text(offer, "categoryId")

    price = maudau.child_text(offer, "price")
    if not price:
        return None, ["Нет цены"], color_fallback_used, color_fallback_changes

    maudau.normalize_name_description(offer)
    maudau.normalize_old_price(offer)
    maudau.enrich_vendor_country_from_params(offer)
    maudau.cleanup_pictures(offer)
    if not maudau.normalize_offer_id(offer):
        return None, ["Нет offer id"], color_fallback_used, color_fallback_changes

    item = ET.Element("item")
    offer_id = maudau.resolve_offer_id_raw(offer)
    make_text_child(item, "id", offer_id or clean_text(offer.get("id")))
    make_text_child(item, "categoryId", category_id)
    make_text_child(item, "code", find_offer_param(offer_params(offer), "Артикул") or maudau.child_text(offer, "vendorCode"))
    make_text_child(item, "vendor", maudau.child_text(offer, "vendor"))
    name_ru = maudau.child_text(offer, "name_ru")
    name_ua = maudau.child_text(offer, "name_ua") or name_ru
    description_ru = html_to_text(maudau.child_text(offer, "description_ru"))
    description_ua = html_to_text(maudau.child_text(offer, "description_ua"))
    if not description_ru:
        description_ru = description_ua
    if not description_ua:
        description_ua = description_ru
    if not description_ru:
        description_ru = name_ru
    if not description_ua:
        description_ua = name_ua
    make_text_child(item, "name_ru", name_ru)
    make_text_child(item, "name_ua", name_ua)
    make_text_child(item, "description_ru", description_ru)
    make_text_child(item, "description_ua", description_ua)
    make_text_child(item, "url", maudau.child_text(offer, "url"))

    picture_count = 0
    for picture in offer.findall("picture"):
        url = clean_text(picture.text)
        if not url:
            continue
        make_text_child(item, "image", url)
        picture_count += 1
        if picture_count >= MAX_PICTURES:
            break
    if picture_count == 0:
        return None, ["Нет фото"], color_fallback_used, color_fallback_changes

    make_text_child(item, "priceRUAH", price)
    old_price = maudau.extract_old_price(offer)
    if old_price and old_price != price:
        make_text_child(item, "oldPriceRUAH", old_price)
    make_text_child(item, "stock", "В наявності" if maudau.extract_available(offer) == "true" else "Немає в наявності")

    guarantee_months = (
        find_offer_param(offer_params(offer), "Гарантийный срок, мес.", "Гарантийный срок, мес", "Гарантия")
        .replace(" мес.", "")
        .replace(" мес", "")
    )
    if guarantee_months:
        guarantee = ET.SubElement(item, "guarantee")
        guarantee.set("type", "manufacturer")
        guarantee.text = clean_text(guarantee_months)

    param_pairs, param_issues, color_fallback_used, color_fallback_changes = build_param_pairs(
        offer,
        source_id,
        source_name,
        target_title,
        target_url,
        allo_categories,
        param_rules,
        value_rules,
    )
    issues.extend(param_issues)
    for name, value in param_pairs:
        param = ET.SubElement(item, "param")
        param.set("name", name)
        param.text = value

    return item, issues, color_fallback_used, color_fallback_changes


def write_report(rows: list[list[str]], summary: dict[str, int]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Проблемы"
    ws.append(["offer_id", "source_category_id", "source_category_name", "target_category", "vendor", "name_ru", "issue"])
    for row in rows:
        ws.append(row)
    ws2 = wb.create_sheet("Сводка")
    ws2.append(["Показатель", "Значение"])
    for key, value in summary.items():
        ws2.append([key, value])
    wb.save(REPORT_XLSX)


def write_color_fix_report(rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Автофиксы_цвета"
    ws.append(
        [
            "offer_id",
            "Артикул",
            "Бренд",
            "Название_RU",
            "Раздел_исходник",
            "Раздел_ALLO",
            "Параметр_ALLO",
            "Было_в_исходнике",
            "Стало_в_ALLO",
        ]
    )
    for row in rows:
        ws.append(row)
    wb.save(COLOR_FIX_REPORT_XLSX)


def main() -> int:
    print("===== СТАРТ ALLO =====")
    print("▶ Источники: URL из MAUDAU/update_maudau.py")
    base_path, base_loaded_from_source, base_fallback_path = download_with_backup(
        maudau.BASE_FEED_URL,
        BASE_XML,
        BASE_BACKUP_XML,
        "АЛЛО XML",
        maudau.BASE_DOWNLOAD_TIMEOUT_SEC,
    )
    rozetka_path, rozetka_loaded_from_source, rozetka_fallback_path = download_with_backup(
        maudau.ROZETKA_FEED_URL,
        ROZETKA_XML,
        ROZETKA_BACKUP_XML,
        "Розетка XML",
        maudau.ROZETKA_DOWNLOAD_TIMEOUT_SEC,
    )

    base_tree = ET.parse(str(base_path))
    base_root = base_tree.getroot()
    rozetka_tree = ET.parse(str(rozetka_path))
    rozetka_idx = maudau.build_rozetka_index(rozetka_tree)
    google_idx, google_loaded_from_source, google_fallback_path = maudau.load_google_table_index()
    sources_state = maudau.load_sources_state()
    if base_loaded_from_source:
        maudau.update_source_success(sources_state, "aquafavorit", base_path)
    else:
        maudau.update_source_failure(sources_state, "aquafavorit")
    if rozetka_loaded_from_source:
        maudau.update_source_success(sources_state, "parserbiz", rozetka_path)
    else:
        maudau.update_source_failure(sources_state, "parserbiz")
    if google_loaded_from_source:
        google_used_path = maudau.GOOGLE_TABLE_BACKUP_CSV
        if google_used_path.exists():
            maudau.update_source_success(sources_state, "google_sheet", google_used_path)
        else:
            maudau.update_source_success(sources_state, "google_sheet", BASE_BACKUP_XML)
    else:
        maudau.update_source_failure(sources_state, "google_sheet")
    maudau.save_sources_state(sources_state)

    stale_issue_rows: list[dict] = []
    stale_defs = [
        ("parserbiz", "Исходник Parser.biz", rozetka_fallback_path or resolve_backup_path(ROZETKA_BACKUP_XML), "parserbiz_last.xml"),
        ("aquafavorit", "Исходник AquaFavorit", base_fallback_path or resolve_backup_path(BASE_BACKUP_XML), "aquafavorit_last.xml"),
        ("google_sheet", "Google Sheet", google_fallback_path or maudau.resolve_google_table_backup_path(), "google_table_last.csv"),
    ]
    for source_key, source_label, backup_path, backup_filename in stale_defs:
        row = maudau.source_stale_info(
            sources_state,
            source_key,
            source_label,
            backup_path=backup_path,
            backup_filename=backup_filename,
        )
        if row:
            stale_issue_rows.append(row)
    backup_issue_active = maudau.write_source_issues_report(stale_issue_rows)
    source_names = source_categories(base_root)
    display_source_names = maudau.build_source_category_names(base_root)
    global HEATER_POWER_AREA_PROFILES, PUMP_POWER_FLOW_PROFILES
    HEATER_POWER_AREA_PROFILES = build_heater_power_area_profiles(base_root)
    PUMP_POWER_FLOW_PROFILES = build_pump_power_flow_profiles(base_root)
    stage1_map, skip_ids = load_stage1_category_map()
    param_rules, value_rules = load_mapping_rules()
    allo_categories, _category_options, _filter_options = allo_map.load_allo_catalog()

    categories_absent_in_rozetka = maudau.detect_source_categories_absent_in_rozetka(base_root, rozetka_idx)

    price = ET.Element("price")
    make_text_child(price, "date", maudau.now_kyiv().strftime("%Y-%m-%d %H:%M"))
    make_text_child(price, "firmName", FIRM_NAME)
    make_text_child(price, "firmId", FIRM_ID)
    make_text_child(price, "rate", DEFAULT_RATE)
    categories_node = ET.SubElement(price, "categories")
    items_node = ET.SubElement(price, "items")

    category_id_by_title: dict[str, str] = {}
    category_titles_used: set[str] = set()
    exported_offer_ids: set[str] = set()
    report_rows: list[list[str]] = []
    color_fix_rows: list[list[str]] = []

    def append_report(offer: ET._Element, source_id: str, source_name: str, target_title: str, issue: str) -> None:
        report_rows.append(
            [
                maudau.resolve_offer_id_raw(offer),
                source_id,
                source_name,
                target_title,
                maudau.child_text(offer, "vendor"),
                maudau.child_text(offer, "name_ru") or maudau.child_text(offer, "name"),
                issue,
            ]
        )

    counters = Counter()
    counters["base_loaded_from_url"] = int(base_loaded_from_source)
    counters["rozetka_loaded_from_url"] = int(rozetka_loaded_from_source)

    for offer in list(base_root.xpath("//offer")):
        counters["offers_seen"] += 1
        source_id = maudau.child_text(offer, "categoryId")
        source_name = display_source_names.get(source_id) or source_names.get(source_id, source_id)
        source_key = normalize_key(source_name)
        params = offer_params(offer)
        offer_probe = normalize_key(
            " ".join(
                [
                    maudau.child_text(offer, "name"),
                    maudau.child_text(offer, "name_ru"),
                    maudau.child_text(offer, "description"),
                    maudau.child_text(offer, "description_ru"),
                    " ".join(f"{k} {v}" for k, v in params.items()),
                ]
            )
        )
        if source_id in skip_ids:
            counters["skip_stage1"] += 1
            continue
        if should_force_skip_source_category(source_name):
            counters["skip_source_forced_by_user_rules"] += 1
            continue
        if any(rule.skip_feed for rule in matching_section_rules(source_name)):
            counters["skip_section_rule"] += 1
            continue

        base_category_title = stage1_map.get(source_id, "")
        if not base_category_title:
            counters["skip_no_category"] += 1
            append_report(offer, source_id, source_name, "", "Нет ALLO-категории в этапе 1")
            continue

        # Аксессуары смесителей: в фид только картриджи и аэраторы.
        if "аксессуары смесители" in source_key:
            if "картридж" not in offer_probe and "аэратор" not in offer_probe and "аератор" not in offer_probe:
                counters["skip_mixer_accessories_other"] += 1
                continue

        if source_key == "шланги":
            hose_kind = normalize_key(find_offer_param(params, "Вид", "Тип"))
            if any(token in hose_kind for token in ("фитинги для садового шланга", "садовый", "газовый")):
                counters["skip_hose_not_for_allo"] += 1
                continue

        if source_key == "ведра для мусора":
            volume_probe = clean_text(find_offer_param(params, "Объем", "Об'єм"))
            if not volume_probe and not re.search(r"\d+(?:[.,]\d+)?\s*(?:л|литр)", offer_probe, flags=re.IGNORECASE):
                counters["skip_trash_bin_without_volume"] += 1
                continue

        if source_key == "писсуары":
            if not clean_text(find_offer_param(params, "Монтаж")):
                counters["skip_urinal_without_mount"] += 1
                continue

        if source_key == "аксессуары полотенцесушители радиаторы":
            if not clean_text(find_offer_param(params, "Тип", "Вид")):
                counters["skip_radiator_accessories_without_type"] += 1
                continue

        key = maudau.resolve_offer_id_key(offer)
        rz = rozetka_idx.get(key)
        google_row = google_idx.get(key)
        vendor = maudau.normalize_key(maudau.child_text(offer, "vendor"))
        keep_without_rozetka = (
            bool(google_row)
            or source_id in maudau.KEEP_WITHOUT_ROZETKA_SOURCE_CATEGORIES
            or source_id in categories_absent_in_rozetka
            or vendor in maudau.ALLOWED_VENDORS
        )
        if rz is None and not keep_without_rozetka:
            counters["skip_no_rozetka_or_google"] += 1
            continue
        changed_price, changed_other = apply_price_and_availability(offer, rz, google_row)
        if changed_price:
            counters["changed_price"] += 1
        if changed_other:
            counters["changed_old_price_or_available"] += 1
        if google_row and not rz:
            counters["used_google_fallback"] += 1

        target_title, target_url = resolve_offer_category(offer, source_id, source_name, base_category_title)
        if not target_title:
            counters["skip_no_target_title"] += 1
            append_report(offer, source_id, source_name, "", "Не удалось определить ALLO-категорию")
            continue
        if normalize_key(target_title) in {"шторки для ванной", "шторки для ванн"}:
            counters["skip_category_disabled"] += 1
            continue

        if target_title == "Арматура для унитазов":
            arm_type = normalize_key(find_offer_param(params, "Тип", "Вид"))
            if not any(token in arm_type for token in ("комплект арматуры для бачка", "впускной клапан", "сливной клапан")):
                counters["skip_toilet_armature_noncanonical"] += 1
                continue

        category_id = category_id_by_title.setdefault(target_title, stable_category_id(target_title))
        category_titles_used.add(target_title)
        item, issues, color_fallback_used, color_fallback_changes = offer_to_allo_item(
            offer,
            category_id,
            source_name,
            target_title,
            target_url,
            allo_categories,
            param_rules,
            value_rules,
        )
        if item is None:
            counters["skip_invalid_item"] += 1
            for issue in issues:
                append_report(offer, source_id, source_name, target_title, issue)
            continue

        if target_title == "Мусорные ведра":
            item_param_names = {(p.get("name") or "").strip() for p in item.findall("param")}
            if "Объем" not in item_param_names and "Об'єм" not in item_param_names:
                counters["skip_trash_bin_without_volume_after_mapping"] += 1
                continue
        if source_key == "шланги":
            item_param_names = {(p.get("name") or "").strip() for p in item.findall("param")}
            if "Диаметр резьбы" not in item_param_names:
                counters["skip_hose_without_thread_diameter_after_mapping"] += 1
                continue
        if source_key == "комплектующие к инсталляциям инсталляции":
            if any(
                issue.startswith("Не заполнен обязательный параметр ALLO:")
                or issue.startswith("Не сопоставлено обязательное значение:")
                for issue in issues
            ):
                counters["skip_installation_components_with_required_issues"] += 1
                continue
        if source_key == "кухонные аксессуары кухня":
            if any(
                issue.startswith("Не заполнен обязательный параметр ALLO:")
                or issue.startswith("Не сопоставлено обязательное значение:")
                for issue in issues
            ):
                counters["skip_kitchen_accessories_with_required_issues"] += 1
                continue
        if should_drop_by_required_issues(source_name, issues):
            counters["skip_selected_categories_with_required_issues"] += 1
            continue

        item_id = normalize_key(item.findtext("id"))
        if item_id and item_id in exported_offer_ids:
            counters["skip_duplicate_offer_id"] += 1
            append_report(offer, source_id, source_name, target_title, "Дубль offer id после канонизации")
            continue
        if item_id:
            exported_offer_ids.add(item_id)
        items_node.append(item)
        counters["items_exported"] += 1
        if color_fallback_used:
            counters["items_color_fallback_fixed"] += 1
            item_offer_id = clean_text(item.findtext("id"))
            item_code = clean_text(item.findtext("code"))
            item_vendor = clean_text(item.findtext("vendor"))
            item_name_ru = clean_text(item.findtext("name_ru"))
            for filter_name, old_value, new_value in color_fallback_changes:
                color_fix_rows.append(
                    [
                        item_offer_id,
                        item_code,
                        item_vendor,
                        item_name_ru,
                        source_name,
                        target_title,
                        filter_name,
                        old_value,
                        new_value,
                    ]
                )
        for issue in issues:
            counters["items_with_param_issues"] += 1
            append_report(offer, source_id, source_name, target_title, issue)

    for title in sorted(category_titles_used):
        category = ET.SubElement(categories_node, "category")
        make_text_child(category, "id", category_id_by_title[title])
        make_text_child(category, "name", title)

    OUTPUT_XML.parent.mkdir(parents=True, exist_ok=True)
    tree = ET.ElementTree(price)
    tree.write(str(OUTPUT_XML), encoding="UTF-8", xml_declaration=True, pretty_print=True)

    write_report(report_rows, dict(counters))
    write_color_fix_report(color_fix_rows)
    size_mb = OUTPUT_XML.stat().st_size / (1024 * 1024)
    print(f"✅ ALLO XML создан: {OUTPUT_XML}")
    print(f"📋 Отчёт: {REPORT_XLSX}")
    print(f"🎨 Отчёт автофиксов цвета: {COLOR_FIX_REPORT_XLSX}")
    print(f"📦 Товаров в фиде: {counters['items_exported']}")
    print(f"📂 Категорий в фиде: {len(category_titles_used)}")
    print(f"🎨 Автофиксов цвета по канонам ALLO: {counters['items_color_fallback_fixed']}")
    print(f"📐 Размер: {size_mb:.2f} MB")
    source_header = "\n".join(
        [
            maudau.source_status_block("Розетка XML", rozetka_loaded_from_source, rozetka_fallback_path),
            maudau.source_status_block("АЛЛО XML", base_loaded_from_source, base_fallback_path),
        ]
    )
    backup_issue_line = "⛔️Необходимо обновить Backup!!!\n\n" if backup_issue_active else ""
    report_msg = f"""===== ⭕️АЛЛО⭕️=====
{source_header}
{backup_issue_line}

❌ Удалено из файла (не в Розетке, кроме Мойдодыр/Dusel): {counters['skip_no_rozetka_or_google']}
🆔 Удалено дублей offer id: {counters['skip_duplicate_offer_id']}
💲 Обновлено цен: {counters['changed_price']}
🔁 Обновлено старых цен и наличия: {counters['changed_old_price_or_available']}

📦 Отправляем на АЛЛО товаров: {counters['items_exported']}
📐 Размер итогового файла: {size_mb:.2f} MB
===== ГОТОВО ✅ ====="""
    print(report_msg)
    maudau.send_telegram(report_msg)
    print("===== ГОТОВО ALLO ✅ =====")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        err = f"""===== ⭕️АЛЛО⭕️=====
▶ Загрузка: Розетка XML
⚠ Ошибка: {exc}
===== ОШИБКА ❌ ====="""
        print(err, file=sys.stderr)
        try:
            maudau.send_telegram(err)
        except Exception:
            pass
        raise SystemExit(1)
