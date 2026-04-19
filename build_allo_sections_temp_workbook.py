#!/usr/bin/env python3
from __future__ import annotations

import re
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from difflib import SequenceMatcher

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path("/Users/Kezhunya/Documents/New project")
EXPORT_XLSX = Path("/Volumes/X-Files/Загрузки рабочие/products_prices_04-07-2026_18-34.xlsx")
OLD_ALLO_XML = ROOT / "АЛЛО" / "исходник для rozetka.xml"
NEW_BASE_XML = ROOT / "АЛЛО" / "исходник на maudau без розетки.xml"
ALLO_377_XLSX = ROOT / "АЛЛО" / "Категории 377.xlsx"
ALLO_MAPPING_XLSX = ROOT / "АЛЛО" / "ALLO_сопоставление.xlsx"
ALLO_FULL_MAPPING_XLSX = ROOT / "АЛЛО" / "Мапинг.xlsx"
OUTPUT_XLSX = ROOT / "АЛЛО" / "Временное_сопоставление_разделов_ALLO.xlsx"
MIN_CATEGORY_SIZE = 14

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")


@dataclass
class ExportRow:
    site_sku: str
    allo_sku: str
    partner_sku: str
    partner_sku_norm: str
    name_ru: str
    availability: str
    price: object
    status: str


@dataclass
class OfferInfo:
    key: str
    category_id: str
    category_name: str
    offer_id: str
    vendor_code: str
    name_ru: str


@dataclass
class ExistingTempEdit:
    hint_377: str
    confirm_377: str


SPECIAL_FILTER_RULES = {
    "1067": "Тип=Сифоны; Применение=Для ванной",
    "1118": "Тип=Сифоны; Применение=Для поддона",
    "1257": "Вид=Труба",
    "1276": "Тип=Панели",
    "1255": "Если тип товара Коллектор/Коллекторная группа -> категория Коллекторы; если Коллекторный шкаф -> категория Коллекторные шкафы",
}

MANUAL_CONFIRM_OVERRIDES = {
    "1175": "Чаши Генуя",
}

TITLE_ALIASES = {
    "Аксессуары к кухонным мойкам": "Аксессуары к кухонным мойкам",
    "Комплектующие к радиаторам": "Комплектующие к радиаторам",
    "Комплектующие к инсталляциям": "Комплектующие к инсталляциям",
    "Инсталляции": "Инсталляции",
    "Сифоны и трапы": "Сифоны и трапы",
    "Фитинги для канализационных труб": "Фитинги для канализационных труб",
    "Фитинги для водопроводных труб": "Фитинги для водопроводных труб",
    "Керамические обогреватели": "Керамические обогреватели",
    "Комплектующая запорная арматура": "Комплектующая запорная арматура",
    "Ведра и корзины для ванной комнаты": "Мусорные ведра",
    "Дозаторы для жидкого мыла": "Дозаторы (диспенсеры) для ванной комнаты",
    "Крючки для ванной": "Крючки для ванной",
    "Держатели для ванной комнаты": "Держатели для ванной комнаты",
    "Тумбы для ванной комнаты": "Тумбы для ванной комнаты",
    "Вытяжные вентиляторы": "Вытяжные вентиляторы",
    "Электрические котлы": "Электрические котлы",
    "Газовые котлы": "Газовые котлы",
    "Твердотопливные котлы": "Твердотопливные котлы",
    "Водопроводные трубы": "Водопроводные трубы",
    "Чаша Генуя": "Чаши Генуя",
    "Комплектующие": "Комплектующие",
}


def norm(value: object | None) -> str:
    return str(value or "").strip()


def sku_norm(value: object | None) -> str:
    text = norm(value)
    if text.endswith("-0377"):
        return text[:-5]
    return text


def normalize_key(value: object | None) -> str:
    return " ".join(norm(value).lower().replace("ё", "е").split())


def similarity(a: str, b: str) -> float:
    ak = normalize_key(a)
    bk = normalize_key(b)
    if not ak or not bk:
        return 0.0
    if ak == bk:
        return 1.0
    if ak in bk or bk in ak:
        return 0.92
    return SequenceMatcher(None, ak, bk).ratio()


def add_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autosize(ws) -> None:
    for column in ws.columns:
        max_len = 0
        letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def build_display_category_names(path: Path) -> dict[str, str]:
    root = ET.parse(path).getroot()
    shop = root.find("shop")
    items: dict[str, str] = {}
    parent_by_id: dict[str, str] = {}

    for category in shop.find("categories").findall("category"):
        cid = norm(category.get("id"))
        if not cid:
            continue
        items[cid] = norm(category.text)
        parent = norm(category.get("parentId"))
        if parent:
            parent_by_id[cid] = parent

    def is_generic(name: str) -> bool:
        key = normalize_key(name)
        return ("комплектующ" in key) or ("аксессуар" in key)

    def resolve_parent_section(cid: str) -> str:
        cur = parent_by_id.get(cid, "")
        seen: set[str] = set()
        while cur and cur not in seen:
            seen.add(cur)
            candidate = items.get(cur, "")
            if candidate and not is_generic(candidate):
                return candidate
            cur = parent_by_id.get(cur, "")
        return ""

    result: dict[str, str] = {}
    for cid, name in items.items():
        display = name
        if is_generic(name):
            section = resolve_parent_section(cid)
            if section:
                display = f"{name} ({section})"
        result[cid] = display

    count_by_name: dict[str, int] = defaultdict(int)
    for value in result.values():
        count_by_name[value] += 1

    for cid, value in list(result.items()):
        if count_by_name.get(value, 0) <= 1:
            continue
        if "(" in value and ")" in value:
            continue
        section = resolve_parent_section(cid)
        if section:
            result[cid] = f"{value} ({section})"
    return result


def extract_offer_keys(offer: ET.Element, include_offer_id: bool) -> set[str]:
    keys = {
        sku_norm(offer.findtext("vendorCode")),
    }
    if include_offer_id:
        keys.add(sku_norm(offer.get("id")))
    for param in offer.findall("param"):
        if norm(param.get("name")) == "Артикул":
            keys.add(sku_norm(param.text))
    return {value for value in keys if value}


def load_export_rows() -> list[ExportRow]:
    wb = load_workbook(EXPORT_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[ExportRow] = []
    for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if index == 1:
            continue
        values = list(row)
        if not any(value is not None and str(value).strip() for value in values):
            continue
        rows.append(
            ExportRow(
                site_sku=norm(values[0]),
                allo_sku=norm(values[1]),
                partner_sku=norm(values[2]),
                partner_sku_norm=sku_norm(values[2]),
                name_ru=norm(values[3]),
                availability=norm(values[5]),
                price=values[7],
                status=norm(values[11]),
            )
        )
    return rows


def load_xml_offers(path: Path, include_offer_id: bool, category_display_names: dict[str, str] | None = None) -> tuple[dict[str, str], dict[str, list[OfferInfo]], Counter]:
    root = ET.parse(path).getroot()
    shop = root.find("shop")
    categories = {
        norm(category.get("id")): norm(category.text)
        for category in shop.find("categories").findall("category")
    }
    by_key: dict[str, list[OfferInfo]] = defaultdict(list)
    category_totals: Counter = Counter()
    for offer in shop.find("offers").findall("offer"):
        category_id = norm(offer.findtext("categoryId"))
        category_name = (category_display_names or {}).get(category_id, categories.get(category_id, ""))
        info = OfferInfo(
            key="",
            category_id=category_id,
            category_name=category_name,
            offer_id=norm(offer.get("id")),
            vendor_code=norm(offer.findtext("vendorCode")),
            name_ru=norm(offer.findtext("name_ru") or offer.findtext("name")),
        )
        category_totals[(category_id, category_name)] += 1
        for key in extract_offer_keys(offer, include_offer_id=include_offer_id):
            by_key[key].append(
                OfferInfo(
                    key=key,
                    category_id=info.category_id,
                    category_name=info.category_name,
                    offer_id=info.offer_id,
                    vendor_code=info.vendor_code,
                    name_ru=info.name_ru,
                )
            )
    return categories, by_key, category_totals


def collapse_unique_offers(by_key: dict[str, list[OfferInfo]]) -> tuple[dict[str, OfferInfo], dict[str, list[OfferInfo]]]:
    unique: dict[str, OfferInfo] = {}
    ambiguous: dict[str, list[OfferInfo]] = {}
    for key, offers in by_key.items():
        unique_categories = {(item.category_id, item.category_name) for item in offers}
        if len(unique_categories) == 1:
            unique[key] = offers[0]
        else:
            ambiguous[key] = offers
    return unique, ambiguous


def load_allo_377_categories() -> list[tuple[str, str]]:
    wb = load_workbook(ALLO_377_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, title = row[:2]
        if title:
            result.append((norm(code), norm(title)))
    return result


def load_current_mapping_suggestions() -> dict[str, str]:
    if not ALLO_MAPPING_XLSX.exists():
        return {}
    wb = load_workbook(ALLO_MAPPING_XLSX, read_only=True, data_only=True)
    if "Категории" not in wb.sheetnames:
        return {}
    ws = wb["Категории"]
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        source_id = norm(row[0])
        suggested = norm(row[6]) or norm(row[4])
        if source_id and suggested:
            result[source_id] = suggested
    return result


def load_existing_temp_edits() -> dict[str, ExistingTempEdit]:
    if not OUTPUT_XLSX.exists():
        return {}
    wb = load_workbook(OUTPUT_XLSX, data_only=False)
    result: dict[str, ExistingTempEdit] = {}
    for sheet_name in ["Новые_разделы_сайта", "Отброшено_<14"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            source_id = norm(row[0])
            if not source_id:
                continue
            current = result.get(source_id, ExistingTempEdit("", ""))
            result[source_id] = ExistingTempEdit(
                hint_377=norm(row[12]) or current.hint_377,
                confirm_377=norm(row[13]) or current.confirm_377,
            )
    return result


def extract_urls(value: str) -> list[str]:
    return re.findall(r"https?://[^\s\n]+", value or "")


def load_allo_filters_by_url() -> dict[str, dict]:
    wb = load_workbook(ALLO_FULL_MAPPING_XLSX, read_only=True, data_only=True)
    ws = wb["Фильтры"]
    result: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        category_title, url, _order, code, filter_name, value, _value_url, _count, value_type, is_required = row
        url = norm(url)
        filter_name = norm(filter_name)
        if not url or not filter_name:
            continue
        data = result.setdefault(
            url,
            {
                "title": norm(category_title),
                "filters": defaultdict(lambda: {"code": "", "required": False, "values": []}),
            },
        )
        info = data["filters"][filter_name]
        info["code"] = norm(code)
        if is_required:
            info["required"] = True
        clean_value = norm(value)
        if clean_value and clean_value not in info["values"]:
            info["values"].append(clean_value)
    return result


def normalize_url(url: str) -> str:
    clean = norm(url).split("?", 1)[0].split("#", 1)[0]
    return clean.rstrip("/") + "/" if clean else ""


def load_allo_url_titles() -> dict[str, str]:
    wb = load_workbook(ALLO_FULL_MAPPING_XLSX, read_only=True, data_only=True)
    result: dict[str, str] = {}

    if "Фильтры" in wb.sheetnames:
        ws = wb["Фильтры"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            category_title, url = row[:2]
            if url and category_title:
                result.setdefault(normalize_url(str(url)), norm(category_title))

    if "Категории" in wb.sheetnames:
        ws = wb["Категории"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            source_url, source_title, _section, child_title, child_url = row[:5]
            if source_url and source_title:
                result.setdefault(normalize_url(str(source_url)), norm(source_title))
            if child_url and child_title:
                result.setdefault(normalize_url(str(child_url)), norm(child_title))
    return result


def map_title_to_377(title: str, allo_377: list[tuple[str, str]]) -> str:
    clean_title = norm(title)
    if not clean_title:
        return ""
    alias = TITLE_ALIASES.get(clean_title, clean_title)
    for _code, candidate in allo_377:
        if normalize_key(candidate) == normalize_key(alias):
            return candidate
    return ""


def infer_confirm_from_hint(hint_text: str, allo_377: list[tuple[str, str]], url_titles: dict[str, str]) -> str:
    hint = norm(hint_text)
    if not hint:
        return ""

    urls = extract_urls(hint)
    matched_titles: list[str] = []
    for url in urls:
        normalized = normalize_url(url)
        if normalized in url_titles:
            matched_titles.append(url_titles[normalized])
            continue
        best_title = ""
        best_len = 0
        for known_url, title in url_titles.items():
            if normalized.startswith(known_url) and len(known_url) > best_len:
                best_len = len(known_url)
                best_title = title
        if best_title:
            matched_titles.append(best_title)

    matched_titles = [title for title in matched_titles if title]
    unique_titles: list[str] = []
    for title in matched_titles:
        if title not in unique_titles:
            unique_titles.append(title)

    if len(unique_titles) == 1:
        return map_title_to_377(unique_titles[0], allo_377)
    if urls:
        return ""

    for line in hint.splitlines():
        direct = map_title_to_377(line, allo_377)
        if direct:
            return direct
    return map_title_to_377(hint, allo_377)


def suggest_377_category(current_mapping: dict[str, str], source_category_id: str, old_category_name: str, allo_377: list[tuple[str, str]]) -> str:
    mapped = current_mapping.get(source_category_id, "")
    if mapped:
        return mapped
    for _code, title in allo_377:
        if normalize_key(old_category_name) == normalize_key(title):
            return title
    return ""


def status_fill(status: str) -> PatternFill:
    if status == "Точно по товарам":
        return GREEN_FILL
    if status == "Подтверждено вручную":
        return GREEN_FILL
    if status == "Нужно проверить":
        return YELLOW_FILL
    return RED_FILL


def main() -> Path:
    export_rows = load_export_rows()
    allo_377 = load_allo_377_categories()
    mapping_suggestions = load_current_mapping_suggestions()
    existing_edits = load_existing_temp_edits()
    new_display_names = build_display_category_names(NEW_BASE_XML)
    allo_filters_by_url = load_allo_filters_by_url()
    allo_url_titles = load_allo_url_titles()

    _old_categories, old_by_key, old_category_totals = load_xml_offers(OLD_ALLO_XML, include_offer_id=True)
    _new_categories, new_by_key, new_category_totals = load_xml_offers(
        NEW_BASE_XML,
        include_offer_id=False,
        category_display_names=new_display_names,
    )
    old_unique, old_ambiguous = collapse_unique_offers(old_by_key)
    new_unique, new_ambiguous = collapse_unique_offers(new_by_key)

    detail_rows: list[dict[str, object]] = []
    export_count_by_new_category: Counter = Counter()
    export_with_old_by_new_category: Counter = Counter()
    old_counter_by_new_category: dict[tuple[str, str], Counter] = defaultdict(Counter)
    sample_rows_by_new_category: dict[tuple[str, str], list[str]] = defaultdict(list)
    missing_old_feed = 0
    missing_new_source = 0
    matched_old_export = 0
    matched_new_export = 0

    for row in export_rows:
        sku = row.partner_sku_norm
        old_offer = old_unique.get(sku)
        new_offer = new_unique.get(sku)
        old_options = old_ambiguous.get(sku, [])
        new_options = new_ambiguous.get(sku, [])

        old_category_id = old_offer.category_id if old_offer else ""
        old_category_name = old_offer.category_name if old_offer else ""
        new_category_id = new_offer.category_id if new_offer else ""
        new_category_name = new_offer.category_name if new_offer else ""

        if new_offer:
            export_count_by_new_category[(new_category_id, new_category_name)] += 1
            matched_new_export += 1
        if new_offer and old_offer:
            export_with_old_by_new_category[(new_category_id, new_category_name)] += 1
            old_counter_by_new_category[(new_category_id, new_category_name)][(old_category_id, old_category_name)] += 1
            if len(sample_rows_by_new_category[(new_category_id, new_category_name)]) < 5:
                sample_rows_by_new_category[(new_category_id, new_category_name)].append(f"{sku} | {row.name_ru}")
        if old_offer:
            matched_old_export += 1

        if not old_offer and not old_options:
            missing_old_feed += 1
        if not new_offer and not new_options:
            missing_new_source += 1

        if old_offer and new_offer:
            row_status = "Есть связка"
        elif new_options:
            row_status = "Несколько новых категорий"
        elif old_options:
            row_status = "Несколько старых категорий"
        elif not old_offer:
            row_status = "Нет в старом ALLO фиде"
        elif not new_offer:
            row_status = "Нет в новом источнике"
        else:
            row_status = "Проверить"

        detail_rows.append(
            {
                "partner_sku": row.partner_sku,
                "partner_sku_norm": sku,
                "name_ru": row.name_ru,
                "old_category_id": old_category_id,
                "old_category_name": old_category_name,
                "old_category_variants": "\n".join(
                    sorted({f"{item.category_id} | {item.category_name}" for item in old_options})
                ),
                "new_category_id": new_category_id,
                "new_category_name": new_category_name,
                "new_category_variants": "\n".join(
                    sorted({f"{item.category_id} | {item.category_name}" for item in new_options})
                ),
                "availability": row.availability,
                "price": row.price,
                "status": row_status,
            }
        )

    summary_rows: list[dict[str, object]] = []
    for (new_category_id, new_category_name), total_in_source in sorted(
        new_category_totals.items(),
        key=lambda item: (-item[1], item[0][1], item[0][0]),
    ):
        export_now = export_count_by_new_category.get((new_category_id, new_category_name), 0)
        linked_now = export_with_old_by_new_category.get((new_category_id, new_category_name), 0)
        old_counter = old_counter_by_new_category.get((new_category_id, new_category_name), Counter())
        top_old_id = ""
        top_old_name = ""
        top_old_count = 0
        if old_counter:
            (top_old_id, top_old_name), top_old_count = old_counter.most_common(1)[0]
        share = top_old_count / linked_now if linked_now else 0.0
        if linked_now == 0:
            status = "Нет опоры в ALLO"
        elif top_old_count >= 5 and share >= 0.85:
            status = "Точно по товарам"
        elif top_old_count >= 3 and share >= 0.60:
            status = "Нужно проверить"
        else:
            status = "Смешано/вручную"

        suggested_377 = suggest_377_category(mapping_suggestions, new_category_id, top_old_name or new_category_name, allo_377)
        preserved = existing_edits.get(new_category_id, ExistingTempEdit("", ""))
        inferred_from_hint = infer_confirm_from_hint(preserved.hint_377, allo_377, allo_url_titles)
        all_old = "\n".join(
            f"{cid} | {cname} ({count})"
            for (cid, cname), count in old_counter.most_common(8)
        )
        summary_rows.append(
            {
                "new_category_id": new_category_id,
                "new_category_name": new_category_name,
                "total_in_source": total_in_source,
                "export_now": export_now,
                "new_not_in_allo": max(total_in_source - export_now, 0),
                "linked_now": linked_now,
                "top_old_id": top_old_id,
                "top_old_name": top_old_name,
                "top_old_count": top_old_count,
                "top_old_share": round(share, 4),
                "old_variants": all_old,
                "mapping_hint": mapping_suggestions.get(new_category_id, ""),
                "suggested_377": preserved.hint_377 or suggested_377,
                "confirm_377": MANUAL_CONFIRM_OVERRIDES.get(new_category_id, "")
                or inferred_from_hint
                or (preserved.confirm_377 if not preserved.hint_377 else "")
                or (suggested_377 if status == "Точно по товарам" and not preserved.hint_377 else ""),
                "examples": "\n".join(sample_rows_by_new_category.get((new_category_id, new_category_name), [])),
                "status": status,
            }
        )

    for row in summary_rows:
        if norm(row["confirm_377"]):
            row["status"] = "Подтверждено вручную"

    kept_summary_rows = [
        row
        for row in summary_rows
        if int(row["total_in_source"]) >= MIN_CATEGORY_SIZE
        or row["status"] in {"Точно по товарам", "Подтверждено вручную"}
    ]
    dropped_summary_rows = [
        row
        for row in summary_rows
        if int(row["total_in_source"]) < MIN_CATEGORY_SIZE
        and row["status"] not in {"Точно по товарам", "Подтверждено вручную"}
    ]

    filter_rows: list[list[object]] = []
    seen_filter_keys: set[tuple[str, str, str, str]] = set()
    for item in summary_rows:
        hint_text = norm(item["suggested_377"])
        confirm_text = norm(item["confirm_377"])
        urls = []
        for raw in [hint_text, confirm_text]:
            urls.extend(extract_urls(raw))
        dedup_urls = []
        for url in urls:
            if url not in dedup_urls:
                dedup_urls.append(url)
        if not dedup_urls and item["new_category_id"] not in SPECIAL_FILTER_RULES:
            continue
        if not dedup_urls:
            dedup_urls = [""]
        for url in dedup_urls:
            category_filters = allo_filters_by_url.get(url, {})
            filter_block = category_filters.get("filters", {})
            if not filter_block:
                key = (str(item["new_category_id"]), url, "", "")
                if key in seen_filter_keys:
                    continue
                seen_filter_keys.add(key)
                filter_rows.append(
                    [
                        item["new_category_id"],
                        item["new_category_name"],
                        url,
                        category_filters.get("title", ""),
                        SPECIAL_FILTER_RULES.get(str(item["new_category_id"]), ""),
                        "",
                        "",
                        "",
                        hint_text,
                    ]
                )
                continue
            for filter_name in sorted(filter_block.keys()):
                info = filter_block[filter_name]
                values_preview = ", ".join(info["values"][:20])
                key = (str(item["new_category_id"]), url, filter_name, values_preview)
                if key in seen_filter_keys:
                    continue
                seen_filter_keys.add(key)
                filter_rows.append(
                    [
                        item["new_category_id"],
                        item["new_category_name"],
                        url,
                        category_filters.get("title", ""),
                        SPECIAL_FILTER_RULES.get(str(item["new_category_id"]), ""),
                        filter_name,
                        "Да" if info["required"] else "",
                        values_preview,
                        hint_text,
                    ]
                )

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Новые_разделы_сайта"
    ws_dropped = wb.create_sheet("Отброшено_<14")
    ws_detail = wb.create_sheet("Товары_связка")
    ws_filters = wb.create_sheet("Фильтры_ALLO")
    ws_377 = wb.create_sheet("Справочник_377")
    ws_info = wb.create_sheet("Сводка")
    ws_lists = wb.create_sheet("Lists")

    add_header(ws_info, ["Показатель", "Значение"])
    for label, value in [
        ("Товаров в экспорте ALLO", len(export_rows)),
        ("SKU из экспорта найдены в старом ALLO-фиде", matched_old_export),
        ("SKU из экспорта найдены в новом источнике сайта", matched_new_export),
        ("SKU из экспорта без совпадения в старом ALLO-фиде", missing_old_feed),
        ("SKU из экспорта без совпадения в новом источнике сайта", missing_new_source),
        ("Категорий в справочнике 377", len(allo_377)),
        ("Разделов нового сайта в основном листе", len(kept_summary_rows)),
        ("Отброшено категорий < 14 товаров", len(dropped_summary_rows)),
    ]:
        ws_info.append([label, value])

    summary_headers = [
        "ID категории сайта",
        "Новая категория сайта",
        "Всего товаров в новом источнике",
        "Уже есть в экспорте ALLO",
        "Новых товаров сверх текущего ALLO",
        "Есть точная связка со старым ALLO",
        "Текущий раздел ALLO ID",
        "Текущий раздел ALLO (из старого фида)",
        "Сколько товаров уже лежит там",
        "Доля главного раздела среди связанных товаров",
        "Все текущие разделы ALLO по товарам",
        "Подсказка из текущего сопоставления",
        "Подсказка категории 377",
        "Подтвердить категорию 377",
        "Примеры товаров",
        "Статус",
    ]
    add_header(ws_summary, summary_headers)
    add_header(ws_dropped, summary_headers)
    for target_ws, rows_set in [(ws_summary, kept_summary_rows), (ws_dropped, dropped_summary_rows)]:
        for item in rows_set:
            target_ws.append(
                [
                    item["new_category_id"],
                    item["new_category_name"],
                    item["total_in_source"],
                    item["export_now"],
                    item["new_not_in_allo"],
                    item["linked_now"],
                    item["top_old_id"],
                    item["top_old_name"],
                    item["top_old_count"],
                    item["top_old_share"],
                    item["old_variants"],
                    item["mapping_hint"],
                    item["suggested_377"],
                    item["confirm_377"],
                    item["examples"],
                    item["status"],
                ]
            )
            fill = status_fill(str(item["status"]))
            for cell in target_ws[target_ws.max_row]:
                cell.fill = fill

    add_header(
        ws_filters,
        [
            "ID категории сайта",
            "Новая категория сайта",
            "URL категории ALLO",
            "Категория ALLO",
            "Какой фильтр/правило выбрать",
            "Фильтр ALLO",
            "Обязательный",
            "Значения фильтра ALLO",
            "Комментарий из таблицы",
        ],
    )
    for row in filter_rows:
        ws_filters.append(row)

    add_header(
        ws_detail,
        [
            "SKU Партнера",
            "SKU Партнера без -0377",
            "Название товара RU",
            "Старая категория ALLO ID",
            "Старая категория ALLO (из старого фида)",
            "Варианты старой категории",
            "Новая категория сайта ID",
            "Новая категория сайта",
            "Варианты новой категории сайта",
            "Наличие",
            "Цена",
            "Статус",
        ],
    )
    for item in detail_rows:
        ws_detail.append(
            [
                item["partner_sku"],
                item["partner_sku_norm"],
                item["name_ru"],
                item["old_category_id"],
                item["old_category_name"],
                item["old_category_variants"],
                item["new_category_id"],
                item["new_category_name"],
                item["new_category_variants"],
                item["availability"],
                item["price"],
                item["status"],
            ]
        )

    add_header(ws_377, ["Код категории 377", "Категория 377"])
    for code, title in allo_377:
        ws_377.append([code, title])

    ws_lists["A1"] = "Категории 377"
    for idx, (_code, title) in enumerate(allo_377, start=2):
        ws_lists.cell(idx, 1).value = title
    ws_lists.sheet_state = "hidden"

    validation_main = DataValidation(
        type="list",
        formula1=f"=Lists!$A$2:$A${len(allo_377) + 1}",
        allow_blank=True,
    )
    validation_dropped = DataValidation(
        type="list",
        formula1=f"=Lists!$A$2:$A${len(allo_377) + 1}",
        allow_blank=True,
    )
    ws_summary.add_data_validation(validation_main)
    validation_main.add(f"N2:N{ws_summary.max_row}")
    ws_dropped.add_data_validation(validation_dropped)
    validation_dropped.add(f"N2:N{ws_dropped.max_row}")

    for ws in [ws_info, ws_summary, ws_dropped, ws_detail, ws_filters, ws_377]:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        autosize(ws)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    return OUTPUT_XLSX


if __name__ == "__main__":
    path = main()
    print(path)
