#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import build_allo_color_review_workbook as review


ROOT = Path("/Users/Kezhunya/Documents/New project/АЛЛО")
OUT_WORKBOOK = ROOT / "ALLO_цвета_ручное_сопоставление.xlsx"


def clean(value) -> str:
    return " ".join(str(value or "").strip().split())


def normalize(value) -> str:
    return review.normalize(value)


def write_manual_workbook(rows: list[dict[str, str]]) -> None:
    color_canons = review.load_color_canons()

    wb = Workbook()
    ws = wb.active
    ws.title = "Цвета_к_сопоставлению"
    headers = [
        "Раздел_ALLO",
        "Параметр_ALLO",
        "Было_в_исходнике",
        "Подтвердить_канон_ALLO",
        "Автоподстановка_сейчас",
        "Разделы_исходника",
        "Примеры_артикулов",
        "Статус",
    ]
    ws.append(headers)

    for row in rows:
        ws.append(
            [
                clean(row.get("Раздел_ALLO", "")),
                clean(row.get("Параметр_ALLO", "")),
                clean(row.get("Было_в_исходнике", "")),
                clean(row.get("Автоподстановка_сейчас", "")),
                clean(row.get("Автоподстановка_сейчас", "")),
                clean(row.get("Разделы_исходника", "")),
                clean(row.get("Примеры_артикулов", "")),
                "Требует проверки",
            ]
        )

    ws.freeze_panes = "A2"
    green = PatternFill(fill_type="solid", start_color="E2F0D9", end_color="E2F0D9")
    yellow = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).fill = green
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=4).fill = yellow
        ws.cell(row=r, column=8).value = (
            f'=IF(D{r}=E{r},"Проверить автоподстановку","Изменено вручную")'
        )

    lists = wb.create_sheet("Lists")
    col = 1
    validation_ranges: dict[tuple[str, str], str] = {}
    for key, vals in sorted(color_canons.items(), key=lambda x: (x[0][0], x[0][1])):
        if not vals:
            continue
        cat_key, param_key = key
        lists.cell(row=1, column=col, value=f"{cat_key}::{param_key}")
        for i, v in enumerate(vals, start=2):
            lists.cell(row=i, column=col, value=v)
        end_row = len(vals) + 1
        letter = lists.cell(row=1, column=col).column_letter
        validation_ranges[key] = f"'Lists'!${letter}$2:${letter}${end_row}"
        col += 1
    lists.sheet_state = "hidden"

    for r in range(2, ws.max_row + 1):
        cat = clean(ws.cell(row=r, column=1).value)
        param = clean(ws.cell(row=r, column=2).value)
        rng = validation_ranges.get((normalize(cat), normalize(param)))
        if not rng:
            continue
        dv = DataValidation(type="list", formula1=f"={rng}", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=r, column=4))

    widths = {
        1: 34,
        2: 20,
        3: 24,
        4: 28,
        5: 24,
        6: 42,
        7: 32,
        8: 24,
    }
    for cidx, w in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=cidx).column_letter].width = w

    wb.save(OUT_WORKBOOK)


def main() -> int:
    rows = review.collect_rows()
    write_manual_workbook(rows)
    print(f"✅ Создан файл: {OUT_WORKBOOK}")
    print(f"📌 Строк для ручного сопоставления: {len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

