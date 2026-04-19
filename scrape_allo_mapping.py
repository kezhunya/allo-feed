from __future__ import annotations

import csv
import re
import sys
import time
from collections import OrderedDict, deque
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


ROOT_URL = "https://allo.ua/ru/santehnika/"
SHEET_CSV_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1aBFskFB-e70xvWBqfe71_5thXBQ9Aqly7PebsVTkq1M/export?format=csv&gid=0"
)
OUTPUT_PATH = Path("/Users/Kezhunya/Documents/New project/АЛЛО/Мапинг.xlsx")
REQUEST_TIMEOUT = 40
REQUEST_DELAY = 0.35
GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0 Safari/537.36"
)


@dataclass
class NavLink:
    source_url: str
    source_title: str
    section_title: str
    child_label: str
    child_url: str
    child_kind: str
    child_page_type: str


def normalize_space(value: str | None) -> str:
    return " ".join((value or "").split())


def normalize_url(url: str) -> str:
    parsed = urlparse(url)
    path = parsed.path or "/"
    if not path.endswith("/"):
        path += "/"
    return f"{parsed.scheme}://{parsed.netloc}{path}"


def normalize_path(url: str) -> str:
    path = urlparse(url).path.strip().lower()
    path = re.sub(r"^/(ru|ua)/", "/", path)
    path = re.sub(r"/+", "/", path)
    path = path.rstrip("/") or "/"
    return path


def path_parts(url: str) -> list[str]:
    path = normalize_path(url)
    return [part for part in path.strip("/").split("/") if part]


def is_allo_category_url(url: str) -> bool:
    parsed = urlparse(url)
    if parsed.netloc and parsed.netloc != "allo.ua":
        return False
    if not parsed.path.startswith("/ru/"):
        return False
    if "/proizvoditel-" in parsed.path:
        return False
    return True


def only_direct_children(parent: Tag, selector: str) -> Iterable[Tag]:
    return parent.find_all(selector, recursive=False)


class AlloMapper:
    def __init__(self) -> None:
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})
        self.visited_pages: dict[str, dict] = {}
        self.navigation_rows: list[NavLink] = []
        self.filter_rows: list[dict] = []
        self.required_rows: list[dict] = []
        self.required_by_path: dict[str, set[tuple[str, str]]] = {}

    def fetch_html(self, url: str) -> str:
        response = self.session.get(url, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        response.encoding = response.encoding or "utf-8"
        time.sleep(REQUEST_DELAY)
        return response.text

    def load_required_filters(self) -> None:
        response = self.session.get(SHEET_CSV_URL, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        response.encoding = "utf-8"
        rows = csv.DictReader(response.text.splitlines())
        for row in rows:
            url = normalize_space(row.get("url", ""))
            if not url.startswith("http"):
                continue
            path = normalize_path(url)
            filter_code = normalize_space(row.get("Код фильтра", ""))
            filter_name = normalize_space(row.get("Название фильтра", ""))
            partner_value = normalize_space(row.get("Фильтр для партнера", ""))
            required = "обяз" in partner_value.lower()
            prepared = {
                "category_name": normalize_space(row.get("Название категории", "")),
                "url": url,
                "path": path,
                "filter_code": filter_code,
                "filter_name": filter_name,
                "partner_flag": partner_value,
                "is_required": required,
            }
            self.required_rows.append(prepared)
            if required:
                self.required_by_path.setdefault(path, set()).add((filter_code, filter_name.lower()))

    def parse_portal_links(self, soup: BeautifulSoup, source_url: str, source_title: str) -> list[NavLink]:
        rows: list[NavLink] = []

        for block in soup.select(".portal-category"):
            header = block.select_one(".portal-category__title")
            section_title = normalize_space(header.get_text(" ", strip=True)) if header else "Категории"
            if "производ" in section_title.lower():
                continue
            for link in block.select(".portal-category__item-link"):
                child_url = normalize_url(link.get("href", ""))
                if not is_allo_category_url(child_url):
                    continue
                rows.append(
                    NavLink(
                        source_url=source_url,
                        source_title=source_title,
                        section_title=section_title,
                        child_label=normalize_space(link.get_text(" ", strip=True)),
                        child_url=child_url,
                        child_kind="category",
                        child_page_type="unknown",
                    )
                )

        for group in soup.select(".portal-group__item"):
            title_link = group.select_one(".portal-group__title-link")
            if title_link:
                child_url = normalize_url(title_link.get("href", ""))
                if is_allo_category_url(child_url):
                    rows.append(
                        NavLink(
                            source_url=source_url,
                            source_title=source_title,
                            section_title="Группы",
                            child_label=normalize_space(title_link.get_text(" ", strip=True)),
                            child_url=child_url,
                            child_kind="group",
                            child_page_type="unknown",
                        )
                    )

            for card_item in group.select(".portal-card__item"):
                card_title = card_item.select_one(".portal-card__title")
                section_title = normalize_space(card_title.get_text(" ", strip=True)) if card_title else "Категории"
                for link in card_item.select(".portal-card__link"):
                    child_url = normalize_url(link.get("href", ""))
                    if not is_allo_category_url(child_url):
                        continue
                    label = normalize_space(link.get_text(" ", strip=True))
                    if "все товары" in label.lower():
                        continue
                    kind = "all_items" if "все товары" in label.lower() else "shortcut"
                    rows.append(
                        NavLink(
                            source_url=source_url,
                            source_title=source_title,
                            section_title=section_title,
                            child_label=label,
                            child_url=child_url,
                            child_kind=kind,
                            child_page_type="unknown",
                        )
                    )
        deduped: OrderedDict[tuple[str, str, str, str], NavLink] = OrderedDict()
        for row in rows:
            key = (row.source_url, row.section_title, row.child_label, row.child_url)
            deduped[key] = row
        return list(deduped.values())

    def parse_listing_filters(self, soup: BeautifulSoup, page_title: str, page_url: str) -> list[dict]:
        navigation = soup.select_one(".v-catalog__navigation")
        if not navigation:
            return []

        rows: list[dict] = []
        path = normalize_path(page_url)
        required_for_page = self.required_by_path.get(path, set())

        for index, accordion in enumerate(navigation.select(".accordion"), start=1):
            header = accordion.select_one(".header-title")
            if not header:
                continue

            filter_code = normalize_space(header.get("data-id", ""))
            filter_name = normalize_space(header.get_text(" ", strip=True))
            if not filter_name:
                continue

            is_required = (
                (filter_code, filter_name.lower()) in required_for_page
                or any(name == filter_name.lower() for _, name in required_for_page)
            )

            body = accordion.select_one(".accordion__body")
            if body is None:
                continue

            values_found = False

            for value_link in body.select(".f-check, .f-tile, .f-radio"):
                raw_text = value_link.get_text(" ", strip=True)
                raw_text = re.sub(r"\(\d+\)", "", raw_text).strip()
                value_text = normalize_space(raw_text)
                if not value_text:
                    continue
                amount = ""
                amount_tag = value_link.select_one(".f-check__amount")
                if amount_tag:
                    amount = normalize_space(amount_tag.get_text(" ", strip=True)).strip("()")
                value_url = normalize_url(value_link.get("href", page_url)) if value_link.get("href") else ""
                rows.append(
                    {
                        "category_title": page_title,
                        "category_url": page_url,
                        "page_path": path,
                        "filter_order": index,
                        "filter_code": filter_code,
                        "filter_name": filter_name,
                        "value": value_text,
                        "value_url": value_url,
                        "value_count": amount,
                        "value_type": "choice",
                        "is_required": is_required,
                    }
                )
                values_found = True

            if body.select_one(".f-range"):
                rows.append(
                    {
                        "category_title": page_title,
                        "category_url": page_url,
                        "page_path": path,
                        "filter_order": index,
                        "filter_code": filter_code,
                        "filter_name": filter_name,
                        "value": "Диапазон",
                        "value_url": "",
                        "value_count": "",
                        "value_type": "range",
                        "is_required": is_required,
                    }
                )
                values_found = True

            if not values_found:
                rows.append(
                    {
                        "category_title": page_title,
                        "category_url": page_url,
                        "page_path": path,
                        "filter_order": index,
                        "filter_code": filter_code,
                        "filter_name": filter_name,
                        "value": "",
                        "value_url": "",
                        "value_count": "",
                        "value_type": "unknown",
                        "is_required": is_required,
                    }
                )

        return rows

    def classify_page(self, soup: BeautifulSoup) -> str:
        if soup.select_one(".v-catalog__navigation .header-title"):
            return "listing"
        if soup.select_one(".portal-group__item") or soup.select_one(".portal-category__item-link"):
            return "portal"
        return "other"

    def page_title(self, soup: BeautifulSoup) -> str:
        title = soup.select_one("h1")
        if title:
            return normalize_space(title.get_text(" ", strip=True))
        title_tag = soup.select_one("title")
        if title_tag:
            return normalize_space(title_tag.get_text(" ", strip=True))
        return ""

    def crawl(self) -> None:
        queue: deque[str] = deque([normalize_url(ROOT_URL)])

        while queue:
            url = queue.popleft()
            if url in self.visited_pages:
                continue

            print(f"Fetching {url}", file=sys.stderr)
            html = self.fetch_html(url)
            soup = BeautifulSoup(html, "lxml")
            title = self.page_title(soup)
            page_type = self.classify_page(soup)
            self.visited_pages[url] = {"title": title, "page_type": page_type}

            if page_type == "portal":
                links = self.parse_portal_links(soup, url, title)
                for link in links:
                    self.navigation_rows.append(link)
                    if self.should_recurse(link) and link.child_url not in self.visited_pages:
                        queue.append(link.child_url)
            if page_type == "listing":
                self.filter_rows.extend(self.parse_listing_filters(soup, title, url))

        page_types = {url: data["page_type"] for url, data in self.visited_pages.items()}
        for row in self.navigation_rows:
            row.child_page_type = page_types.get(row.child_url, "unknown")

    def should_recurse(self, link: NavLink) -> bool:
        if link.child_kind in {"group", "category", "all_items"}:
            return True

        parts = path_parts(link.child_url)
        if len(parts) == 1:
            return True
        if len(parts) == 2 and parts[0] == "products":
            return True
        return False

    def relevant_required_rows(self) -> list[dict]:
        seen_paths = {normalize_path(row["category_url"]) for row in self.filter_rows}
        return [row for row in self.required_rows if row["path"] in seen_paths]

    def autosize(self, worksheet) -> None:
        for column_cells in worksheet.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column].width = min(max_length + 2, 60)

    def write_workbook(self) -> None:
        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()

        ws_summary = workbook.active
        ws_summary.title = "Сводка"
        summary_rows = [
            ("Корневая категория", ROOT_URL),
            ("Уникальных страниц", len(self.visited_pages)),
            ("Навигационных связей", len(self.navigation_rows)),
            ("Категорий с фильтрами", len({row["category_url"] for row in self.filter_rows})),
            ("Строк фильтров", len(self.filter_rows)),
            ("Релевантных обязательных фильтров", len(self.relevant_required_rows())),
            ("Источник обязательных фильтров", SHEET_CSV_URL),
        ]
        for row in summary_rows:
            ws_summary.append(row)
        self.autosize(ws_summary)

        ws_nav = workbook.create_sheet("Категории")
        nav_headers = [
            "Источник URL",
            "Источник название",
            "Секция",
            "Дочерний элемент",
            "Дочерний URL",
            "Тип ссылки",
            "Тип страницы",
        ]
        ws_nav.append(nav_headers)
        for cell in ws_nav[1]:
            cell.fill = HEADER_FILL
        for row in self.navigation_rows:
            ws_nav.append(
                [
                    row.source_url,
                    row.source_title,
                    row.section_title,
                    row.child_label,
                    row.child_url,
                    row.child_kind,
                    row.child_page_type,
                ]
            )
        self.autosize(ws_nav)

        ws_filters = workbook.create_sheet("Фильтры")
        filter_headers = [
            "Категория",
            "URL категории",
            "Порядок",
            "Код фильтра",
            "Название фильтра",
            "Значение",
            "URL значения",
            "Количество",
            "Тип значения",
            "Обязательный",
        ]
        ws_filters.append(filter_headers)
        for cell in ws_filters[1]:
            cell.fill = HEADER_FILL
        for row in self.filter_rows:
            values = [
                row["category_title"],
                row["category_url"],
                row["filter_order"],
                row["filter_code"],
                row["filter_name"],
                row["value"],
                row["value_url"],
                row["value_count"],
                row["value_type"],
                "Да" if row["is_required"] else "",
            ]
            ws_filters.append(values)
            if row["is_required"]:
                for cell in ws_filters[ws_filters.max_row]:
                    cell.fill = GREEN_FILL
        self.autosize(ws_filters)

        ws_required = workbook.create_sheet("Обязательные")
        required_headers = [
            "Название категории",
            "URL",
            "Нормализованный путь",
            "Код фильтра",
            "Название фильтра",
            "Флаг",
        ]
        ws_required.append(required_headers)
        for cell in ws_required[1]:
            cell.fill = HEADER_FILL
        for row in self.relevant_required_rows():
            ws_required.append(
                [
                    row["category_name"],
                    row["url"],
                    row["path"],
                    row["filter_code"],
                    row["filter_name"],
                    row["partner_flag"],
                ]
            )
            for cell in ws_required[ws_required.max_row]:
                cell.fill = GREEN_FILL
        self.autosize(ws_required)

        workbook.save(OUTPUT_PATH)


def main() -> int:
    mapper = AlloMapper()
    mapper.load_required_filters()
    mapper.crawl()
    mapper.write_workbook()
    print(f"Saved {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
